"""
Update Checkout Page sheet in SunnyDiamonds_v2.xlsx using openpyxl.
Preserves all formatting, styles, colors, fonts, borders.
Only updates Actual Result and Status columns for matching Test Case IDs.
"""
import json
import os
from openpyxl import load_workbook

EXCEL = os.path.join(os.path.dirname(__file__), 'TestCase', 'SunnyDiamonds_v2.xlsx')
RESULTS = os.path.join(os.path.dirname(__file__), 'checkout-not-tested-results.json')

with open(RESULTS, 'r', encoding='utf-8') as f:
    results = json.load(f)

print(f"Loaded {len(results)} test results.")

wb = load_workbook(EXCEL)
ws = wb['Checkout Page']

# Find header row and column indices
header_row = None
tc_col = None
actual_col = None
status_col = None

for row in ws.iter_rows(min_row=1, max_row=10):
    for cell in row:
        if cell.value == 'Test Case ID':
            header_row = cell.row
            tc_col = cell.column
        if cell.value == 'Actual Result':
            actual_col = cell.column
        if cell.value == 'Status':
            status_col = cell.column
    if header_row:
        break

print(f"Header row: {header_row}, TC col: {tc_col}, Actual col: {actual_col}, Status col: {status_col}")

# Build lookup from results
result_map = {r['tcId']: r for r in results}

updated = 0
for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
    tc_id = row[tc_col - 1].value
    if tc_id and tc_id in result_map:
        current_status = row[status_col - 1].value
        if current_status and str(current_status).strip().lower() == 'not tested':
            r = result_map[tc_id]
            row[actual_col - 1].value = r['actualResult']
            row[status_col - 1].value = r['status']
            updated += 1
            print(f"  {tc_id} -> {r['status']}")

wb.save(EXCEL)
print(f"\nUpdated {updated} rows. Saved to {EXCEL}")
print("All styles, colors, and formatting preserved.")
