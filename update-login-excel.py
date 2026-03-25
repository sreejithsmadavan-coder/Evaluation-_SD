"""
Updates ONLY the 'Login Page' sheet in SunnyDiamonds_v2.xlsx with test results.
Maps test case IDs to rows and writes Actual Result (col G) and Status (col H).
Does NOT modify any other sheet or data.
"""
import json
import openpyxl
import os

RESULTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'login-test-results.json')
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TestCase', 'SunnyDiamonds_v2.xlsx')
SHEET_NAME = 'Login Page'

# Login Page column layout (1-indexed):
# A(1)=Test Case ID, B(2)=Module, C(3)=Description, D(4)=Preconditions,
# E(5)=Test Steps, F(6)=Expected Result, G(7)=Actual Result, H(8)=Status,
# I(9)=Test Type, J(10)=Priority, K(11)=Remarks
ACTUAL_RESULT_COL = 7   # Column G
STATUS_COL = 8           # Column H
TC_ID_COL = 1            # Column A

def main():
    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        results = json.load(f)

    results_map = {r['tcId']: r for r in results}
    print(f"Loaded {len(results)} test results")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    print(f"Opened sheet: {SHEET_NAME}")

    # Find header row (look for "Test Case ID" in column A)
    header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=TC_ID_COL).value
        if cell_val and 'Test Case ID' in str(cell_val):
            header_row = row
            break

    if header_row is None:
        print("ERROR: Could not find header row")
        return

    print(f"Header at row {header_row}, data starts at row {header_row + 1}")

    updated_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=TC_ID_COL).value
        if tc_id and str(tc_id).strip() in results_map:
            tc_id_str = str(tc_id).strip()
            result = results_map[tc_id_str]

            ws.cell(row=row, column=ACTUAL_RESULT_COL).value = result['actualResult']
            ws.cell(row=row, column=STATUS_COL).value = result['status']

            updated_count += 1
            print(f"  Row {row}: {tc_id_str} -> {result['status']}")

    wb.save(EXCEL_FILE)
    print(f"\nUpdated {updated_count} test cases in '{SHEET_NAME}' sheet")
    print(f"Excel file saved: {EXCEL_FILE}")

if __name__ == '__main__':
    main()
