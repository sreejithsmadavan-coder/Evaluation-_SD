"""
Update Checkout Page in SunnyDiamonds_v2.xlsx using openpyxl (preserves all styles).
1. Renames old TC IDs (VIDEO_*, SD_*, S0*) to TC_CHECKOUT_076-112
2. Writes test results (Actual Result + Status) for all renamed rows
3. Inserts new guest checkout rows (TC_CHECKOUT_113-137) before last row
4. Renames last row to TC_CHECKOUT_138 and writes its result
"""
import json, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from copy import copy

EXCEL = os.path.join(os.path.dirname(__file__), 'TestCase', 'SunnyDiamonds_v2.xlsx')
RESULTS = os.path.join(os.path.dirname(__file__), 'checkout-not-tested-results.json')

with open(RESULTS, 'r', encoding='utf-8') as f:
    results = json.load(f)
result_map = {r['tcId']: r for r in results}
print(f"Loaded {len(results)} test results.")

wb = load_workbook(EXCEL)
ws = wb['Checkout Page']

# ── Step 1: Find all "Not Tested" rows and map them ──
not_tested_rows = []
last_row = None
for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
    tc_id = row[0].value
    status = row[7].value
    if tc_id and status and str(status).strip().lower() == 'not tested':
        not_tested_rows.append(row[0].row)
    if tc_id:
        last_row = row[0].row

print(f"Found {len(not_tested_rows)} 'Not Tested' rows.")
print(f"Last TC row: {last_row}")

# Not Tested rows map to TC_CHECKOUT_076 through TC_CHECKOUT_112
# (37 Not Tested rows -> 076 to 112)
# Last row (117, TC_CHECKOUT_077) maps to TC_CHECKOUT_138

# ── Step 2: Update the 37 existing Not Tested rows ──
updated = 0
for i, row_idx in enumerate(not_tested_rows):
    new_id = f'TC_CHECKOUT_{76 + i:03d}'
    if new_id in result_map:
        r = result_map[new_id]
        ws.cell(row=row_idx, column=1).value = new_id      # TC ID
        ws.cell(row=row_idx, column=7).value = r['actualResult']  # Actual Result
        ws.cell(row=row_idx, column=8).value = r['status']        # Status
        updated += 1
        print(f"  Row {row_idx}: {new_id} -> {r['status']}")
    else:
        ws.cell(row=row_idx, column=1).value = new_id
        print(f"  Row {row_idx}: {new_id} (renamed, no result)")

# ── Step 3: Insert guest checkout rows (TC_113-137) before last row ──
guest_ids = [f'TC_CHECKOUT_{n:03d}' for n in range(113, 138)]
guest_results = [result_map[tid] for tid in guest_ids if tid in result_map]

print(f"\nInserting {len(guest_results)} guest checkout rows before row {last_row}...")

if guest_results:
    ws.insert_rows(last_row, len(guest_results))

    # Copy style from a nearby existing row (reference: first Not Tested row)
    ref_row = not_tested_rows[0] if not_tested_rows else 80
    for i, tc in enumerate(guest_results):
        new_row = last_row + i
        # Copy cell styles from reference row
        for col in range(1, 12):
            src = ws.cell(row=ref_row, column=col)
            dst = ws.cell(row=new_row, column=col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)

        ws.cell(row=new_row, column=1).value = tc['tcId']
        ws.cell(row=new_row, column=2).value = 'Checkout Page'
        ws.cell(row=new_row, column=7).value = tc['actualResult']
        ws.cell(row=new_row, column=8).value = tc['status']
        updated += 1
        print(f"  Row {new_row}: {tc['tcId']} -> {tc['status']}")

# ── Step 4: Update last row (now shifted down) -> TC_CHECKOUT_138 ──
shifted_last = last_row + len(guest_results)
if 'TC_CHECKOUT_138' in result_map:
    r = result_map['TC_CHECKOUT_138']
    ws.cell(row=shifted_last, column=1).value = 'TC_CHECKOUT_138'
    ws.cell(row=shifted_last, column=7).value = r['actualResult']
    ws.cell(row=shifted_last, column=8).value = r['status']
    updated += 1
    print(f"  Row {shifted_last}: TC_CHECKOUT_138 -> {r['status']}")

# Also update TC_CHECKOUT_075 result if we have it
if 'TC_CHECKOUT_075' in result_map:
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        if row[0].value == 'TC_CHECKOUT_075':
            r = result_map['TC_CHECKOUT_075']
            row[6].value = r['actualResult']
            row[7].value = r['status']
            updated += 1
            print(f"  Row {row[0].row}: TC_CHECKOUT_075 -> {r['status']}")
            break

wb.save(EXCEL)
print(f"\n=== Done === Updated {updated} rows in {EXCEL}")
print("All styles, colors, and formatting preserved.")
