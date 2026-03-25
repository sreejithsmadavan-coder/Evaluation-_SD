"""
Updates the SunnyDiamonds.xlsx Excel file with test execution results.
Maps test case IDs to rows and writes Actual Result (column H) and Status (column I).
"""
import json
import openpyxl
import os

RESULTS_FILE = os.path.join(os.path.dirname(__file__), 'test-results.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'TestCase', 'SunnyDiamonds.xlsx')
SHEET_NAME = 'Registration Test Cases'

# Column mapping (1-indexed):
# A=Module Name, B=Test Case ID, C=Description, D=Preconditions,
# E=Test Steps, F=Test Data, G=Expected Result,
# H=Actual Result, I=Status, J=Test Type, K=Priority, L=Remarks
ACTUAL_RESULT_COL = 8   # Column H
STATUS_COL = 9           # Column I

def main():
    # Load test results
    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        results = json.load(f)

    # Create a lookup by test case ID
    results_map = {r['tcId']: r for r in results}
    print(f"Loaded {len(results)} test results")

    # Load Excel workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    print(f"Opened sheet: {SHEET_NAME}")

    # Find header row (look for "Test Case ID" in column B)
    header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and 'Test Case ID' in str(cell_val):
            header_row = row
            break

    if header_row is None:
        print("ERROR: Could not find header row with 'Test Case ID'")
        return

    print(f"Header row found at row {header_row}")

    # Iterate through data rows and update results
    updated_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        tc_id = ws.cell(row=row, column=2).value
        if tc_id and str(tc_id).strip() in results_map:
            tc_id_str = str(tc_id).strip()
            result = results_map[tc_id_str]

            # Write Actual Result (column H)
            ws.cell(row=row, column=ACTUAL_RESULT_COL).value = result['actualResult']

            # Write Status (column I)
            ws.cell(row=row, column=STATUS_COL).value = result['status']

            updated_count += 1
            print(f"  Row {row}: {tc_id_str} -> {result['status']}")

    # Save the workbook
    wb.save(EXCEL_FILE)
    print(f"\nUpdated {updated_count} test cases in {EXCEL_FILE}")
    print("Excel file saved successfully!")

if __name__ == '__main__':
    main()
