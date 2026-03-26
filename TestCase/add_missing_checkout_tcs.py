import shutil, os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_work.xlsx"

# ── colours ────────────────────────────────────────────────────────────────────
GREEN  = "E2EFDA"   # Positive
RED    = "FCE4D6"   # Negative
YELLOW = "FFF2CC"   # Edge Case
HEADER_BG = "2F4F4F"

thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col, value, bg=None, bold=False, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", size=9, bold=bold)
    c.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal="left")
    c.border = border
    return c

# ── missing TCs mapped to QA Checklist ─────────────────────────────────────────
# Format: (TC_ID, Module, Description, Preconditions, Steps, Expected, ActualResult, Status, TestType, Priority, Remarks)
# Preconditions = "" means use Common Preconditions row

NAV = "Navigate back to https://qa-sunnydiamonds.webc.in/checkout"

MISSING_TCS = [
    # ── Text Fields — Leading Space (Checklist §5, Contact form §2) ─────────────
    (
        "TC_CHECKOUT_053",
        "Checkout - Shipping Address",
        "First Name — Enter value with leading space — should be blocked or trimmed",
        "",
        "1. " + NAV + "\n2. Click on First Name field\n3. Type ' John' (with leading space)\n4. Click on another field\n5. Observe field value and error message",
        "Leading space is blocked or auto-trimmed; no leading space in submitted value; field shows error or trims to 'John'",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist §5: Leading and trailing whitespace should not be allowed in text fields. Contact form §2: Leading spaces should not be allowed."
    ),
    (
        "TC_CHECKOUT_054",
        "Checkout - Shipping Address",
        "Last Name — Enter value with leading space — should be blocked or trimmed",
        "",
        "1. " + NAV + "\n2. Click on Last Name field\n3. Type ' Smith' (with leading space)\n4. Click on another field\n5. Observe field value and error message",
        "Leading space is blocked or auto-trimmed; no leading space in submitted value",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Contact form §2: Leading spaces should not be allowed in First Name and Last Name."
    ),
    (
        "TC_CHECKOUT_055",
        "Checkout - Shipping Address",
        "Address field — Enter value with leading space — should be blocked or trimmed",
        "",
        "1. " + NAV + "\n2. Click on Address field\n3. Type '  123 Main Street' (with leading spaces)\n4. Tab out\n5. Observe value and validation",
        "Leading spaces are blocked or trimmed; address value starts with alphanumeric character",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §5: Leading and trailing whitespaces in text fields should be handled appropriately."
    ),
    # ── Text Fields — Special Chars in Name (Contact form §2, Checklist §5) ─────
    (
        "TC_CHECKOUT_056",
        "Checkout - Shipping Address",
        "First Name — Enter special characters (e.g., 'John@#Doe') — should be rejected",
        "",
        "1. " + NAV + "\n2. Enter 'John@#Doe' in First Name field\n3. Tab out\n4. Observe error message",
        "Field rejects special characters; error message displayed: 'First Name should contain alphabets only' or similar",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Contact form §2: Only alphabets should be allowed in First Name and Last Name fields."
    ),
    (
        "TC_CHECKOUT_057",
        "Checkout - Shipping Address",
        "Last Name — Enter special characters (e.g., 'Smith!@#') — should be rejected",
        "",
        "1. " + NAV + "\n2. Enter 'Smith!@#' in Last Name field\n3. Tab out\n4. Observe error message",
        "Field rejects special characters; error message displayed indicating alphabets only",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Contact form §2: Only alphabets should be allowed in Last Name. Alpha-only validation."
    ),
    # ── Last Name BVA (Contact form §2: max 56) ───────────────────────────────
    (
        "TC_CHECKOUT_058",
        "Checkout - Shipping Address",
        "Last Name BVA — Enter exactly 56 characters (maximum allowed)",
        "",
        "1. " + NAV + "\n2. Enter a 56-character alphabetic string in Last Name field (e.g., 'A'*56)\n3. Tab out\n4. Observe acceptance",
        "Last Name field accepts exactly 56 characters; no error displayed; value saved correctly",
        "", "Not Tested", "Positive", "Medium",
        "BVA max boundary for Last Name. Mapped from checklist Contact form §2: Max 56 characters for Last Name."
    ),
    (
        "TC_CHECKOUT_059",
        "Checkout - Shipping Address",
        "Last Name BVA — Enter 57 characters (one above maximum of 56)",
        "",
        "1. " + NAV + "\n2. Enter a 57-character alphabetic string in Last Name field\n3. Tab out\n4. Observe validation",
        "Field rejects input beyond 56 characters; either input is truncated at 56 or error message is displayed",
        "", "Not Tested", "Negative", "High",
        "BVA max+1 for Last Name. Mapped from checklist Contact form §2: Maximum characters for Last Name should be 56."
    ),
    # ── Consecutive Whitespace (Checklist §5) ────────────────────────────────
    (
        "TC_CHECKOUT_060",
        "Checkout - Shipping Address",
        "First Name — Enter consecutive whitespace (e.g., 'John  Doe' with double space) — should be collapsed or blocked",
        "",
        "1. " + NAV + "\n2. Enter 'John  Doe' (two consecutive spaces) in First Name field\n3. Tab out\n4. Observe behavior",
        "Consecutive whitespace is either collapsed to single space or an error is shown; field does not accept multiple consecutive spaces",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §5: Check if consecutive whitespaces are handled appropriately in text fields."
    ),
    # ── Emoji/Unicode in text fields (Checklist §5) ──────────────────────────
    (
        "TC_CHECKOUT_061",
        "Checkout - Shipping Address",
        "First Name — Enter emoji characters (e.g., 'John😊') — system should handle or reject",
        "",
        "1. " + NAV + "\n2. Enter 'John\U0001f60a' in First Name field\n3. Tab out\n4. Observe behavior and error",
        "System rejects emoji input with appropriate error message, OR displays validation error indicating only alphabets are accepted",
        "", "Not Tested", "Edge Case", "Medium",
        "Mapped from checklist §5: Test the text field for inclusion of emoji and Unicode characters — ensure system handles correctly."
    ),
    # ── City Field Validation (Checklist §12) ─────────────────────────────────
    (
        "TC_CHECKOUT_062",
        "Checkout - Shipping Address",
        "City field — Enter numeric digits only (e.g., '12345') — should be rejected",
        "",
        "1. " + NAV + "\n2. Enter '12345' in City field\n3. Tab out\n4. Observe error message",
        "City field rejects numeric-only input; error message displayed indicating valid city name required",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §12: Test for proper validation of the city field; confirm valid city names are accepted."
    ),
    (
        "TC_CHECKOUT_063",
        "Checkout - Shipping Address",
        "City field — Enter leading space (e.g., ' Mumbai') — should be blocked or trimmed",
        "",
        "1. " + NAV + "\n2. Enter ' Mumbai' (with leading space) in City field\n3. Tab out\n4. Observe behavior",
        "Leading space is blocked or trimmed; city value starts with a letter",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §12 and §5: Address field validation — leading whitespace should be handled."
    ),
    # ── Phone — Dashes & Whitespace (Checklist §11) ───────────────────────────
    (
        "TC_CHECKOUT_064",
        "Checkout - Shipping Address",
        "Phone Number — Enter with dashes (e.g., '98765-43210') — verify acceptance or rejection",
        "",
        "1. " + NAV + "\n2. Enter '98765-43210' in Phone Number field\n3. Tab out\n4. Observe validation behavior",
        "System either accepts the dashed format and normalizes it to '9876543210', or displays error asking for 10-digit format without dashes",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §11: Handle different phone formats (e.g., with or without dashes). Verify system behavior."
    ),
    (
        "TC_CHECKOUT_065",
        "Checkout - Shipping Address",
        "Phone Number — Enter with leading/trailing whitespace (e.g., ' 9876543210 ') — should be blocked",
        "",
        "1. " + NAV + "\n2. Enter ' 9876543210 ' (with spaces before and after) in Phone Number field\n3. Tab out\n4. Observe behavior",
        "Leading and trailing whitespace is blocked or trimmed; phone field does not accept spaces around the number",
        "", "Not Tested", "Negative", "Medium",
        "Mapped from checklist §11: Validate how the system handles leading and trailing whitespaces in phone numbers."
    ),
    # ── Pin Code BVA max+1 (Checklist §4 Numeric, §12 Postal Code) ──────────
    (
        "TC_CHECKOUT_066",
        "Checkout - Shipping Address",
        "Pin Code BVA — Enter 7 digits (one above maximum 6-digit limit)",
        "",
        "1. " + NAV + "\n2. Enter '1234567' (7 digits) in Pin Code field\n3. Tab out\n4. Observe validation",
        "Field rejects 7-digit input; either truncates at 6 digits or error message displayed for invalid PIN code",
        "", "Not Tested", "Negative", "High",
        "BVA: max+1 for PIN code. Mapped from checklist §4 (numeric max) and §12 (postal code validation). Complements TC_CHECKOUT_040."
    ),
    # ── State Dropdown — Options Verification (Checklist §6) ─────────────────
    (
        "TC_CHECKOUT_067",
        "Checkout - Shipping Address",
        "State dropdown — Verify options are sorted alphabetically and all Indian states/UTs are present",
        "",
        "1. " + NAV + "\n2. Click on State dropdown\n3. Observe all options listed\n4. Verify alphabetical sort order\n5. Verify count matches 28 states + 8 UTs = 36 options",
        "State dropdown contains all Indian states and UTs (36 total); options are sorted alphabetically from 'Andaman and Nicobar Islands' to 'West Bengal'",
        "", "Not Tested", "Positive", "Medium",
        "Mapped from checklist §6: Verify options are sorted alphabetically; check correct number of options are available in dropdown."
    ),
    # ── Tab-out Error Messages (Checklist §16) ────────────────────────────────
    (
        "TC_CHECKOUT_068",
        "Checkout - Shipping Address",
        "Email field — Verify error message appears on tab-out when invalid email entered",
        "",
        "1. " + NAV + "\n2. Click Email Address field\n3. Enter 'invalidemail' (no @)\n4. Press Tab to move to next field\n5. Observe immediate error display",
        "Error message 'Invalid email address' or similar appears immediately on tab-out, displayed near the Email field, without needing to click PAY NOW",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist §16: Ensure validation messages are displayed while tabout or data entry. Verify inline field-level error."
    ),
    (
        "TC_CHECKOUT_069",
        "Checkout - Shipping Address",
        "Phone Number field — Verify error message appears on tab-out when invalid phone entered",
        "",
        "1. " + NAV + "\n2. Click Phone Number field\n3. Enter '123' (too short)\n4. Press Tab to move to next field\n5. Observe immediate error display",
        "Error message appears immediately on tab-out near the Phone Number field indicating invalid format or minimum length required",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist §16: Validation messages displayed on tab-out. Complements TC_CHECKOUT_023 (First Name tab-out)."
    ),
    (
        "TC_CHECKOUT_070",
        "Checkout - Shipping Address",
        "Pin Code field — Verify error message appears on tab-out when non-numeric or short value entered",
        "",
        "1. " + NAV + "\n2. Click Pin Code field\n3. Enter 'ABC' (alphabetic)\n4. Press Tab to move to next field\n5. Observe immediate error display",
        "Error message appears immediately on tab-out near Pin Code field: 'Please enter a valid 6-digit PIN code' or similar",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist §16: Ensure validation messages displayed on tab-out. Complements TC_CHECKOUT_029."
    ),
    # ── CAPTCHA (Checklist §13) ───────────────────────────────────────────────
    (
        "TC_CHECKOUT_071",
        "Checkout - Security",
        "Verify reCAPTCHA is present on checkout page and blocks submission if not completed",
        "",
        "1. " + NAV + "\n2. Fill all mandatory shipping address fields with valid data\n3. Select a payment method\n4. Do NOT complete reCAPTCHA verification\n5. Click PAY NOW button\n6. Observe system response",
        "reCAPTCHA widget is visible on checkout page; clicking PAY NOW without completing CAPTCHA displays error message preventing form submission",
        "", "Not Tested", "Negative", "Critical",
        "Mapped from checklist §13: Implement and validate CAPTCHA or other bot protection mechanisms. reCAPTCHA (g-recaptcha-response) is present on checkout form."
    ),
    # ── Form Validation Sheet — Success / Error / Button Disable ─────────────
    (
        "TC_CHECKOUT_072",
        "Checkout - Order Placement",
        "Verify success message displayed after successful order placement (COD)",
        "",
        "1. " + NAV + "\n2. Fill all mandatory fields with valid data\n3. Select 'Cash on Delivery'\n4. Complete reCAPTCHA\n5. Click PAY NOW\n6. Observe success message and page redirect",
        "Success message displayed (e.g., 'Order placed successfully!' or order confirmation screen shown); user redirected to order confirmation page with order ID",
        "", "Not Tested", "Positive", "Critical",
        "Mapped from Form Validations sheet: 'Form submitted successfully!' message after submitting form. Order success verification."
    ),
    (
        "TC_CHECKOUT_073",
        "Checkout - Order Placement",
        "PAY NOW button — Verify button is disabled or shows loader after first click to prevent duplicate submission",
        "",
        "1. " + NAV + "\n2. Fill all fields with valid data\n3. Select COD payment\n4. Click PAY NOW\n5. Immediately attempt to click PAY NOW again before page transitions\n6. Observe button state",
        "PAY NOW button becomes disabled or shows loading spinner after first click; second click has no effect; only one order is created",
        "", "Not Tested", "Negative", "Critical",
        "Mapped from Form Validations sheet: 'Remove multiple click on submit button by disabling or adding loader'. Prevents duplicate orders."
    ),
    (
        "TC_CHECKOUT_074",
        "Checkout - Payment",
        "Verify error message displayed when online payment fails via Razorpay",
        "",
        "1. " + NAV + "\n2. Fill all mandatory fields with valid data\n3. Select 'Pay Online' (Razorpay)\n4. Click PAY NOW\n5. In Razorpay popup, simulate payment failure (use test card for failure)\n6. Observe error handling on checkout page",
        "Payment failure is handled gracefully; error message displayed: 'An error occurred while processing payment. Please try again.' or similar; user remains on checkout page with form data preserved",
        "", "Not Tested", "Negative", "Critical",
        "Mapped from Form Validations sheet: 'An error occurred while submitting the form. Please try again.' Error handling for payment gateway failure."
    ),
    (
        "TC_CHECKOUT_075",
        "Checkout - Network",
        "Simulate network error during checkout — verify appropriate error message displayed",
        "",
        "1. " + NAV + "\n2. Fill all fields with valid data\n3. Throttle network to offline using DevTools (F12 > Network > Offline)\n4. Click PAY NOW\n5. Observe error handling",
        "System displays network error message: 'Network error. Please check your internet connection and try again.' User is not charged; form data preserved after reconnection",
        "", "Not Tested", "Edge Case", "High",
        "Mapped from Form Validations sheet: 'Network error. Please check your internet connection and try again.' Network resilience testing."
    ),
]

# ── Load workbook and sheet ────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]

# ── Find TC_CHECKOUT_052 row (Valid Login — must stay LAST) ───────────────────
last_tc_row = None
valid_login_row_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value == "TC_CHECKOUT_052":
        last_tc_row = row[0].row
        valid_login_row_data = [c.value for c in row]
        break

if last_tc_row is None:
    # No TC_052, just find last data row
    last_tc_row = ws.max_row + 1
    valid_login_row_data = None

print(f"TC_CHECKOUT_052 (Valid Login) found at row: {last_tc_row}")
print(f"Current max row: {ws.max_row}")

# ── Helper: get style from header row ─────────────────────────────────────────
# Find a sample TC row to copy fill colors
def get_row_fill(tc_type):
    if tc_type == "Positive":
        return GREEN
    elif tc_type == "Negative":
        return RED
    else:
        return YELLOW  # Edge Case

# ── Append missing TCs starting from row after current last row (overwriting TC_052 slot) ──
# We will overwrite TC_052 row with first missing TC, then append rest, then add Valid Login last

start_row = last_tc_row  # Start writing at TC_052's row
new_tc_number = 53

for i, tc in enumerate(MISSING_TCS):
    write_row = start_row + i
    tc_id, module, desc, precond, steps, expected, actual, status, tc_type, priority, remarks = tc
    bg = get_row_fill(tc_type)

    col_vals = [
        tc_id, module, desc, precond, steps, expected, actual, status, tc_type, priority, remarks
    ]
    for col_idx, val in enumerate(col_vals, start=1):
        cell_style(ws, write_row, col_idx, val, bg=bg)

print(f"Added {len(MISSING_TCS)} missing TCs (rows {start_row} to {start_row + len(MISSING_TCS) - 1})")

# ── Re-append Valid Login TC as the LAST TC ────────────────────────────────────
final_row = start_row + len(MISSING_TCS)
valid_login_tc = [
    "TC_CHECKOUT_076",
    "Checkout - Authentication",
    "Valid Login — Verify authenticated user (sreejith.s+4@webandcrafts.com) can access and complete Checkout page",
    "",
    "1. Open https://qa-sunnydiamonds.webc.in/login\n2. Enter email: sreejith.s+4@webandcrafts.com\n3. Enter password: Password\n4. Click Login\n5. Navigate to /jewellery PLP\n6. Add 2 products to cart\n7. Go to Cart and click CHECKOUT SECURELY\n8. Verify Checkout page loads with all sections: Shipping Address, Order Summary, Coupon, Payment\n9. Fill all mandatory fields with valid data\n10. Click PAY NOW",
    "Authenticated user successfully reaches and completes the checkout page; all sections load correctly; order placed successfully with valid credentials",
    "", "Not Tested", "Positive", "Critical",
    "Valid login TC — ALWAYS LAST per QA standard (sreejith.s+4@webandcrafts.com / Password). Confirms full auth -> checkout -> order flow."
]

bg = get_row_fill("Positive")
for col_idx, val in enumerate(valid_login_tc, start=1):
    cell_style(ws, final_row, col_idx, val, bg=bg)

print(f"Valid Login TC written at row: {final_row} as TC_CHECKOUT_076")

# ── Update metadata row (Row 2) with new total TC count ──────────────────────
total_tcs = final_row - 5  # rows 6 to final_row inclusive
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            old_val = cell.value
            # Replace total TC count
            import re
            new_val = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total_tcs}', old_val)
            cell.value = new_val
            print(f"Updated metadata: {new_val[:80]}")
            break

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(TMP)
print("Saved to TMP file.")

try:
    os.replace(TMP, FILE)
    print(f"SUCCESS: SunnyDiamonds_v2.xlsx updated with {total_tcs} total Checkout TCs")
except PermissionError:
    print(f"PermissionError: File is open in Excel. Saved as: {TMP}")
    print("Please close SunnyDiamonds_v2.xlsx in Excel, then this file will be copied over.")
