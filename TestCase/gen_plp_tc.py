from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import shutil, os, tempfile

FILE  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET = "PLP Page"

# Work on a temp copy so we can save even if the original is open in Excel
_tmp = FILE.replace(".xlsx", "_plp_tmp.xlsx")
shutil.copy2(FILE, _tmp)
wb = load_workbook(_tmp)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GOLD_LIGHT  = "FFF3CD"
POS_GREEN   = "E2EFDA"
NEG_RED     = "FCE4D6"
EDGE_YELLOW = "FFF2CC"
WHITE       = "FFFFFF"
GREY_MID    = "D9D9D9"
GREY_LIGHT  = "F2F2F2"

def S(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

thin  = Border(left=S(), right=S(), top=S(), bottom=S())
thick = Border(left=S(), right=S(), top=S(), bottom=S("medium", "1F3864"))

# ── Row 1 Banner ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
c = ws.cell(row=1, column=1,
            value="SUNNY DIAMONDS — JEWELLERY PLP (PRODUCT LISTING PAGE) TEST CASES")
c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=15)
c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Row 2 Metadata ────────────────────────────────────────────────────────────
meta = [
    ("A2","Project URL:","B2","https://qa-sunnydiamonds.webc.in/jewellery"),
    ("D2","Module:","E2","PLP – Product Listing Page"),
    ("G2","Created By:","H2","Sreejith S Madavan"),
    ("J2","Created Date:","K2","25-Mar-2026"),
]
for lc,lv,vc,vv in meta:
    l=ws[lc]; l.value=lv
    l.font=Font(name="Arial",bold=True,color=DARK_BLUE,size=10)
    l.fill=PatternFill("solid",fgColor=GREY_MID)
    l.alignment=Alignment(horizontal="left",vertical="center"); l.border=thin
    v=ws[vc]; v.value=vv
    v.font=Font(name="Arial",color=DARK_BLUE,size=10)
    v.fill=PatternFill("solid",fgColor=GREY_LIGHT)
    v.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    v.border=thin
ws.merge_cells("B2:C2"); ws.merge_cells("E2:F2")
ws.merge_cells("H2:I2")
ws.row_dimensions[2].height = 22

for col in range(1, 12):
    ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=DARK_BLUE)
ws.row_dimensions[3].height = 6

# ── Row 4 Headers ─────────────────────────────────────────────────────────────
HEADERS = [
    "Test Case ID", "Module Name", "Test Case Description",
    "Preconditions", "Test Steps", "Expected Result",
    "Actual Result", "Status", "Test Type", "Priority", "Remarks"
]
for i, h in enumerate(HEADERS, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thick
ws.row_dimensions[4].height = 30

# ── Shared text blocks ────────────────────────────────────────────────────────
PRE = ("1. Application accessible at https://qa-sunnydiamonds.webc.in/jewellery\n"
       "2. User is on the Jewellery PLP\n"
       "3. Application is up and running\n"
       "4. Minimum 1 product exists in the catalogue")

PRE_LOGIN = ("1. Application accessible at https://qa-sunnydiamonds.webc.in/jewellery\n"
             "2. Account with email 'sreejith.s+4@webandcrafts.com' is registered\n"
             "3. User is NOT logged in initially")

# Navigation note embedded in steps
NAV_FIRST  = "NAVIGATION: page.goto('https://qa-sunnydiamonds.webc.in/jewellery')\n"
NAV_BACK   = "NAVIGATION: page.goBack()  [use goto() as fallback if goBack() fails]\n"

RESET_BACK = ("POST-TEST RESET:\n"
              "await page.goBack()  // returns to PLP\n"
              "[Fallback: await page.goto('https://qa-sunnydiamonds.webc.in/jewellery')]")

# ── Test Cases ────────────────────────────────────────────────────────────────
# Tuple: (ID, Module, Description, Preconditions, Steps, Expected,
#          ActualResult, Status, TestType, Priority, Remarks)
TC = [

    # ══════════════════ POSITIVE ════════════════════════════════════════════
    (
        "TC_PLP_001", "PLP Page",
        "Verify Jewellery PLP loads correctly with all UI elements visible",
        PRE,
        NAV_FIRST +
        "1. Navigate to https://qa-sunnydiamonds.webc.in/jewellery\n"
        "2. Wait for page to fully load\n"
        "3. Inspect all visible UI elements",
        "PLP page loads successfully\n"
        "Following elements are visible:\n"
        "- Page heading: 'Jewellery'\n"
        "- Product description text\n"
        "- Total product count label (e.g. '2141 Products')\n"
        "- Filter panel on the left\n"
        "- Sort By dropdown\n"
        "- Product grid with cards (image, name, price, Add to Cart)\n"
        "- Pagination bar (Previous | 1 | 2 | ... | 90 | Next)\n"
        "- Breadcrumb: Home > Jewellery\n"
        "- No console errors",
        "", "Pass/Fail", "Positive", "High",
        "TC_PLP_001: FIRST TC — uses page.goto(). Full UI completeness check."
    ),
    (
        "TC_PLP_002", "PLP Page",
        "Verify total product count is displayed and matches actual catalogue size",
        PRE,
        NAV_BACK +
        "1. Observe the product count label above the product grid\n"
        "2. Note the number displayed (e.g. '2141 Products')\n"
        "3. Verify the count is a positive integer",
        "Product count label is displayed (e.g. '2141 Products')\n"
        "Count is a positive non-zero number\n"
        "Count is consistent before and after page reload",
        "", "Pass/Fail", "Positive", "Medium",
        "Product count accuracy validation"
    ),
    (
        "TC_PLP_003", "PLP Page",
        "Filter by Category: Rings — verify only Ring products are displayed",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, click category 'Rings'\n"
        "2. URL updates to ?category=rings or /jewellery/rings\n"
        "3. Observe the product grid",
        "Only Ring products are displayed\n"
        "Product count decreases from total catalogue\n"
        "URL reflects the applied filter\n"
        "Filter label 'Rings' appears as active/selected",
        "", "Pass/Fail", "Positive", "High",
        "Category filter — Rings"
    ),
    (
        "TC_PLP_004", "PLP Page",
        "Filter by Category: Earrings — verify only Earring products are displayed",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, click category 'Earrings'\n"
        "2. Observe the product grid",
        "Only Earring products are displayed\n"
        "Product count updates correctly\n"
        "URL reflects ?category=earrings or /jewellery/earrings",
        "", "Pass/Fail", "Positive", "High",
        "Category filter — Earrings"
    ),
    (
        "TC_PLP_005", "PLP Page",
        "Filter by Category: Pendants — verify only Pendant products are displayed",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, click category 'Pendants'\n"
        "2. Observe the product grid",
        "Only Pendant products are displayed\n"
        "Product count updates correctly\n"
        "URL reflects the filter parameter",
        "", "Pass/Fail", "Positive", "High",
        "Category filter — Pendants"
    ),
    (
        "TC_PLP_006", "PLP Page",
        "Filter by Metal Color: Yellow Gold — verify filtered products are yellow gold",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Metal Color 'Yellow gold'\n"
        "2. URL updates to ?variants.color=yellow-gold\n"
        "3. Observe the product grid",
        "Products displayed are filtered to Yellow Gold metal color\n"
        "Product count updates\n"
        "URL contains '?variants.color=yellow-gold'",
        "", "Pass/Fail", "Positive", "High",
        "Metal color filter — Yellow Gold"
    ),
    (
        "TC_PLP_007", "PLP Page",
        "Filter by Metal Color: Rose Gold — verify filtered products are rose gold",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Metal Color 'Rose gold'\n"
        "2. Observe the product grid and URL",
        "Products are filtered to Rose Gold\n"
        "URL contains '?variants.color=rose-gold'\n"
        "Product count updates accordingly",
        "", "Pass/Fail", "Positive", "High",
        "Metal color filter — Rose Gold"
    ),
    (
        "TC_PLP_008", "PLP Page",
        "Filter by Metal Color: White Gold — verify filtered products are white gold",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Metal Color 'White gold'\n"
        "2. Observe the product grid and URL",
        "Products are filtered to White Gold\n"
        "URL contains '?variants.color=white-gold'\n"
        "Product count updates accordingly",
        "", "Pass/Fail", "Positive", "High",
        "Metal color filter — White Gold"
    ),
    (
        "TC_PLP_009", "PLP Page",
        "Filter by Metal Purity: 18K — verify only 18K products are shown",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Metal Purity '18K'\n"
        "2. URL updates to ?variants.metal_purity=18k\n"
        "3. Observe the product grid",
        "Only 18K products are displayed\n"
        "URL contains '?variants.metal_purity=18k'\n"
        "Product count reflects 18K items only",
        "", "Pass/Fail", "Positive", "High",
        "Metal purity filter — 18K"
    ),
    (
        "TC_PLP_010", "PLP Page",
        "Filter by Metal Purity: 22K — verify only 22K products are shown",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Metal Purity '22K'\n"
        "2. URL updates to ?variants.metal_purity=22k\n"
        "3. Observe the product grid",
        "Only 22K products are displayed\n"
        "URL contains '?variants.metal_purity=22k'\n"
        "Product count reflects 22K items only",
        "", "Pass/Fail", "Positive", "High",
        "Metal purity filter — 22K"
    ),
    (
        "TC_PLP_011", "PLP Page",
        "Filter by Price Range: Below \u20b910k — verify products are within that range",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Price 'Below \u20b910k'\n"
        "2. URL updates to ?range=0TO10000\n"
        "3. Observe the product grid and prices displayed",
        "Only products priced below \u20b910,000 are displayed\n"
        "URL contains '?range=0TO10000'\n"
        "All visible product prices are \u2264 \u20b99,999\n"
        "Product count updates",
        "", "Pass/Fail", "Positive", "High",
        "Price filter — Below \u20b910k"
    ),
    (
        "TC_PLP_012", "PLP Page",
        "Filter by Price Range: \u20b910k\u2013\u20b930k — verify products are within that price band",
        PRE,
        NAV_BACK +
        "1. In the Filter panel, select Price '\u20b910k - \u20b930k'\n"
        "2. URL updates to ?range=10000TO30000\n"
        "3. Observe product prices",
        "Only products priced between \u20b910,000 and \u20b930,000 are displayed\n"
        "URL contains '?range=10000TO30000'\n"
        "Product count updates",
        "", "Pass/Fail", "Positive", "High",
        "Price filter — \u20b910k\u2013\u20b930k"
    ),
    (
        "TC_PLP_013", "PLP Page",
        "Sort By: Price Low to High — verify first product has the lowest price",
        PRE,
        NAV_BACK +
        "1. Click the Sort By dropdown\n"
        "2. Select 'Price: Low to High'\n"
        "3. Wait for products to re-order\n"
        "4. Compare the price of the first and second product card",
        "Products are sorted in ascending price order\n"
        "First product has the lowest price\n"
        "Each subsequent product has a price >= the previous one",
        "", "Pass/Fail", "Positive", "High",
        "Sort — Price ascending"
    ),
    (
        "TC_PLP_014", "PLP Page",
        "Sort By: Price High to Low — verify first product has the highest price",
        PRE,
        NAV_BACK +
        "1. Click the Sort By dropdown\n"
        "2. Select 'Price: High to Low'\n"
        "3. Wait for products to re-order\n"
        "4. Compare prices of first vs subsequent cards",
        "Products are sorted in descending price order\n"
        "First product has the highest price\n"
        "Each subsequent product has a price <= the previous one",
        "", "Pass/Fail", "Positive", "High",
        "Sort — Price descending"
    ),
    (
        "TC_PLP_015", "PLP Page",
        "Sort By: Name A to Z — verify products are sorted alphabetically",
        PRE,
        NAV_BACK +
        "1. Click the Sort By dropdown\n"
        "2. Select 'Name: A to Z'\n"
        "3. Compare the names of the first few product cards",
        "Products are sorted alphabetically A\u2192Z\n"
        "First product name starts closest to 'A'\n"
        "Names are in ascending alphabetical order",
        "", "Pass/Fail", "Positive", "Medium",
        "Sort — Name A to Z"
    ),
    (
        "TC_PLP_016", "PLP Page",
        "Sort By: Name Z to A — verify products are sorted in reverse alphabetical order",
        PRE,
        NAV_BACK +
        "1. Click the Sort By dropdown\n"
        "2. Select 'Name: Z to A'\n"
        "3. Compare the names of the first few product cards",
        "Products are sorted alphabetically Z\u2192A\n"
        "First product name starts closest to 'Z'\n"
        "Names are in descending alphabetical order",
        "", "Pass/Fail", "Positive", "Medium",
        "Sort — Name Z to A"
    ),
    (
        "TC_PLP_017", "PLP Page",
        "Pagination: Navigate to Page 2 using the Next button",
        PRE,
        NAV_BACK +
        "1. Scroll to bottom of product grid\n"
        "2. Click the 'Next' button in the pagination bar\n"
        "3. Wait for page 2 to load\n"
        "4. Verify current page indicator shows '2'",
        "Page 2 of products loads successfully\n"
        "Page 2 indicator is active/highlighted in pagination\n"
        "Different set of products is displayed compared to page 1\n"
        "URL or page state reflects page 2",
        "", "Pass/Fail", "Positive", "High",
        "Pagination — Next button"
    ),
    (
        "TC_PLP_018", "PLP Page",
        "Pagination: Navigate back to Page 1 using the Previous button",
        PRE,
        NAV_BACK +
        "1. Navigate to page 2 (click Next)\n"
        "2. Click the 'Previous' button\n"
        "3. Wait for page 1 to load",
        "Page 1 of products loads successfully\n"
        "Page 1 indicator is active in pagination\n"
        "Same products as original load are displayed",
        "", "Pass/Fail", "Positive", "High",
        "Pagination — Previous button"
    ),
    (
        "TC_PLP_019", "PLP Page",
        "Pagination: Click on a specific page number (e.g. page 5) directly",
        PRE,
        NAV_BACK +
        "1. In the pagination bar, click page number '2' (or any available number)\n"
        "2. Wait for the page to load\n"
        "3. Verify the correct page is shown",
        "The selected page (e.g. 2) loads correctly\n"
        "Selected page number is highlighted in pagination\n"
        "Products on that page are displayed",
        "", "Pass/Fail", "Positive", "High",
        "Pagination — Direct page number click"
    ),
    (
        "TC_PLP_020", "PLP Page",
        "Product card click navigates to the Product Detail Page (PDP)",
        PRE,
        NAV_BACK +
        "1. Click on any product card image or product name on the PLP\n"
        "2. Observe the navigation\n"
        "--- " + RESET_BACK,
        "User is navigated to the Product Detail Page (PDP)\n"
        "PDP URL is different from PLP URL (e.g. /{product-slug}?variant_id=X)\n"
        "Product name on PDP matches the card clicked on PLP\n"
        "Back navigation (goBack) returns to PLP",
        "", "Pass/Fail", "Positive", "High",
        "Product card click \u2192 PDP navigation. goBack() tested as reset."
    ),
    (
        "TC_PLP_021", "PLP Page",
        "Add to Cart button on product card adds the product to the cart",
        PRE,
        NAV_BACK +
        "1. Note the current cart icon count\n"
        "2. Click 'ADD TO CART' on any product card\n"
        "3. Observe the cart icon and any confirmation",
        "Product is added to the cart successfully\n"
        "Cart icon count increments by 1\n"
        "Confirmation toast/message is displayed\n"
        "Cart page shows the added product",
        "", "Pass/Fail", "Positive", "High",
        "Add to Cart from PLP product card"
    ),
    (
        "TC_PLP_022", "PLP Page",
        "Breadcrumb: Click 'Home' in breadcrumb navigates to Home Page",
        PRE,
        NAV_BACK +
        "1. Locate breadcrumb: Home > Jewellery\n"
        "2. Click on 'Home' in the breadcrumb\n"
        "--- " + RESET_BACK,
        "User is redirected to the Home Page\n"
        "URL changes to https://qa-sunnydiamonds.webc.in/\n"
        "goBack() returns to PLP",
        "", "Pass/Fail", "Positive", "Medium",
        "Breadcrumb navigation — Home link"
    ),
    (
        "TC_PLP_023", "PLP Page",
        "Multi-filter: Category (Rings) + Metal Color (Yellow Gold) applied together",
        PRE,
        NAV_BACK +
        "1. Select category 'Rings' from the filter\n"
        "2. Also select Metal Color 'Yellow gold'\n"
        "3. Observe the product grid and URL",
        "Products displayed are Yellow Gold Rings only\n"
        "Product count reflects the combined filter\n"
        "URL contains both filter parameters\n"
        "No unrelated products are shown",
        "", "Pass/Fail", "Positive", "High",
        "Multi-filter combination — Category + Metal Color"
    ),
    (
        "TC_PLP_024", "PLP Page",
        "Multi-filter: Metal Purity (18K) + Price Range (Below \u20b910k) applied together",
        PRE,
        NAV_BACK +
        "1. Select Metal Purity '18K' from filter\n"
        "2. Select Price 'Below \u20b910k'\n"
        "3. Observe product grid",
        "Products are 18K AND priced below \u20b910,000\n"
        "URL reflects both filter parameters\n"
        "Product count reflects combined filter",
        "", "Pass/Fail", "Positive", "High",
        "Multi-filter combination — Metal Purity + Price"
    ),
    (
        "TC_PLP_025", "PLP Page",
        "Verify PLP page title and meta description are correct for SEO",
        PRE,
        NAV_BACK +
        "1. Navigate to https://qa-sunnydiamonds.webc.in/jewellery\n"
        "2. Check browser tab title\n"
        "3. Inspect <title> and <meta name='description'> in page source",
        "Browser tab title includes 'Jewellery' and 'Sunny Diamonds'\n"
        "Meta description is present and relevant to jewellery\n"
        "Page heading H1 says 'Jewellery'",
        "", "Pass/Fail", "Positive", "Medium",
        "SEO — Page title and meta description"
    ),

    # ══════════════════ NEGATIVE ════════════════════════════════════════════
    (
        "TC_PLP_026", "PLP Page",
        "Apply filter combination that returns zero products — verify empty state",
        PRE,
        NAV_BACK +
        "1. Select an extreme price filter (e.g. 'Below \u20b910k')\n"
        "2. Also select a rare gemstone type filter (e.g. 'Aqua')\n"
        "3. Observe the product grid",
        "Empty state message is displayed (e.g. 'No products found')\n"
        "Product count shows 0\n"
        "No broken layout or error page\n"
        "User can clear filters to return to full listing",
        "", "Pass/Fail", "Negative", "High",
        "Zero-result filter combination — empty state handling"
    ),
    (
        "TC_PLP_027", "PLP Page",
        "Direct URL with invalid/non-existent category — graceful handling",
        PRE,
        NAV_BACK +
        "1. Manually type URL: https://qa-sunnydiamonds.webc.in/jewellery/invalidcategory\n"
        "2. Press Enter and observe the page",
        "System does NOT crash\n"
        "Either:\n"
        "a) 404 page is displayed with 'Back to Home' option\n"
        "OR\n"
        "b) Redirected to /jewellery with no category filter\n"
        "No unhandled exception or blank page",
        "", "Pass/Fail", "Negative", "High",
        "Invalid category URL — graceful 404 or redirect"
    ),
    (
        "TC_PLP_028", "PLP Page",
        "Navigate to PLP with invalid filter query parameter — verify fallback",
        PRE,
        NAV_BACK +
        "1. Manually type URL: https://qa-sunnydiamonds.webc.in/jewellery?variants.color=invalidcolor\n"
        "2. Observe the page response",
        "System handles invalid parameter gracefully\n"
        "Either:\n"
        "a) PLP loads with no products (empty state)\n"
        "OR\n"
        "b) PLP loads ignoring the invalid parameter\n"
        "No application error or crash",
        "", "Pass/Fail", "Negative", "High",
        "Invalid filter query param — graceful fallback"
    ),
    (
        "TC_PLP_029", "PLP Page",
        "Pagination: Navigate beyond last page (e.g. page=9999) via URL — verify handling",
        PRE,
        NAV_BACK +
        "1. On PLP, manually append page param: ?page=9999 to URL\n"
        "2. Press Enter and observe",
        "System handles out-of-range page number gracefully\n"
        "Either:\n"
        "a) Redirected back to page 1\n"
        "OR\n"
        "b) 'No products found' or empty state shown\n"
        "No application crash or 500 error",
        "", "Pass/Fail", "Negative", "High",
        "Out-of-range page number — graceful handling"
    ),

    # ══════════════════ EDGE CASES ══════════════════════════════════════════
    (
        "TC_PLP_030", "PLP Page",
        "Browser back navigation from PDP returns to PLP with filter state preserved",
        PRE,
        NAV_BACK +
        "1. Apply a filter on PLP (e.g. Category: Rings)\n"
        "2. Click a product card to navigate to PDP\n"
        "3. On PDP, call page.goBack()\n"
        "4. Observe the PLP filter state",
        "User returns to PLP via goBack()\n"
        "Previously applied filter (Rings) is still active\n"
        "Scroll position is preserved or user is returned to the product they clicked\n"
        "No full page reload required",
        "", "Pass/Fail", "Edge Case", "High",
        "goBack() from PDP \u2192 PLP filter state preservation"
    ),
    (
        "TC_PLP_031", "PLP Page",
        "PLP filter URL is bookmarkable/shareable — state preserved on direct access",
        PRE,
        NAV_BACK +
        "1. Apply filter: Yellow Gold (?variants.color=yellow-gold)\n"
        "2. Copy the URL from the browser\n"
        "3. Open the URL in a new browser tab\n"
        "4. Observe the filter state",
        "New tab loads with the same filter applied\n"
        "Yellow Gold filter is active\n"
        "Product count and grid match the filtered state\n"
        "URL is shareable and reproducible",
        "", "Pass/Fail", "Edge Case", "High",
        "Filter URL shareability and bookmarking"
    ),
    (
        "TC_PLP_032", "PLP Page",
        "Rapid consecutive filter toggling — no duplicate API calls or UI freeze",
        PRE,
        NAV_BACK +
        "1. Rapidly click different category filters one after another:\n"
        "   Rings > Earrings > Pendants > Nosepins (within 2 seconds)\n"
        "2. Observe the UI stability",
        "UI remains stable and does not freeze\n"
        "Final selected filter is applied correctly\n"
        "No duplicate API calls or incorrect product display\n"
        "Loading indicator appears and resolves",
        "", "Pass/Fail", "Edge Case", "High",
        "Rapid filter toggling — debounce and UI stability"
    ),
    (
        "TC_PLP_033", "PLP Page",
        "Sort + Filter combined — results are correctly filtered AND sorted",
        PRE,
        NAV_BACK +
        "1. Apply Category filter: 'Rings'\n"
        "2. Apply Sort: 'Price: Low to High'\n"
        "3. Compare prices of first 3 product cards",
        "Only Ring products are displayed\n"
        "Products are sorted by price in ascending order within the Rings category\n"
        "Both filter and sort are simultaneously active",
        "", "Pass/Fail", "Edge Case", "High",
        "Simultaneous filter + sort — combined result accuracy"
    ),
    (
        "TC_PLP_034", "PLP Page",
        "PLP on mobile viewport (375x812) — responsive layout verification",
        PRE,
        NAV_BACK +
        "1. Resize browser to 375x812 (iPhone viewport)\n"
        "2. Navigate to https://qa-sunnydiamonds.webc.in/jewellery\n"
        "3. Inspect layout: filter panel, product grid, sort dropdown, pagination",
        "PLP renders correctly on mobile viewport\n"
        "Filter panel is accessible (hamburger/drawer on mobile)\n"
        "Product grid adapts to single/two column layout\n"
        "Sort dropdown is usable\n"
        "Pagination is visible and tappable\n"
        "No horizontal overflow or content cut-off",
        "", "Pass/Fail", "Edge Case", "High",
        "Responsive layout — mobile viewport 375x812"
    ),
    (
        "TC_PLP_035", "PLP Page",
        "Last page of pagination (page 90) loads correctly with partial product count",
        PRE,
        NAV_BACK +
        "1. In pagination bar, click the last page number shown (e.g. 90)\n"
        "2. Wait for the page to load\n"
        "3. Count the number of products displayed",
        "Last page (90) loads correctly\n"
        "Products are displayed (may be fewer than per-page default)\n"
        "'Next' button is disabled or hidden on the last page\n"
        "No error or blank page",
        "", "Pass/Fail", "Edge Case", "High",
        "Pagination — last page boundary check"
    ),
    (
        "TC_PLP_036", "PLP Page",
        "Add the same product to cart twice from PLP — quantity increments correctly",
        PRE,
        NAV_BACK +
        "1. Note current cart count\n"
        "2. Click 'ADD TO CART' on a product card\n"
        "3. Click 'ADD TO CART' on the SAME product card again",
        "Either:\n"
        "a) Cart quantity for that product increments to 2\n"
        "OR\n"
        "b) System shows 'Already in cart' message\n"
        "No duplicate separate line items for the same product",
        "", "Pass/Fail", "Edge Case", "Medium",
        "Duplicate Add to Cart — quantity vs duplicate item handling"
    ),
    (
        "TC_PLP_037", "PLP Page",
        "PLP HTTPS enforcement — page served over secure connection",
        PRE,
        NAV_BACK +
        "1. Navigate to https://qa-sunnydiamonds.webc.in/jewellery\n"
        "2. Inspect browser address bar\n"
        "3. Check for padlock icon and HTTPS\n"
        "4. Open DevTools > Console for mixed-content warnings",
        "Page URL starts with 'https://'\n"
        "Padlock icon is visible in the browser address bar\n"
        "No mixed-content warnings in console\n"
        "No 'Not Secure' warning",
        "", "Pass/Fail", "Edge Case", "High",
        "Security — HTTPS enforcement on PLP"
    ),
    (
        "TC_PLP_038", "PLP Page",
        "Price range filter: Above \u20b980k — verify only high-value products are shown",
        PRE,
        NAV_BACK +
        "1. Select Price filter 'Above \u20b980k'\n"
        "2. URL updates to ?range=80000TO10000000\n"
        "3. Inspect the prices of displayed product cards",
        "Only products priced above \u20b980,000 are displayed\n"
        "URL contains '?range=80000TO10000000'\n"
        "No products priced below \u20b980,000 appear in the grid",
        "", "Pass/Fail", "Edge Case", "High",
        "Price filter edge — Above \u20b980k (max range boundary)"
    ),

    # ══════════════════ POSITIVE — VALID LOGIN (MUST BE LAST) ══════════════
    (
        "TC_PLP_039", "PLP Page",
        "Login with valid credentials and verify Add to Cart on PLP works with authenticated session",
        PRE_LOGIN,
        NAV_BACK +
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' — verify successful login\n"
        "5. Navigate to https://qa-sunnydiamonds.webc.in/jewellery\n"
        "6. Click 'ADD TO CART' on any product card\n"
        "7. Verify cart update\n"
        "POST-TEST RESET:\n"
        "- Click Profile > Log Out > Confirm LOG OUT",
        "Login is successful\n"
        "User is authenticated (profile icon visible)\n"
        "On PLP, clicking 'ADD TO CART' adds the product to cart\n"
        "Cart icon count increments\n"
        "Confirmation message is displayed\n"
        "No 'Login required' redirect occurs",
        "", "Pass/Fail", "Positive", "High",
        "VALID LOGIN — PLACED LAST as required. "
        "Email: sreejith.s+4@webandcrafts.com | Password: Password. "
        "Verifies authenticated Add to Cart on PLP."
    ),
]

# ── Write test cases to sheet ─────────────────────────────────────────────────
TYPE_COLORS = {"Positive": POS_GREEN, "Negative": NEG_RED, "Edge Case": EDGE_YELLOW}
PRI_COLORS  = {"High": "FFD7D7", "Critical": "FFB3B3", "Medium": "FFF3CD", "Low": "E8F5E9"}

for r_idx, tc in enumerate(TC):
    row = 5 + r_idx
    ws.row_dimensions[row].height = 130
    (tc_id, mod, desc, pre, steps, exp, act, status, ttype, pri, rem) = tc
    bg = TYPE_COLORS.get(ttype, WHITE)
    for c_idx, val in enumerate([tc_id, mod, desc, pre, steps, exp, act, status, ttype, pri, rem], 1):
        c = ws.cell(row=row, column=c_idx, value=val)
        c.font      = Font(name="Arial", size=9, color="000000", bold=(c_idx == 1))
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border    = thin
    # Test Type text colour
    ws.cell(row=row, column=9).font = Font(name="Arial", size=9, bold=True,
        color=("1F6B2E" if ttype == "Positive" else
               "7B1E0C" if ttype == "Negative" else "7B5B00"))
    # Priority colour
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=PRI_COLORS.get(pri, WHITE))
    ws.cell(row=row, column=10).font = Font(name="Arial", size=9, bold=True,
        color=("990000" if pri == "Critical" else "000000"))

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14, 14, 42, 34, 60, 44, 18, 12, 14, 10, 38]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C5"

wb.save(_tmp)

# Try to replace the original; if locked, keep as _plp_tmp.xlsx
try:
    if os.path.exists(FILE):
        os.replace(_tmp, FILE)
    else:
        shutil.copy2(_tmp, FILE)
        os.remove(_tmp)
    _saved_to = FILE
except PermissionError:
    _saved_to = _tmp

# Summary
pos  = sum(1 for t in TC if t[8] == "Positive")
neg  = sum(1 for t in TC if t[8] == "Negative")
edge = sum(1 for t in TC if t[8] == "Edge Case")
print(f"SUCCESS — '{SHEET}' sheet written to: {_saved_to}")
print(f"Total TCs : {len(TC)}")
print(f"Positive  : {pos}")
print(f"Negative  : {neg}")
print(f"Edge Case : {edge}")
print(f"Last TC   : {TC[-1][0]} | Type: {TC[-1][8]}")
print(f"Sheets    : {wb.sheetnames}")
