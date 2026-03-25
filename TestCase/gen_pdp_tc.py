# -*- coding: utf-8 -*-
"""
FRIDAY — Senior QA Agent
PDP (Product Detail Page) Test Case Generator
Target : SunnyDiamonds_v2.xlsx  →  sheet "PDP Page"
URL    : https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=45
"""

import shutil, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── File / Sheet config ────────────────────────────────────────────────────────
FILE  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET = "PDP Page"
PDP_URL = "https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=45"

# Work on a temp copy so we can save even if the original is open in Excel
_tmp = FILE.replace(".xlsx", "_pdp_tmp.xlsx")
shutil.copy2(FILE, _tmp)
wb = load_workbook(_tmp)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GOLD_LIGHT  = "FFF3CD"
POS_GREEN   = "E2EFDA"
NEG_RED     = "FCE4D6"
EDGE_YELLOW = "FFF2CC"
WHITE       = "FFFFFF"
HEADER_FONT_COLOR = "FFFFFF"

PRI_COLORS  = {
    "Critical": "FFD7D7",
    "High"    : "FFE5CC",
    "Medium"  : "FFFACC",
    "Low"     : "E2EFDA",
}

# ── Navigation helpers ─────────────────────────────────────────────────────────
NAV_FIRST = f"NAVIGATION: page.goto('{PDP_URL}')\n"
NAV_BACK  = "NAVIGATION: page.goBack()  [use goto() as fallback if goBack() fails]\n"

# ── Common preconditions ───────────────────────────────────────────────────────
PRE_OPEN  = (
    "- Application is accessible at https://qa-sunnydiamonds.webc.in\n"
    "- Browser is open and running\n"
    f"- PDP URL: {PDP_URL}\n"
    "- User is NOT logged in (guest mode)"
)
PRE_LOGIN = (
    "- Application is accessible at https://qa-sunnydiamonds.webc.in\n"
    "- Browser is open and running\n"
    "- Valid user credentials available:\n"
    "    Email   : sreejith.s+4@webandcrafts.com\n"
    "    Password: Password\n"
    "- User is NOT yet logged in"
)

# ── Test Cases  (TC_PDP_001 … TC_PDP_044) ─────────────────────────────────────
# Tuple layout:
# (id, module, description, preconditions, steps, expected, actual, status,
#  test_type, priority, remarks)

TC = [

    # ─── POSITIVE — Page Load & Display ──────────────────────────────────────
    ("TC_PDP_001", "PDP Page",
     "Verify PDP page loads successfully for a valid product URL",
     PRE_OPEN,
     NAV_FIRST +
     "1. Navigate to: " + PDP_URL + "\n"
     "2. Wait for page to fully load\n"
     "3. Observe page title, product name, price, and images",
     "- Page loads without errors (HTTP 200)\n"
     "- Product title '18 K ROSE GOLD MIA DIAMOND PENDANT' is visible\n"
     "- Product price ₹47,419 is displayed\n"
     "- Product images are loaded correctly\n"
     "- No 404 or error page is shown",
     "", "Pass/Fail", "Positive", "Critical",
     "Smoke TC — critical baseline for all PDP tests"),

    ("TC_PDP_002", "PDP Page",
     "Verify product title and SKU are displayed correctly on PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Observe the product heading\n"
     "2. Verify SKU number is displayed below the title",
     "- Title: '18 K ROSE GOLD MIA DIAMOND PENDANT'\n"
     "- SKU: '1120821883027' is visible\n"
     "- Title font is prominent and readable",
     "", "Pass/Fail", "Positive", "High",
     "Validates product identity fields"),

    ("TC_PDP_003", "PDP Page",
     "Verify product price (MRP) is displayed with 'Inclusive of all taxes' label",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the price section on PDP\n"
     "2. Verify price amount and tax label",
     "- Price ₹47,419 is prominently displayed\n"
     "- Label '(MRP Inclusive of all taxes)' appears below or beside price\n"
     "- Price format is correct (₹ symbol + Indian number format)",
     "", "Pass/Fail", "Positive", "High",
     "Price transparency validation"),

    ("TC_PDP_004", "PDP Page",
     "Verify product rating (4.5 stars) is displayed on PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the star rating section near the product title\n"
     "2. Verify rating value is visible",
     "- Rating '(4.5)' is visible\n"
     "- Star icons or numeric rating is displayed\n"
     "- Rating is positioned near the product title/SKU",
     "", "Pass/Fail", "Positive", "Medium",
     "Social proof element validation"),

    ("TC_PDP_005", "PDP Page",
     "Verify breadcrumb navigation is displayed correctly (Home > Product Name)",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the breadcrumb at the top of the page\n"
     "2. Verify breadcrumb links and structure",
     "- Breadcrumb shows: Home > 18 K Rose Gold Mia Diamond Pendant\n"
     "- 'Home' link is clickable and navigates to homepage\n"
     "- Product name segment is the current active page",
     "", "Pass/Fail", "Positive", "Medium",
     "Navigation hierarchy validation"),

    # ─── POSITIVE — Product Gallery ──────────────────────────────────────────
    ("TC_PDP_006", "PDP Page",
     "Verify product image gallery loads with multiple images",
     PRE_OPEN,
     NAV_BACK +
     "1. Observe the product image gallery/slider section\n"
     "2. Count available images\n"
     "3. Check image quality and loading state",
     "- Multiple product images are loaded (31 images expected)\n"
     "- Images are clear, not broken, and product-specific\n"
     "- Gallery is interactive (carousel/slider navigation visible)",
     "", "Pass/Fail", "Positive", "High",
     "Gallery contains 31 product images per crawl analysis"),

    ("TC_PDP_007", "PDP Page",
     "Verify product gallery thumbnail navigation works correctly",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the thumbnail strip in the product gallery\n"
     "2. Click on a different thumbnail image\n"
     "3. Observe the main image change",
     "- Clicking a thumbnail updates the main product image\n"
     "- Active thumbnail is visually highlighted\n"
     "- Image transition is smooth (no errors)",
     "", "Pass/Fail", "Positive", "Medium",
     "Thumbnail-to-main image sync test"),

    # ─── POSITIVE — Price Breakup ─────────────────────────────────────────────
    ("TC_PDP_008", "PDP Page",
     "Verify Price Breakup toggle expands and shows detailed breakdown",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'Price Breakup' button/toggle on PDP\n"
     "2. Click the 'Price Breakup' button\n"
     "3. Verify the expanded breakdown details",
     "- Price Breakup section expands on click\n"
     "- Metal Price: ₹9,372 is visible\n"
     "- Diamond Price: ₹33,450 is visible\n"
     "- Making Charge: ₹3,215.46 is visible\n"
     "- GST (3%): ₹1,381.124 is visible\n"
     "- Total: ₹47,419 is visible\n"
     "- Disclaimer note about actual weight variation is shown",
     "", "Pass/Fail", "Positive", "High",
     "Price transparency — detailed breakup toggle"),

    ("TC_PDP_009", "PDP Page",
     "Verify Price Breakup section collapses when toggled again",
     PRE_OPEN,
     NAV_BACK +
     "1. Click 'Price Breakup' button to expand\n"
     "2. Click 'Price Breakup' button again to collapse\n"
     "3. Verify the section is hidden",
     "- Price Breakup section collapses on second click\n"
     "- Detailed pricing rows are no longer visible\n"
     "- Only main price (₹47,419) remains visible",
     "", "Pass/Fail", "Positive", "Medium",
     "Toggle collapse behavior validation"),

    ("TC_PDP_010", "PDP Page",
     "Verify Fair Pricing comparison section shows Sunny Diamonds vs Estimated Retail",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'Fair Pricing' section on the PDP\n"
     "2. Observe the price comparison",
     "- 'Fair Pricing' label or 'Learn More' link is visible\n"
     "- Sunny Diamonds price: ₹47,419 is shown\n"
     "- Estimated Retail price: ₹52,160 is shown\n"
     "- Comparison visually demonstrates savings",
     "", "Pass/Fail", "Positive", "Medium",
     "Competitive pricing transparency section"),

    # ─── POSITIVE — Quantity Input ────────────────────────────────────────────
    ("TC_PDP_011", "PDP Page",
     "Verify default quantity is 1 and increment button works correctly",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the Quantity section on PDP\n"
     "2. Verify default quantity value is '1'\n"
     "3. Click the '+' (increment) button\n"
     "4. Observe the quantity value",
     "- Default quantity is 1\n"
     "- Clicking '+' increments quantity to 2\n"
     "- Quantity field updates correctly\n"
     "- No page reload occurs",
     "", "Pass/Fail", "Positive", "High",
     "Quantity stepper — increment validation"),

    ("TC_PDP_012", "PDP Page",
     "Verify quantity decrement button reduces quantity by 1",
     PRE_OPEN,
     NAV_BACK +
     "1. First increment quantity to 3 by clicking '+' twice\n"
     "2. Click the '-' (decrement) button\n"
     "3. Observe quantity value",
     "- Quantity decreases from 3 to 2 on click\n"
     "- Decrement works correctly\n"
     "- Quantity field displays updated value",
     "", "Pass/Fail", "Positive", "Medium",
     "Quantity stepper — decrement validation"),

    # ─── POSITIVE — Check Availability (Pincode) ─────────────────────────────
    ("TC_PDP_013", "PDP Page",
     "Verify 'Check Availability' section accepts valid pincode and returns result",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'Check availability' section on PDP\n"
     "2. Click on the pincode input field\n"
     "3. Enter a valid 6-digit pincode: '682035'\n"
     "4. Click the check/submit button\n"
     "5. Observe the availability result",
     "- Pincode input accepts 6-digit numeric value\n"
     "- On submit, delivery availability result is displayed\n"
     "- Result shows estimated delivery date or availability message\n"
     "- No error is thrown for valid pincode",
     "", "Pass/Fail", "Positive", "High",
     "Valid pincode: 682035 (Ernakulam, Kerala)"),

    # ─── POSITIVE — Metal Details Accordion ──────────────────────────────────
    ("TC_PDP_014", "PDP Page",
     "Verify 'Metal Details' accordion expands and shows product metal specifications",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'METAL DETAILS' accordion button on PDP\n"
     "2. Click to expand the Metal Details section\n"
     "3. Verify the displayed metal information",
     "- Metal Details section expands\n"
     "- Metal Weight: 1 (gm) is displayed\n"
     "- Setting Type: OPEN is displayed\n"
     "- Metal Purity: 18K is displayed\n"
     "- Metal Type: GOLD is displayed\n"
     "- Metal Color: ROSE GOLD is displayed",
     "", "Pass/Fail", "Positive", "High",
     "Accordion: Metal specifications validation"),

    ("TC_PDP_015", "PDP Page",
     "Verify 'Diamond Details' accordion expands and shows diamond specifications",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'DIAMOND DETAILS' accordion button on PDP\n"
     "2. Click to expand the Diamond Details section\n"
     "3. Verify the displayed diamond information",
     "- Diamond Details section expands\n"
     "- Diamond specifications (cut, weight, clarity, etc.) are displayed\n"
     "- Diamond Price ₹33,450 referenced information is shown\n"
     "- Section layout is readable and properly formatted",
     "", "Pass/Fail", "Positive", "High",
     "Accordion: Diamond specifications validation"),

    ("TC_PDP_016", "PDP Page",
     "Verify 'Manufactured By' accordion expands and shows manufacturer details",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'MANUFACTURED BY' accordion button on PDP\n"
     "2. Click to expand the Manufactured By section\n"
     "3. Verify the manufacturer information",
     "- Manufactured By section expands\n"
     "- Manufacturer details are displayed\n"
     "- Section collapses when clicked again",
     "", "Pass/Fail", "Positive", "Medium",
     "Accordion: Manufacturer info validation"),

    # ─── POSITIVE — Add to Cart ───────────────────────────────────────────────
    ("TC_PDP_017", "PDP Page",
     "Verify 'Add to Cart' button is visible and clickable on PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'ADD TO CART' button (primary CTA) on PDP\n"
     "2. Verify button is visible and enabled\n"
     "3. Observe button styling and placement",
     "- 'ADD TO CART' button is prominently displayed\n"
     "- Button is enabled (not disabled/greyed out)\n"
     "- Button style matches brand design (secondary button style)\n"
     "- Both sticky header and product section CTAs are present",
     "", "Pass/Fail", "Positive", "Critical",
     "Primary CTA visibility check (guest mode)"),

    ("TC_PDP_018", "PDP Page",
     "Verify clicking 'Add to Cart' as guest user triggers login redirect or adds to cart",
     PRE_OPEN,
     NAV_BACK +
     "1. Ensure user is NOT logged in\n"
     "2. Locate the 'ADD TO CART' button\n"
     "3. Click the 'ADD TO CART' button\n"
     "4. Observe the behavior (redirect to login OR cart update)",
     "- Either: User is redirected to login page with a redirect-back URL\n"
     "- Or: Item is added to guest cart and cart count updates\n"
     "- No unhandled error or blank screen occurs\n"
     "- Appropriate feedback (toast/modal) is shown to the user",
     "", "Pass/Fail", "Positive", "Critical",
     "Guest Add to Cart flow — login redirect or guest cart behavior"),

    # ─── POSITIVE — Buy Now ───────────────────────────────────────────────────
    ("TC_PDP_019", "PDP Page",
     "Verify 'Buy Now' button is visible and clickable on PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'BUY NOW' button on PDP\n"
     "2. Verify button is visible and enabled\n"
     "3. Observe button placement relative to Add to Cart",
     "- 'BUY NOW' button is visible\n"
     "- Button is enabled and clickable\n"
     "- Button is positioned alongside the ADD TO CART button\n"
     "- Visual differentiation between BUY NOW and ADD TO CART is clear",
     "", "Pass/Fail", "Positive", "Critical",
     "Secondary CTA — Buy Now visibility"),

    ("TC_PDP_020", "PDP Page",
     "Verify clicking 'Buy Now' as guest redirects to login or initiates checkout",
     PRE_OPEN,
     NAV_BACK +
     "1. Ensure user is NOT logged in\n"
     "2. Click the 'BUY NOW' button on PDP\n"
     "3. Observe the resulting action",
     "- User is redirected to login page (if authentication required)\n"
     "- OR: Checkout flow is initiated\n"
     "- No JavaScript errors or blank screen\n"
     "- User is not stuck or confused by the outcome",
     "", "Pass/Fail", "Positive", "High",
     "Guest Buy Now flow behavior"),

    # ─── POSITIVE — You May Also Like ─────────────────────────────────────────
    ("TC_PDP_021", "PDP Page",
     "Verify 'You May Also Like' section displays related product recommendations",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll down to the 'YOU MAY ALSO LIKE' section\n"
     "2. Observe the recommended products displayed",
     "- 'YOU MAY ALSO LIKE' heading is visible\n"
     "- At least 10 related products are displayed\n"
     "- Each product shows: image, name, price\n"
     "- Each product has an 'ADD TO CART' button",
     "", "Pass/Fail", "Positive", "Medium",
     "Related products section — 10 items expected"),

    ("TC_PDP_022", "PDP Page",
     "Verify clicking a related product in 'You May Also Like' navigates to its PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll to the 'YOU MAY ALSO LIKE' section\n"
     "2. Click on the product name or image of 'Amaury Diamond Earring'\n"
     "3. Observe navigation",
     "- User is navigated to the PDP of 'Amaury Diamond Earring' (₹33,726)\n"
     "- New PDP loads correctly with correct product details\n"
     "- URL updates to reflect the new product slug",
     "", "Pass/Fail", "Positive", "Medium",
     "Cross-sell navigation test"),

    # ─── POSITIVE — Social Share ──────────────────────────────────────────────
    ("TC_PDP_023", "PDP Page",
     "Verify Social Share section is visible on PDP",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'Social Share' section on PDP\n"
     "2. Verify share options are displayed",
     "- 'Social Share' / 'Share this link via' text is visible\n"
     "- At least one sharing option is displayed (Facebook, WhatsApp, copy link, etc.)\n"
     "- Share section is interactive",
     "", "Pass/Fail", "Positive", "Low",
     "Social share widget visibility test"),

    # ─── POSITIVE — Trust Badges & Policies ──────────────────────────────────
    ("TC_PDP_024", "PDP Page",
     "Verify 'Our Promise' trust badges are displayed below product details",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll below the 'You May Also Like' section\n"
     "2. Locate the 'PURE. CERTIFIED. TRUSTED.' section\n"
     "3. Verify all trust badges are visible",
     "- 'PURE. CERTIFIED. TRUSTED.' heading is visible\n"
     "- 8 trust badges displayed:\n"
     "  • Internally Flawless Diamonds\n"
     "  • 100% Money Back on Diamond Value\n"
     "  • Certifications of Diamonds\n"
     "  • BIS Hall Mark for Jewellery\n"
     "  • Brand Assured Quality\n"
     "  • 15 Days Return Policy\n"
     "  • Cash On Delivery\n"
     "  • Pan India Free Shipping",
     "", "Pass/Fail", "Positive", "Medium",
     "Trust signals — all 8 badges verified"),

    # ─── POSITIVE — Newsletter Subscribe ─────────────────────────────────────
    ("TC_PDP_025", "PDP Page",
     "Verify newsletter subscription with a valid email address",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll to the footer newsletter section\n"
     "2. Locate the email input field with placeholder 'Your email *'\n"
     "3. Enter a valid email: 'testuser@example.com'\n"
     "4. Click the 'SUBSCRIBE' button\n"
     "5. Observe the response",
     "- Email input accepts the valid email address\n"
     "- On SUBSCRIBE click, a success message is displayed\n"
     "- Message such as 'Thank you for subscribing!' or similar\n"
     "- No server error occurs",
     "", "Pass/Fail", "Positive", "Low",
     "Newsletter subscription happy path"),

    # ─── POSITIVE — Breadcrumb Navigation ────────────────────────────────────
    ("TC_PDP_026", "PDP Page",
     "Verify clicking 'Home' breadcrumb navigates back to the homepage",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the breadcrumb at the top of the PDP\n"
     "2. Click the 'Home' link in the breadcrumb\n"
     "3. Observe navigation",
     "- User is navigated to the homepage: https://qa-sunnydiamonds.webc.in/\n"
     "- Homepage loads correctly\n"
     "- URL changes to the root URL",
     "", "Pass/Fail", "Positive", "Medium",
     "Breadcrumb back navigation"),

    # ─── POSITIVE — Cookie Consent ────────────────────────────────────────────
    ("TC_PDP_027", "PDP Page",
     "Verify Cookie Consent banner appears and 'Accept All' button works",
     PRE_OPEN,
     NAV_BACK +
     "1. Open PDP in a fresh browser session (clear cookies first)\n"
     "2. Observe the cookie consent banner\n"
     "3. Click 'Accept All'\n"
     "4. Verify banner disappears",
     "- Cookie consent banner: 'We Value Your Privacy' is visible on first visit\n"
     "- 'Accept All' and 'Decline' buttons are present\n"
     "- Clicking 'Accept All' dismisses the banner\n"
     "- Cookie preference is saved (banner not shown on next visit)",
     "", "Pass/Fail", "Positive", "Low",
     "GDPR/cookie consent flow"),

    ("TC_PDP_028", "PDP Page",
     "Verify clicking 'Decline' on cookie consent dismisses the banner",
     PRE_OPEN,
     NAV_BACK +
     "1. Open PDP in a fresh browser session\n"
     "2. When cookie consent banner appears, click 'Decline'\n"
     "3. Observe banner behavior",
     "- Banner is dismissed after clicking 'Decline'\n"
     "- Non-essential cookies are not set\n"
     "- Page remains functional after declining",
     "", "Pass/Fail", "Positive", "Low",
     "Cookie decline flow — page should remain usable"),

    # ─── NEGATIVE — Invalid Pincode ──────────────────────────────────────────
    ("TC_PDP_029", "PDP Page",
     "Verify error message when invalid pincode (less than 6 digits) is entered",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the 'Check availability' pincode input\n"
     "2. Enter an invalid pincode with only 4 digits: '6820'\n"
     "3. Click the check button\n"
     "4. Observe the error handling",
     "- System shows validation error: 'Please enter a valid 6-digit pincode'\n"
     "- OR: check button is disabled until 6 digits are entered\n"
     "- No server call is made for incomplete input\n"
     "- Pincode field is highlighted in red",
     "", "Pass/Fail", "Negative", "High",
     "Invalid pincode — less than 6 digits"),

    ("TC_PDP_030", "PDP Page",
     "Verify error when non-numeric (alphabetic) value is entered in pincode field",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the pincode input field\n"
     "2. Enter alphabetic characters: 'ABCDEF'\n"
     "3. Click the check button",
     "- Input field rejects or ignores non-numeric characters\n"
     "- OR: Validation error 'Please enter a valid numeric pincode' is shown\n"
     "- System does not crash or show unhandled error",
     "", "Pass/Fail", "Negative", "Medium",
     "Non-numeric pincode input handling"),

    ("TC_PDP_031", "PDP Page",
     "Verify error when an unserviceable/unavailable pincode is submitted",
     PRE_OPEN,
     NAV_BACK +
     "1. Enter a valid-format but unserviceable pincode: '110001' (New Delhi)\n"
     "2. Click the check availability button\n"
     "3. Observe the response message",
     "- System shows a meaningful 'Delivery not available' message\n"
     "- OR: Message like 'Sorry, we do not deliver to this pincode'\n"
     "- No JavaScript error is thrown\n"
     "- User is not left with a blank/loading state",
     "", "Pass/Fail", "Negative", "High",
     "Unserviceable area response"),

    # ─── NEGATIVE — Quantity Input ────────────────────────────────────────────
    ("TC_PDP_032", "PDP Page",
     "Verify quantity field rejects 0 or negative values",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the quantity input field (min=1)\n"
     "2. Manually type '0' in the quantity field\n"
     "3. Click 'ADD TO CART'\n"
     "4. Observe the behavior",
     "- System prevents quantity 0: field resets to 1 or shows validation error\n"
     "- OR: ADD TO CART is disabled when quantity is 0\n"
     "- Negative quantity is not accepted\n"
     "- No product with quantity 0 is added to cart",
     "", "Pass/Fail", "Negative", "High",
     "Quantity min-boundary negative test — min=1"),

    # ─── NEGATIVE — Newsletter ────────────────────────────────────────────────
    ("TC_PDP_033", "PDP Page",
     "Verify newsletter subscription rejects an invalid email format",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll to the footer newsletter section\n"
     "2. Enter invalid email: 'notanemail'\n"
     "3. Click 'SUBSCRIBE'\n"
     "4. Observe validation behavior",
     "- System displays validation error: 'Please enter a valid email address'\n"
     "- OR: HTML5 email validation prevents form submission\n"
     "- Subscription is NOT processed for invalid email\n"
     "- Error message is clearly visible to the user",
     "", "Pass/Fail", "Negative", "Medium",
     "Newsletter — invalid email format rejection"),

    # ─── NEGATIVE — Invalid variant_id in URL ────────────────────────────────
    ("TC_PDP_034", "PDP Page",
     "Verify PDP behavior when accessed with an invalid variant_id in URL",
     PRE_OPEN,
     "NAVIGATION: page.goto('https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=99999')\n"
     "1. Navigate to PDP URL with non-existent variant_id=99999\n"
     "2. Observe the page response",
     "- Page handles the invalid variant gracefully\n"
     "- Either: Default variant is shown\n"
     "- Or: Appropriate 'Variant not found' message is displayed\n"
     "- Application does NOT crash or show a 500 error",
     "", "Pass/Fail", "Negative", "Medium",
     "Invalid variant_id edge — error handling"),

    # ─── EDGE CASES ───────────────────────────────────────────────────────────
    ("TC_PDP_035", "PDP Page",
     "Verify quantity field boundary — maximum allowed integer value",
     PRE_OPEN,
     NAV_BACK +
     "1. Locate the quantity input field\n"
     "2. Clear the field and type an extremely large number: '9999999'\n"
     "3. Click 'ADD TO CART'\n"
     "4. Observe the system response",
     "- System either accepts large quantity or shows max-order limit warning\n"
     "- No JavaScript runtime error\n"
     "- If limit exists, error message 'Maximum quantity allowed is X' is shown\n"
     "- Price is not incorrectly calculated for extreme values",
     "", "Pass/Fail", "Edge Case", "Medium",
     "BVA: Quantity upper boundary — no max defined in HTML"),

    ("TC_PDP_036", "PDP Page",
     "Verify quantity accepts exactly 1 (minimum boundary value) — BVA",
     PRE_OPEN,
     NAV_BACK +
     "1. Set quantity input to exactly '1' (minimum value)\n"
     "2. Click 'ADD TO CART'\n"
     "3. Observe behavior",
     "- Quantity 1 is accepted\n"
     "- ADD TO CART proceeds without validation error\n"
     "- Item is added to cart with quantity 1",
     "", "Pass/Fail", "Edge Case", "Medium",
     "BVA: Quantity min boundary — exactly 1"),

    ("TC_PDP_037", "PDP Page",
     "Verify PDP loads correctly on mobile viewport (375×812 — iPhone SE)",
     PRE_OPEN,
     "NAVIGATION: page.setViewportSize({width: 375, height: 812}), then page.goto('" + PDP_URL + "')\n"
     "1. Set viewport to 375x812 (mobile)\n"
     "2. Navigate to PDP URL\n"
     "3. Observe responsive layout",
     "- Page is fully responsive on mobile viewport\n"
     "- Product title, price, and CTA buttons are visible without horizontal scroll\n"
     "- Images resize correctly\n"
     "- Mobile sticky navbar (Home, Profile, Cart, Search) is visible\n"
     "- ADD TO CART and BUY NOW buttons are accessible",
     "", "Pass/Fail", "Edge Case", "High",
     "Responsive design — mobile 375px viewport"),

    ("TC_PDP_038", "PDP Page",
     "Verify PDP loads on tablet viewport (768×1024 — iPad)",
     PRE_OPEN,
     "NAVIGATION: page.setViewportSize({width: 768, height: 1024}), then page.goto('" + PDP_URL + "')\n"
     "1. Set viewport to 768×1024 (tablet)\n"
     "2. Navigate to PDP\n"
     "3. Observe layout adaptation",
     "- Page adapts to tablet layout without overflow\n"
     "- Product image and details are side-by-side or stacked appropriately\n"
     "- All CTAs, accordions, and sections remain accessible",
     "", "Pass/Fail", "Edge Case", "Medium",
     "Responsive design — tablet 768px viewport"),

    ("TC_PDP_039", "PDP Page",
     "Verify PDP URL with no variant_id parameter loads default product variant",
     PRE_OPEN,
     "NAVIGATION: page.goto('https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant')\n"
     "1. Navigate to PDP URL without any variant_id query parameter\n"
     "2. Observe the product loaded",
     "- Page loads without error\n"
     "- Default product variant is displayed\n"
     "- Product title and price are shown\n"
     "- No 404 or unhandled error page",
     "", "Pass/Fail", "Edge Case", "Medium",
     "URL parameter edge — no variant_id provided"),

    ("TC_PDP_040", "PDP Page",
     "Verify pincode field accepts exactly 6-digit numeric input (BVA — boundary)",
     PRE_OPEN,
     NAV_BACK +
     "1. Enter exactly 6-digit pincode: '682035'\n"
     "2. Click check availability\n"
     "3. Then clear and enter 7-digit: '6820350'\n"
     "4. Click check availability again",
     "- 6-digit pincode: check proceeds without validation error\n"
     "- 7-digit pincode: validation error 'Enter a valid 6-digit pincode'\n"
     "- OR: Input is limited to 6 characters maximum via maxLength attribute",
     "", "Pass/Fail", "Edge Case", "High",
     "BVA: Pincode boundary — exactly 6 digits valid, 7 digits invalid"),

    ("TC_PDP_041", "PDP Page",
     "Verify XSS injection attempt in pincode field is sanitised — Security",
     PRE_OPEN,
     NAV_BACK +
     "1. Click the pincode input field\n"
     "2. Enter XSS payload: '<script>alert(1)</script>'\n"
     "3. Click check availability button\n"
     "4. Observe page behavior",
     "- XSS script is NOT executed in the browser\n"
     "- No JavaScript alert popup appears\n"
     "- Input is sanitised/escaped by the application\n"
     "- Error message for invalid pincode format is shown",
     "", "Pass/Fail", "Edge Case", "Critical",
     "OWASP A03 — XSS injection in pincode field"),

    ("TC_PDP_042", "PDP Page",
     "Verify newsletter email field rejects XSS payload — Security",
     PRE_OPEN,
     NAV_BACK +
     "1. Scroll to newsletter subscribe section\n"
     "2. Enter XSS payload in email field: '<img src=x onerror=alert(1)>'\n"
     "3. Click SUBSCRIBE\n"
     "4. Observe page behavior",
     "- XSS script is NOT executed\n"
     "- HTML tags are sanitised/stripped\n"
     "- Validation error is shown for invalid email format\n"
     "- No alert popup or DOM manipulation occurs",
     "", "Pass/Fail", "Edge Case", "Critical",
     "OWASP A03 — XSS in newsletter email field"),

    ("TC_PDP_043", "PDP Page",
     "Verify browser back/forward navigation preserves PDP state correctly",
     PRE_OPEN,
     NAV_BACK +
     "1. Open PDP via direct URL\n"
     "2. Expand 'Price Breakup' section\n"
     "3. Navigate to a related product in 'You May Also Like'\n"
     "4. Click browser Back button\n"
     "5. Observe original PDP state",
     "- Browser navigates back to original PDP\n"
     "- Product title and price are retained\n"
     "- Page loads without errors on back navigation\n"
     "- No infinite loop or navigation error",
     "", "Pass/Fail", "Edge Case", "High",
     "Browser back navigation state persistence test"),

    # ─── VALID LOGIN (MUST BE LAST) ───────────────────────────────────────────
    ("TC_PDP_044", "PDP Page",
     "Login with valid credentials and verify Add to Cart on PDP works with authenticated session",
     PRE_LOGIN,
     NAV_BACK +
     "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
     "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
     "3. Enter Password: 'Password'\n"
     "4. Click the Sign In / Login button\n"
     "5. Verify successful login (redirect away from login page)\n"
     "6. Navigate to PDP: " + PDP_URL + "\n"
     "7. Click 'ADD TO CART' button\n"
     "8. Observe cart behavior for authenticated user",
     "- Login is successful (user is redirected to homepage or dashboard)\n"
     "- PDP loads correctly for the authenticated user\n"
     "- 'ADD TO CART' click adds the product to cart\n"
     "- Cart icon/count updates to reflect the added item\n"
     "- Success toast/notification: 'Item added to cart' (or similar) is shown\n"
     "- No authentication error or redirect-to-login occurs",
     "", "Pass/Fail", "Positive", "Critical",
     "VALID LOGIN — PLACED LAST as required | Email: sreejith.s+4@webandcrafts.com"),
]

# ── Banner / Metadata rows ────────────────────────────────────────────────────
import datetime
TODAY = datetime.date(2026, 3, 25).strftime("%d-%B-%Y")

TOTAL_COLS = 11
BANNER_FILL   = PatternFill("solid", fgColor=DARK_BLUE)
META_FILL     = PatternFill("solid", fgColor=MID_BLUE)
SUBHDR_FILL   = PatternFill("solid", fgColor=LIGHT_BLUE)
SEP_FILL      = PatternFill("solid", fgColor=GOLD_LIGHT)

thin = Side(style="thin", color="B0B0B0")
thick = Side(style="medium", color="1F3864")
THIN_BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

def banner_row(ws, row_num, text, fill, font_color=HEADER_FONT_COLOR, size=13, bold=True):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=TOTAL_COLS)
    cell = ws.cell(row=row_num, column=1)
    cell.value = text
    cell.fill  = fill
    cell.font  = Font(name="Arial", size=size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row 1 — Project banner
banner_row(ws, 1, "🔶  SUNNY DIAMONDS — QA TEST SUITE  🔶",
           PatternFill("solid", fgColor=DARK_BLUE), size=14)
ws.row_dimensions[1].height = 28

# Row 2 — Module banner
banner_row(ws, 2, "MODULE: PDP PAGE (Product Detail Page)  |  URL: " + PDP_URL,
           PatternFill("solid", fgColor=MID_BLUE), size=11)
ws.row_dimensions[2].height = 22

# Row 3 — Metadata
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
meta_cell = ws.cell(row=3, column=1)
meta_cell.value = (
    f"Test Cases Created By: Sreejith S Madavan     |     "
    f"Created Date: {TODAY}     |     "
    f"Total TCs: {len(TC)}     |     "
    f"Environment: QA     |     "
    f"Tool: Playwright (Chromium)"
)
meta_cell.fill = META_FILL
meta_cell.font = Font(name="Arial", size=10, bold=False, color=HEADER_FONT_COLOR)
meta_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[3].height = 18

# Row 4 — Column headers
HEADERS = [
    "Test Case ID", "Module Name", "Test Case Description",
    "Preconditions", "Test Steps", "Expected Result",
    "Actual Result", "Status", "Test Type", "Priority", "Remarks"
]
HDR_FILL = PatternFill("solid", fgColor="0D2B5E")
for col_idx, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = hdr
    cell.fill  = HDR_FILL
    cell.font  = Font(name="Arial", size=10, bold=True, color=HEADER_FONT_COLOR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
ws.row_dimensions[4].height = 22

# ── Type → fill colour map ────────────────────────────────────────────────────
TYPE_FILL = {
    "Positive" : PatternFill("solid", fgColor=POS_GREEN),
    "Negative" : PatternFill("solid", fgColor=NEG_RED),
    "Edge Case": PatternFill("solid", fgColor=EDGE_YELLOW),
}

# ── Write test case rows ──────────────────────────────────────────────────────
for row_offset, tc in enumerate(TC):
    row = 5 + row_offset
    (tc_id, module, desc, pre, steps, expected,
     actual, status, tc_type, priority, remarks) = tc

    values = [tc_id, module, desc, pre, steps, expected,
              actual, status, tc_type, priority, remarks]
    row_fill = TYPE_FILL.get(tc_type, PatternFill("solid", fgColor=WHITE))

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.fill  = row_fill
        cell.font  = Font(name="Arial", size=9)
        cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[row].height = 90

    # Test Case ID — bold
    ws.cell(row=row, column=1).font = Font(name="Arial", size=9, bold=True)
    # Module — centred
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    # Status — centred
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    # Test Type — centred + bold
    ws.cell(row=row, column=9).font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    # Priority colour
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor=PRI_COLORS.get(priority, WHITE))
    ws.cell(row=row, column=10).font = Font(name="Arial", size=9, bold=True,
        color=("990000" if priority == "Critical" else "000000"))

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14, 14, 44, 36, 64, 46, 18, 12, 14, 10, 40]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C5"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(_tmp)

try:
    if os.path.exists(FILE):
        os.replace(_tmp, FILE)
    else:
        shutil.copy2(_tmp, FILE)
        os.remove(_tmp)
    _saved_to = FILE
except PermissionError:
    _saved_to = _tmp

# ── Summary ───────────────────────────────────────────────────────────────────
pos  = sum(1 for t in TC if t[8] == "Positive")
neg  = sum(1 for t in TC if t[8] == "Negative")
edge = sum(1 for t in TC if t[8] == "Edge Case")
last = TC[-1]
print(f"SUCCESS — '{SHEET}' sheet written to: {_saved_to}")
print(f"Total TCs : {len(TC)}")
print(f"Positive  : {pos}")
print(f"Negative  : {neg}")
print(f"Edge Case : {edge}")
print(f"Last TC   : {last[0]} | Type: {last[8]}")
print(f"Sheets    : {wb.sheetnames}")
