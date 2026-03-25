from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds.xlsx"
SHEET = "Login Page"

wb = load_workbook(FILE)

# Remove existing sheet if present to rebuild cleanly
if SHEET in wb.sheetnames:
    del wb[SHEET]

ws = wb.create_sheet(SHEET)

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
GOLD_LIGHT   = "FFF3CD"
POS_GREEN    = "E2EFDA"
NEG_RED      = "FCE4D6"
EDGE_YELLOW  = "FFF2CC"
WHITE        = "FFFFFF"
GREY_MID     = "D9D9D9"
GREY_LIGHT   = "F2F2F2"

def S(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

thin  = Border(left=S(), right=S(), top=S(), bottom=S())
thick = Border(left=S(), right=S(), top=S(), bottom=S("medium","1F3864"))

# ── Row 1 – Banner ────────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
c = ws.cell(row=1, column=1, value="SUNNY DIAMONDS — LOGIN MODULE QA TEST CASES")
c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=16)
c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Row 2 – Metadata ──────────────────────────────────────────────────────────
meta = [
    ("A2","Project URL:","B2","https://qa-sunnydiamonds.webc.in/login"),
    ("D2","Module:","E2","Login Page"),
    ("G2","Created By:","H2","Sreejith S Madavan"),
    ("J2","Created Date:","K2","25-Mar-2026"),
]
for lc,lv,vc,vv in meta:
    l = ws[lc]; l.value=lv
    l.font=Font(name="Arial",bold=True,color=DARK_BLUE,size=10)
    l.fill=PatternFill("solid",fgColor=GREY_MID)
    l.alignment=Alignment(horizontal="left",vertical="center")
    l.border=thin
    v = ws[vc]; v.value=vv
    v.font=Font(name="Arial",color=DARK_BLUE,size=10)
    v.fill=PatternFill("solid",fgColor=GREY_LIGHT)
    v.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    v.border=thin

ws.merge_cells("B2:C2"); ws.merge_cells("E2:F2")
ws.merge_cells("H2:I2"); ws.merge_cells("K2:K2")
ws.row_dimensions[2].height = 22

for col in range(1,12):
    ws.cell(row=3,column=col).fill = PatternFill("solid",fgColor=DARK_BLUE)
ws.row_dimensions[3].height = 6

# ── Row 4 – Column headers ─────────────────────────────────────────────────────
HEADERS = [
    "Test Case ID","Module Name","Test Case Description",
    "Preconditions","Test Steps","Expected Result",
    "Actual Result","Status","Test Type","Priority","Remarks"
]
for i,h in enumerate(HEADERS,1):
    c = ws.cell(row=4,column=i,value=h)
    c.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=MID_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=thick
ws.row_dimensions[4].height = 30

# ── Shared preconditions ──────────────────────────────────────────────────────
PRE = ("1. Application is accessible at https://qa-sunnydiamonds.webc.in/login\n"
       "2. User is on the Login page\n"
       "3. Application is up and running")

PRE_REG = ("1. Application is accessible at https://qa-sunnydiamonds.webc.in/login\n"
           "2. Account with email 'sreejith.s+4@webandcrafts.com' is registered\n"
           "3. User is on the Login page")

PRE_UNRG = ("1. Application is accessible at https://qa-sunnydiamonds.webc.in/login\n"
            "2. Email 'notregistered@test.com' does NOT exist in the system\n"
            "3. User is on the Login page")

LOGOUT = ("POST-TEST RESET:\n"
          "1. If logged in, click Profile icon\n"
          "2. Click 'Log Out' → Confirm 'LOG OUT'\n"
          "3. Navigate back to https://qa-sunnydiamonds.webc.in/login")

# ── Test Cases ────────────────────────────────────────────────────────────────
# (ID, Module, Description, Preconditions, Steps, Expected, ActualResult, Status, Type, Priority, Remarks)
TC = [

    # ════════════════════ POSITIVE ════════════════════════════════════════════
    (
        "TC_LOGIN_001","Login Page",
        "Verify Login page loads correctly with all UI elements visible",
        PRE,
        "1. Open browser\n"
        "2. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "3. Inspect the page for all expected UI elements",
        "Login page loads successfully\n"
        "Following elements are visible:\n"
        "- 'Welcome Back' heading\n"
        "- Email input field (*)\n"
        "- Password input field (*) with eye-toggle\n"
        "- 'Remember me' checkbox\n"
        "- 'Forgot password?' link\n"
        "- 'Sign In' button\n"
        "- 'Create Account' link\n"
        "- No console errors",
        "","Pass/Fail","Positive","High",
        "UI completeness check — all elements present"
    ),
    (
        "TC_LOGIN_002","Login Page",
        "Verify Password show/hide toggle (eye icon) works on Password field",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter any text in Password field: 'MyPassword'\n"
        "3. Observe password is masked (dots) by default\n"
        "4. Click the eye icon next to Password field\n"
        "5. Observe the password visibility\n"
        "6. Click the eye icon again",
        "Password is masked by default\n"
        "Password becomes visible as plain text on first eye icon click\n"
        "Password is masked again on second eye icon click",
        "","Pass/Fail","Positive","Medium",
        "UI toggle — password show/hide"
    ),
    (
        "TC_LOGIN_003","Login Page",
        "Verify 'Remember me' checkbox can be checked and unchecked",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Click the 'Remember me' checkbox\n"
        "3. Verify it is checked\n"
        "4. Click it again\n"
        "5. Verify it is unchecked",
        "Checkbox toggles correctly between checked and unchecked states\n"
        "Visual indicator (tick) appears when checked\n"
        "Tick disappears when unchecked",
        "","Pass/Fail","Positive","Medium",
        "UI checkbox toggle verification"
    ),
    (
        "TC_LOGIN_004","Login Page",
        "Verify 'Forgot password?' link navigates to Forgot Password page",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Click the 'Forgot password?' link",
        "User is redirected to https://qa-sunnydiamonds.webc.in/forgot-password\n"
        "Forgot Password page loads correctly\n"
        "No error is displayed",
        "","Pass/Fail","Positive","Medium",
        "Navigation — Forgot Password link"
    ),
    (
        "TC_LOGIN_005","Login Page",
        "Verify 'Create Account' link navigates to Registration page",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Click the 'Create Account' link at the bottom of the form",
        "User is redirected to https://qa-sunnydiamonds.webc.in/create\n"
        "Registration page loads correctly\n"
        "No error is displayed",
        "","Pass/Fail","Positive","Medium",
        "Navigation — Create Account link"
    ),

    # ════════════════════ NEGATIVE ════════════════════════════════════════════
    (
        "TC_LOGIN_006","Login Page",
        "Submit Login form with both Email and Password fields empty",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Leave Email field empty\n"
        "3. Leave Password field empty\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error messages displayed:\n"
        "- 'Email is required'\n"
        "- 'Password is required'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Empty form submission — both fields"
    ),
    (
        "TC_LOGIN_007","Login Page",
        "Submit Login form with Email field empty and valid Password entered",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Leave Email field empty\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Email is required'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Empty email — required field validation"
    ),
    (
        "TC_LOGIN_008","Login Page",
        "Submit Login form with valid Email entered and Password field empty",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Leave Password field empty\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Password is required'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Empty password — required field validation"
    ),
    (
        "TC_LOGIN_009","Login Page",
        "Submit Login with invalid email format — missing @ symbol",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'invalidemail.com' (no @ symbol)\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Please enter a valid email address'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Invalid email format — missing @"
    ),
    (
        "TC_LOGIN_010","Login Page",
        "Submit Login with invalid email format — missing domain extension",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'user@domain' (no .com/.in)\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Please enter a valid email address'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Invalid email format — missing domain extension"
    ),
    (
        "TC_LOGIN_011","Login Page",
        "Submit Login with a valid email format but unregistered email address",
        PRE_UNRG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'notregistered@test.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Login fails\n"
        "Error message displayed: 'Invalid email or password. Please try again.'\n"
        "User remains on the Login page\n"
        "No account details are exposed",
        "","Pass/Fail","Negative","High",
        "Unregistered email — authentication failure"
    ),
    (
        "TC_LOGIN_012","Login Page",
        "Submit Login with registered email but incorrect password",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'WrongPassword123'\n"
        "4. Click 'Sign In' button",
        "Login fails\n"
        "Error message displayed: 'Invalid email or password. Please try again.'\n"
        "User remains on the Login page\n"
        "Account is NOT locked after single failed attempt",
        "","Pass/Fail","Negative","High",
        "Wrong password — authentication failure"
    ),
    (
        "TC_LOGIN_013","Login Page",
        "Submit Login with correct email but password in wrong case (case sensitivity check)",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'password' (all lowercase — correct is 'Password')\n"
        "4. Click 'Sign In' button",
        "Login fails\n"
        "Error message: 'Invalid email or password. Please try again.'\n"
        "Confirms passwords are case-sensitive\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Password case sensitivity — 'password' vs 'Password'"
    ),
    (
        "TC_LOGIN_014","Login Page",
        "Submit Login with correct email but password entered with only spaces",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: '         ' (spaces only)\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Password cannot be empty or spaces only'\n"
        "OR Login fails with: 'Invalid email or password'\n"
        "User remains on the Login page",
        "","Pass/Fail","Negative","High",
        "Whitespace-only password field"
    ),

    # ════════════════════ EDGE CASES ══════════════════════════════════════════
    (
        "TC_LOGIN_015","Login Page",
        "SQL Injection attempt in Email field (Security)",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: \"' OR '1'='1\"\n"
        "3. Enter Password: 'anypassword'\n"
        "4. Click 'Sign In' button",
        "Login is NOT successful\n"
        "System shows standard authentication failure message\n"
        "No SQL error exposed in the response\n"
        "Database remains unaffected",
        "","Pass/Fail","Edge Case","Critical",
        "Security — SQL injection prevention (OWASP Top 10)"
    ),
    (
        "TC_LOGIN_016","Login Page",
        "XSS injection attempt in Email field (Security)",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: '<script>alert(\"XSS\")</script>'\n"
        "3. Enter Password: 'anypassword'\n"
        "4. Click 'Sign In' button",
        "No JavaScript alert executes\n"
        "Input is sanitised or rejected with a validation error\n"
        "System shows: 'Please enter a valid email address'\n"
        "No script is rendered in the DOM",
        "","Pass/Fail","Edge Case","Critical",
        "Security — XSS prevention (OWASP Top 10)"
    ),
    (
        "TC_LOGIN_017","Login Page",
        "Submit Login with email containing leading and trailing spaces",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: ' sreejith.s+4@webandcrafts.com ' (with spaces around)\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Option A: System trims whitespace and logs in successfully\n"
        "Option B: Validation error — 'Invalid email address'\n"
        "Behaviour must be consistent and predictable",
        "","Pass/Fail","Edge Case","Medium",
        "Email with leading/trailing spaces"
    ),
    (
        "TC_LOGIN_018","Login Page",
        "Submit Login with extremely long email address (>254 characters — RFC boundary)",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: a string of 260+ characters ending with '@test.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Form is NOT submitted\n"
        "Validation error: 'Invalid email address' or 'Email too long'\n"
        "OR field rejects input beyond maximum length\n"
        "No server crash or 500 error",
        "","Pass/Fail","Edge Case","High",
        "BVA — Email max length boundary (RFC 5321: 254 chars)"
    ),
    (
        "TC_LOGIN_019","Login Page",
        "Submit Login with an extremely long password (boundary test)",
        PRE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: a string of 500+ characters\n"
        "4. Click 'Sign In' button",
        "Login fails gracefully with standard error message\n"
        "No server error (500) or application crash\n"
        "No sensitive information exposed",
        "","Pass/Fail","Edge Case","High",
        "BVA — Extremely long password input"
    ),
    (
        "TC_LOGIN_020","Login Page",
        "Verify email field accepts case-insensitive email (uppercase vs lowercase)",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'SREEJITH.S+4@WEBANDCRAFTS.COM' (all uppercase)\n"
        "3. Enter Password: 'Password'\n"
        "4. Click 'Sign In' button",
        "Option A: Login succeeds (email is case-insensitive — best practice)\n"
        "Option B: Login fails with 'Invalid email or password'\n"
        "Behaviour must be consistent — should not differ from lowercase login",
        "","Pass/Fail","Edge Case","Medium",
        "Email case insensitivity — RFC 5321 compliance"
    ),
    (
        "TC_LOGIN_021","Login Page",
        "Verify Login page is served over HTTPS with secure connection",
        PRE,
        "1. Open browser\n"
        "2. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "3. Check address bar for HTTPS and padlock icon\n"
        "4. Open browser DevTools → Console\n"
        "5. Check for mixed-content warnings",
        "Page URL begins with 'https://'\n"
        "Padlock icon is visible in browser address bar\n"
        "No 'Not Secure' warning displayed\n"
        "No mixed-content errors in browser console",
        "","Pass/Fail","Edge Case","High",
        "Security — HTTPS enforcement on login page"
    ),
    (
        "TC_LOGIN_022","Login Page",
        "Verify double-clicking 'Sign In' does not trigger duplicate login requests",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Rapidly double-click the 'Sign In' button",
        "Only ONE login request is sent to the server\n"
        "Button is disabled or debounced after first click\n"
        "User is redirected to dashboard exactly once\n"
        "No duplicate session or API error occurs",
        "","Pass/Fail","Edge Case","High",
        "Duplicate submission prevention on rapid double-click"
    ),
    (
        "TC_LOGIN_023","Login Page",
        "Verify 'Remember Me' — session persists after browser tab is closed and reopened",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Check the 'Remember me' checkbox\n"
        "5. Click 'Sign In'\n"
        "6. Close the browser tab\n"
        "7. Reopen the browser and navigate to https://qa-sunnydiamonds.webc.in/\n"
        "---\n"+LOGOUT,
        "User is still logged in after reopening the browser tab\n"
        "Session is persistent due to 'Remember me' being selected\n"
        "User does not need to log in again",
        "","Pass/Fail","Edge Case","Medium",
        "Remember Me — persistent session after tab close"
    ),

    # ════════════════════ POSITIVE — VALID LOGIN (MUST BE LAST) ═══════════════
    (
        "TC_LOGIN_024","Login Page",
        "Successful login with valid registered email and correct password",
        PRE_REG,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "3. Enter Password: 'Password'\n"
        "4. Leave 'Remember me' unchecked\n"
        "5. Click 'Sign In' button\n"
        "---\n"+LOGOUT,
        "Login is successful\n"
        "User is redirected to the home/dashboard page\n"
        "Welcome message or user name is displayed\n"
        "User is NOT on the login page (URL changes)\n"
        "No error message is shown",
        "","Pass/Fail","Positive","High",
        "✅ VALID LOGIN — Happy path. PLACED LAST as required. "
        "Email: sreejith.s+4@webandcrafts.com | Password: Password"
    ),
]

# ── Write test cases ───────────────────────────────────────────────────────────
TYPE_COLORS  = {"Positive": POS_GREEN, "Negative": NEG_RED, "Edge Case": EDGE_YELLOW}
PRI_COLORS   = {"High":"FFD7D7","Critical":"FFB3B3","Medium":"FFF3CD","Low":"E8F5E9"}

for r_idx, tc in enumerate(TC):
    row = 5 + r_idx
    ws.row_dimensions[row].height = 130
    (tc_id,mod,desc,pre,steps,exp,act,status,ttype,pri,rem) = tc
    bg = TYPE_COLORS.get(ttype, WHITE)
    for c_idx,val in enumerate([tc_id,mod,desc,pre,steps,exp,act,status,ttype,pri,rem],1):
        c = ws.cell(row=row, column=c_idx, value=val)
        c.font      = Font(name="Arial",size=9,color="000000",bold=(c_idx==1))
        c.fill      = PatternFill("solid",fgColor=bg)
        c.alignment = Alignment(horizontal="left",vertical="top",wrap_text=True)
        c.border    = thin
    # Test Type col (9) — coloured text
    ws.cell(row=row,column=9).font = Font(name="Arial",size=9,bold=True,
        color=("1F6B2E" if ttype=="Positive" else
               "7B1E0C" if ttype=="Negative" else "7B5B00"))
    # Priority col (10)
    p_bg = PRI_COLORS.get(pri,"FFFFFF")
    ws.cell(row=row,column=10).fill=PatternFill("solid",fgColor=p_bg)
    ws.cell(row=row,column=10).font=Font(name="Arial",size=9,bold=True,
        color=("990000" if pri=="Critical" else "000000"))

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14, 14, 40, 34, 58, 42, 18, 12, 14, 10, 35]
for i,w in enumerate(COL_WIDTHS,1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "C5"

wb.save(FILE)
print(f"SUCCESS — '{SHEET}' sheet added to: {FILE}")
print(f"Total Test Cases: {len(TC)}")
pos  = sum(1 for t in TC if t[9]=="Positive")
neg  = sum(1 for t in TC if t[9]=="Negative")
edge = sum(1 for t in TC if t[9]=="Edge Case")
print(f"  Positive : {pos}")
print(f"  Negative : {neg}")
print(f"  Edge Case: {edge}")
print(f"Last TC    : {TC[-1][0]} — {TC[-1][2][:60]}")
