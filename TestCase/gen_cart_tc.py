"""
FRIDAY — QA Agent
Cart Page Test Case Generator
Target: SunnyDiamonds_v2.xlsx  |  Sheet: Cart Page
Total TCs: 42  (Positive: 26 | Negative: 8 | Edge Case: 8)
Last TC: TC_CART_042 — Valid Login (Positive / Critical)
"""
import shutil, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET = "Cart Page"

_tmp = FILE.replace(".xlsx", "_cart_tmp.xlsx")
shutil.copy2(FILE, _tmp)
wb = load_workbook(_tmp)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# ── Colours ─────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GOLD_LIGHT  = "FFF3CD"
POS_GREEN   = "E2EFDA"
NEG_RED     = "FCE4D6"
EDGE_YELLOW = "FFF2CC"
WHITE       = "FFFFFF"
PRI_COLORS  = {"Critical":"FF0000","High":"FFD700","Medium":"90EE90","Low":"ADD8E6"}

def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(row, col, val, bg=MID_BLUE, fg="FFFFFF", size=10, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=size, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin()
    return c

def data_cell(row, col, val, bg=WHITE, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold, color="000000")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    c.border    = thin()
    return c

# ── Row 1: Banner ────────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
c = ws.cell(row=1, column=1,
    value="SUNNY DIAMONDS — QA TEST CASES  |  Cart Page  |  https://qa-sunnydiamonds.webc.in/cart")
c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border    = thin()
ws.row_dimensions[1].height = 28

# ── Row 2: Metadata ──────────────────────────────────────────────────────────
META = [
    ("A","Created By:","B","Sreejith S Madavan"),
    ("C","Module:","D","Cart Page"),
    ("E","Created Date:","F","2026-03-25"),
    ("G","Environment:","H","QA"),
    ("I","Total TCs:","J","42"),
]
for lc,lv,vc,vv in META:
    c = ws[f"{lc}2"]; c.value = lv
    c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=MID_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin()
    c = ws[f"{vc}2"]; c.value = vv
    c.font=Font(name="Arial",size=9,bold=False,color="000000")
    c.fill=PatternFill("solid",fgColor=LIGHT_BLUE)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.border=thin()
ws.row_dimensions[2].height = 18

# ── Row 3: Legend ────────────────────────────────────────────────────────────
ws.merge_cells("A3:K3")
c = ws.cell(row=3, column=1,
    value=("LEGEND:  Green = Positive   Pink/Red = Negative   Yellow = Edge Case  |  "
           "NAVIGATION: TC_CART_001 = Full Setup (Login + Add Items + Navigate to Cart).  "
           "TC_CART_002 onwards = Navigate back to Cart page (https://qa-sunnydiamonds.webc.in/cart)"))
c.font=Font(name="Arial",size=8,color="000000")
c.fill=PatternFill("solid",fgColor=GOLD_LIGHT)
c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
c.border=thin()
ws.row_dimensions[3].height = 22

# ── Row 4: Common Preconditions ──────────────────────────────────────────────
ws.merge_cells("A4:K4")
c = ws.cell(row=4, column=1,
    value=("COMMON PRECONDITIONS (Apply to ALL Test Cases): "
           "1. Application is accessible at https://qa-sunnydiamonds.webc.in  "
           "2. Browser is open and running (Chrome / Firefox / Edge)  "
           "3. Cart Page URL: https://qa-sunnydiamonds.webc.in/cart  "
           "4. Valid Test Account -> Email: sreejith.s+4@webandcrafts.com | Password: Password"))
c.font=Font(name="Arial",size=8,bold=True,color="000000")
c.fill=PatternFill("solid",fgColor=POS_GREEN)
c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
c.border=thin()
ws.row_dimensions[4].height = 30

# ── Row 5: Column Headers ────────────────────────────────────────────────────
HEADERS = ["Test Case ID","Module Name","Test Case Description","Preconditions",
           "Test Steps","Expected Result","Actual Result","Status",
           "Test Type","Priority","Remarks"]
for col, hdr in enumerate(HEADERS, 1):
    hdr_cell(5, col, hdr)
ws.row_dimensions[5].height = 20

# ── Navigation helpers ───────────────────────────────────────────────────────
NAV_001 = (
    "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
    "2. Enter Email: sreejith.s+4@webandcrafts.com\n"
    "3. Enter Password: Password\n"
    "4. Click the Sign In button\n"
    "5. After login, click the Sunny Diamonds logo to go to Home Page\n"
    "6. Click ALL JEWELLERY from the navigation menu\n"
    "7. Click on any product (e.g., Aminah Diamond Ring)\n"
    "8. Click ADD TO CART button\n"
    "9. Click browser Back to return to PLP\n"
    "10. Select another product (e.g., 18K Rose Gold Eden Diamond Ring)\n"
    "11. Click ADD TO CART button\n"
    "12. Click the Cart icon (top-right header) to navigate to Cart page\n"
    "13. Verify the Cart page is displayed"
)
NAV_BACK = "1. Navigate back to the Cart page (https://qa-sunnydiamonds.webc.in/cart)\n"

# ── Precondition helpers ─────────────────────────────────────────────────────
PRE_CART  = "User is logged in; Cart page is open with at least 2 items added"
PRE_SETUP = "Valid account: sreejith.s+4@webandcrafts.com / Password"
PRE_EMPTY = "User is logged in; Cart is empty (all items removed)"
PRE_GUEST = "User is NOT logged in (guest mode; no active session)"
PRE_ONE   = "User is logged in; Cart has at least 1 item with qty = 1"
PRE_CREDS = (
    "User is NOT logged in\n"
    "Valid credentials:\n"
    "  Email   : sreejith.s+4@webandcrafts.com\n"
    "  Password: Password"
)

# ── Test Cases ───────────────────────────────────────────────────────────────
# Tuple: (ID, Module, Description, Preconditions, Steps, Expected, Actual, Status, Type, Priority, Remarks)
TC = [

    # ════════════════════════════════════════════════════════════ POSITIVE (25)
    (
        "TC_CART_001","Cart Page",
        "Verify Cart page loads successfully after login and adding products to cart",
        PRE_SETUP,
        NAV_001,
        "Cart page loads at https://qa-sunnydiamonds.webc.in/cart.\n"
        "Page heading 'My Cart' is visible.\n"
        "Both added products appear as cart items.\n"
        "URL contains /cart.",
        "","Pass/Fail","Positive","Critical",
        "SETUP TC — Establishes login session and cart state for all subsequent TCs."
    ),
    (
        "TC_CART_002","Cart Page",
        "Verify 'My Cart' page title is displayed correctly",
        PRE_CART,
        NAV_BACK +
        "2. Observe the H1 page heading at the top of the cart content area.",
        "Page heading 'My Cart' is displayed prominently.\n"
        "Font, size, and styling are consistent with the design.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_003","Cart Page",
        "Verify cart item count text displays correct number of items",
        PRE_CART,
        NAV_BACK +
        "2. Observe the item count text displayed below the 'My Cart' heading.",
        "Item count text reads 'X items in your cart' where X matches the actual number of items added.\n"
        "Count is accurate and updates dynamically.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_004","Cart Page",
        "Verify product name is displayed correctly for each cart item",
        PRE_CART,
        NAV_BACK +
        "2. Observe the product name heading for each item in the cart list.\n"
        "3. Cross-check names against the products added from PDP.",
        "Each cart item displays the correct product name.\n"
        "Names match exactly as shown on the respective PDP pages.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_005","Cart Page",
        "Verify product SKU is displayed for each cart item",
        PRE_CART,
        NAV_BACK +
        "2. Observe the SKU field displayed below the product name for each item.",
        "SKU label ('SKU:') and its value are visible for each cart item.\n"
        "SKU matches the product's SKU as shown on PDP.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_006","Cart Page",
        "Verify product colour (metal variant) is displayed for each cart item",
        PRE_CART,
        NAV_BACK +
        "2. Observe the 'Color:' field for each item in the cart.",
        "Color label and value (e.g., 'rose-gold') are visible for each item.\n"
        "Color matches the variant selected on PDP.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_007","Cart Page",
        "Verify product unit price is displayed correctly per cart item",
        PRE_CART,
        NAV_BACK +
        "2. Observe the price (with ₹ symbol) displayed for each cart item.\n"
        "3. Cross-check unit price against the price shown on PDP for the same product.",
        "Price is displayed with the ₹ symbol for each item.\n"
        "Unit price matches the PDP price for the same variant.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_008","Cart Page",
        "Verify product thumbnail image is displayed for each cart item",
        PRE_CART,
        NAV_BACK +
        "2. Observe the product image thumbnail on the left side of each cart item row.\n"
        "3. Check that no image placeholder or broken icon is shown.",
        "Product thumbnail image loads correctly for each cart item.\n"
        "Images are not broken or missing.\n"
        "Images correspond to the correct products.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_009","Cart Page",
        "Verify quantity increases by 1 when '+' button is clicked",
        PRE_CART,
        NAV_BACK +
        "2. Locate the quantity control (- [qty] +) on the first cart item.\n"
        "3. Note the current quantity (e.g., 1).\n"
        "4. Click the '+' button once.\n"
        "5. Observe the quantity input field.",
        "Quantity increases by 1 (e.g., 1 becomes 2).\n"
        "The quantity input field updates to the new value immediately.\n"
        "No page reload is required.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_010","Cart Page",
        "Verify Subtotal updates correctly after quantity is increased",
        PRE_CART,
        NAV_BACK +
        "2. Note the current Subtotal in the Price Details section.\n"
        "3. Note the unit price of Item 1.\n"
        "4. Click '+' on Item 1 to increase qty from 1 to 2.\n"
        "5. Observe the Subtotal in the Price Details section.",
        "Subtotal increases by the unit price of Item 1.\n"
        "New Subtotal = Old Subtotal + Item 1 unit price.\n"
        "Total also updates to match the new Subtotal.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_011","Cart Page",
        "Verify quantity decreases by 1 when '-' button is clicked",
        PRE_CART,
        NAV_BACK +
        "2. Ensure Item 1 has qty >= 2 (use '+' to increase if needed).\n"
        "3. Note the current quantity.\n"
        "4. Click the '-' button once.\n"
        "5. Observe the quantity input field.",
        "Quantity decreases by 1.\n"
        "The quantity input field updates immediately.\n"
        "No page reload is required.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_012","Cart Page",
        "Verify Subtotal updates correctly after quantity is decreased",
        PRE_CART,
        NAV_BACK +
        "2. Increase qty of Item 1 to 2 using '+' button.\n"
        "3. Note the current Subtotal.\n"
        "4. Click '-' on Item 1 to decrease qty to 1.\n"
        "5. Observe the Subtotal in Price Details.",
        "Subtotal decreases by the unit price of Item 1.\n"
        "New Subtotal = Old Subtotal - Item 1 unit price.\n"
        "Total updates to match the new Subtotal.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_013","Cart Page",
        "Verify Total equals Subtotal when no discount is applied",
        PRE_CART,
        NAV_BACK +
        "2. Observe both 'Subtotal' and 'Total' values in the Price Details panel.",
        "Subtotal amount equals Total amount.\n"
        "Both values are prefixed with the ₹ symbol.\n"
        "No discount or promo row is visible.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_014","Cart Page",
        "Verify a single item can be removed from the cart",
        PRE_CART,
        NAV_BACK +
        "2. Identify the remove (trash/X) icon on the first cart item.\n"
        "3. Click the remove icon.\n"
        "4. Observe the cart items list.",
        "The removed item disappears from the cart list immediately.\n"
        "Remaining items are still shown correctly.\n"
        "Item count decrements by 1.",
        "","Pass/Fail","Positive","Critical",""
    ),
    (
        "TC_CART_015","Cart Page",
        "Verify cart item count updates correctly after item removal",
        PRE_CART,
        NAV_BACK +
        "2. Note the current item count text (e.g., '2 items in your cart').\n"
        "3. Click the remove icon on any one cart item.\n"
        "4. Observe the item count text and header cart badge.",
        "Item count text updates immediately (e.g., '2 items' becomes '1 item').\n"
        "The cart icon badge in the header also decrements to match.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_016","Cart Page",
        "Verify empty cart state is shown after removing all items",
        PRE_CART,
        NAV_BACK +
        "2. Click the remove icon on Item 1.\n"
        "3. Click the remove icon on Item 2 (and any remaining items).\n"
        "4. Observe the cart page after all items are removed.",
        "Cart page shows an empty state message (e.g., 'Your cart is empty' or similar).\n"
        "No item rows are visible.\n"
        "Cart icon badge in the header shows 0 or is hidden.",
        "","Pass/Fail","Positive","Critical",""
    ),
    (
        "TC_CART_017","Cart Page",
        "Verify 'Continue Shopping' link navigates to the correct page",
        PRE_CART,
        NAV_BACK +
        "2. Locate the 'Continue Shopping' link at the top-left of the cart section.\n"
        "3. Click 'Continue Shopping'.",
        "Browser navigates to the Trending or Jewellery page.\n"
        "URL changes to /trending or /jewellery.\n"
        "Page loads without error.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_018","Cart Page",
        "Verify 'CHECKOUT SECURELY' button is visible and navigates to checkout",
        PRE_CART,
        NAV_BACK +
        "2. Scroll to the Price Details / Order Summary section.\n"
        "3. Verify the 'CHECKOUT SECURELY' button is visible.\n"
        "4. Click the 'CHECKOUT SECURELY' button.",
        "'CHECKOUT SECURELY' button is visible and properly styled.\n"
        "Clicking it navigates to the checkout or payment page.\n"
        "No error is thrown.",
        "","Pass/Fail","Positive","Critical",""
    ),
    (
        "TC_CART_019","Cart Page",
        "Verify cart icon badge in header shows the correct item count",
        PRE_CART,
        NAV_BACK +
        "2. Observe the cart icon in the top navigation header.\n"
        "3. Check the badge/counter displayed on the cart icon.",
        "Cart icon badge displays the correct number matching the items added.\n"
        "Badge count matches the 'X items in your cart' text on the page.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_020","Cart Page",
        "Verify Price Details section shows Subtotal and Total amounts",
        PRE_CART,
        NAV_BACK +
        "2. Locate the 'Price Details' section on the right-hand panel.\n"
        "3. Observe all rows in the price summary.",
        "'Price Details' heading is displayed.\n"
        "Subtotal row shows a ₹ amount greater than 0.\n"
        "Total row shows a ₹ amount equal to Subtotal (no discount).",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_021","Cart Page",
        "Verify 'Our Promise to You' section is visible on the Cart page",
        PRE_CART,
        NAV_BACK +
        "2. Scroll down past the cart items list.\n"
        "3. Observe the trust/promise section below the cart.",
        "'Our Promise to You' section is visible.\n"
        "Promise cards (e.g., Internally Flawless Diamonds, 100% Money Back, 15 Days Return) are shown.\n"
        "Icons and labels load correctly.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_022","Cart Page",
        "Verify mobile sticky checkout bar is present on Cart page",
        PRE_CART,
        NAV_BACK +
        "2. Resize the browser to mobile width (375px) using DevTools or device mode.\n"
        "3. Observe the bottom of the screen.",
        "A sticky 'Checkout Securely' bar is pinned at the bottom of the viewport on mobile.\n"
        "Bar is visible and clickable.\n"
        "Clicking it navigates to the checkout page.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_023","Cart Page",
        "Verify navigating to Cart page via the header cart icon",
        PRE_CART,
        NAV_BACK +
        "2. Navigate to the Home page or PLP.\n"
        "3. Locate the cart icon in the top-right of the navigation header.\n"
        "4. Click the cart icon.",
        "Browser navigates to https://qa-sunnydiamonds.webc.in/cart.\n"
        "Cart page loads with all previously added items intact.\n"
        "Item count and prices are unchanged.",
        "","Pass/Fail","Positive","Medium",""
    ),
    (
        "TC_CART_024","Cart Page",
        "Verify cart data persists after browser page refresh",
        PRE_CART,
        NAV_BACK +
        "2. Note all items in the cart (names, quantities, prices).\n"
        "3. Press F5 or click the browser Refresh button.\n"
        "4. Wait for the page to fully reload.",
        "All cart items are still present after the refresh.\n"
        "Quantities, prices, and item details are unchanged.\n"
        "Subtotal and Total remain the same.",
        "","Pass/Fail","Positive","High",""
    ),
    (
        "TC_CART_025","Cart Page",
        "Verify footer newsletter subscription works from Cart page",
        PRE_CART,
        NAV_BACK +
        "2. Scroll to the footer section of the Cart page.\n"
        "3. Locate the newsletter email field and 'Subscribe' button.\n"
        "4. Enter a valid email: testuser@example.com\n"
        "5. Click the 'Subscribe' button.",
        "Email field accepts the input.\n"
        "A success message or confirmation is shown after clicking Subscribe.\n"
        "No errors are thrown for a valid email.",
        "","Pass/Fail","Positive","Low",""
    ),

    # ════════════════════════════════════════════════════════════ NEGATIVE (8)
    (
        "TC_CART_026","Cart Page",
        "Verify behaviour when '-' button is clicked with quantity already at 1",
        PRE_ONE,
        NAV_BACK +
        "2. Ensure the cart item has quantity = 1.\n"
        "3. Click the '-' (decrease) button on that item.",
        "System does NOT allow quantity to go below 1.\n"
        "Either:\n"
        "  a) Quantity stays at 1 and the '-' button is disabled/greyed out, OR\n"
        "  b) A confirmation dialog appears before removing the item.\n"
        "Cart must NOT show a negative quantity value.",
        "","Pass/Fail","Negative","High",
        "BVA lower boundary — qty = 1 is the minimum."
    ),
    (
        "TC_CART_027","Cart Page",
        "Verify behaviour when quantity input field is manually set to 0",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click inside the quantity input field of any cart item.\n"
        "3. Clear the field and type '0'.\n"
        "4. Press Tab or Enter to confirm.",
        "System does not accept 0 as a valid quantity.\n"
        "Either:\n"
        "  a) Quantity reverts to 1 (minimum allowed), OR\n"
        "  b) Item is removed with a confirmation prompt.\n"
        "Subtotal does not display ₹0 for a product.",
        "","Pass/Fail","Negative","High",""
    ),
    (
        "TC_CART_028","Cart Page",
        "Verify behaviour when a negative value is entered in the quantity field",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click inside the quantity input field.\n"
        "3. Clear the field and type '-1'.\n"
        "4. Press Tab or Enter.",
        "Negative value is rejected.\n"
        "Quantity reverts to the last valid value (e.g., 1).\n"
        "Subtotal remains positive.\n"
        "An appropriate error or warning message may appear.",
        "","Pass/Fail","Negative","High",""
    ),
    (
        "TC_CART_029","Cart Page",
        "Verify behaviour when alphabetic characters are entered in the quantity field",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click inside the quantity input field.\n"
        "3. Clear the field and type 'abc'.\n"
        "4. Press Tab or Enter.",
        "Alphabetic input is not accepted by the quantity field.\n"
        "Field does not display letters.\n"
        "Quantity reverts to the last valid numeric value.\n"
        "No JS console errors or page crashes.",
        "","Pass/Fail","Negative","Medium",""
    ),
    (
        "TC_CART_030","Cart Page",
        "Verify behaviour when special characters are entered in the quantity field",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click inside the quantity input field.\n"
        "3. Clear and type '@#$%'.\n"
        "4. Press Tab or Enter.",
        "Special characters are rejected by the quantity field.\n"
        "Quantity reverts to the last valid numeric value.\n"
        "No XSS execution or unhandled errors occur.",
        "","Pass/Fail","Negative","Medium",""
    ),
    (
        "TC_CART_031","Cart Page",
        "Verify Cart page behaviour when accessed directly by a guest (unauthenticated) user",
        PRE_GUEST,
        "1. Open browser in private/incognito mode (no active session).\n"
        "2. Navigate directly to https://qa-sunnydiamonds.webc.in/cart.",
        "Guest user is:\n"
        "  a) Redirected to the Login page, OR\n"
        "  b) Shown an empty cart with a prompt to log in.\n"
        "Cart items from another user's session are NOT exposed.\n"
        "No unauthorised data is displayed.",
        "","Pass/Fail","Negative","Critical",
        "Security: direct URL access without authentication."
    ),
    (
        "TC_CART_032","Cart Page",
        "Verify newsletter subscription with invalid email format in footer",
        "User is logged in; Cart page is open",
        NAV_BACK +
        "2. Scroll to the footer newsletter section.\n"
        "3. Enter an invalid email: 'notanemail' or 'test@'\n"
        "4. Click the 'Subscribe' button.",
        "Validation error is shown for the invalid email format.\n"
        "Subscription is NOT submitted.\n"
        "Error message clearly indicates the issue.",
        "","Pass/Fail","Negative","Medium",""
    ),
    (
        "TC_CART_033","Cart Page",
        "Verify 'CHECKOUT SECURELY' behaviour when cart is empty",
        PRE_EMPTY,
        "1. Navigate to the Cart page (all items already removed).\n"
        "2. Observe the page state and checkout button visibility.",
        "When cart is empty:\n"
        "  a) Checkout button is hidden or disabled, OR\n"
        "  b) Clicking it shows 'Your cart is empty. Add items to proceed.'\n"
        "User is NOT navigated to the payment step with an empty cart.",
        "","Pass/Fail","Negative","High",
        "Empty cart checkout prevention."
    ),

    # ════════════════════════════════════════════════════════════ EDGE CASE (8)
    (
        "TC_CART_034","Cart Page",
        "BVA: Verify quantity = 1 is accepted as the minimum valid quantity",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Ensure at least one item has quantity = 1.\n"
        "3. Verify the quantity input shows '1'.\n"
        "4. Verify no error or warning is displayed.",
        "Quantity 1 is accepted as the minimum valid value.\n"
        "Subtotal reflects the single-unit price.\n"
        "No error or warning is shown.",
        "","Pass/Fail","Edge Case","High",
        "BVA lower boundary — minimum valid quantity."
    ),
    (
        "TC_CART_035","Cart Page",
        "BVA: Verify behaviour when a very large quantity (999) is entered",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click the quantity input field of a cart item.\n"
        "3. Clear the field and type '999'.\n"
        "4. Press Tab or Enter.\n"
        "5. Observe the updated cart and price.",
        "System either:\n"
        "  a) Accepts 999 and updates Subtotal correctly (999 x unit price), OR\n"
        "  b) Shows a maximum quantity limit error message.\n"
        "No crash, JS exception, or incorrect price calculation occurs.",
        "","Pass/Fail","Edge Case","Medium",
        "BVA upper boundary — large quantity stress test."
    ),
    (
        "TC_CART_036","Cart Page",
        "Verify price calculation accuracy for multiple items with different unit prices",
        PRE_CART,
        NAV_BACK +
        "2. Note the unit price and quantity for each item in the cart.\n"
        "3. Manually calculate: Expected Subtotal = sum of (unit price x qty) for all items.\n"
        "4. Compare the manually calculated total against the displayed Subtotal.",
        "Displayed Subtotal = sum of (each item's unit price x its quantity).\n"
        "Total equals Subtotal (no discount applied).\n"
        "No rounding error greater than ₹1.",
        "","Pass/Fail","Edge Case","High",
        "Price accuracy and arithmetic validation."
    ),
    (
        "TC_CART_037","Cart Page",
        "Verify Cart page renders correctly on mobile viewport (375 x 812 px)",
        PRE_CART,
        NAV_BACK +
        "2. Open browser DevTools and set viewport to 375 x 812 px (iPhone SE/12).\n"
        "3. Refresh the Cart page.\n"
        "4. Scroll through the entire page and check layout.",
        "Cart items display in a single-column mobile layout.\n"
        "Sticky checkout bar is visible at the bottom.\n"
        "Quantity controls are usable at touch-target size.\n"
        "No horizontal overflow or broken layout.",
        "","Pass/Fail","Edge Case","Medium",
        "Mobile responsiveness validation."
    ),
    (
        "TC_CART_038","Cart Page",
        "Verify Cart page renders correctly on tablet viewport (768 x 1024 px)",
        PRE_CART,
        NAV_BACK +
        "2. Open browser DevTools and set viewport to 768 x 1024 px (iPad).\n"
        "3. Refresh the Cart page.\n"
        "4. Observe layout, alignment, and element visibility.",
        "Cart renders correctly at tablet viewport.\n"
        "No overlapping UI elements.\n"
        "Price Details panel is accessible.\n"
        "All interactive elements (buttons, inputs) are usable.",
        "","Pass/Fail","Edge Case","Medium",
        "Tablet responsiveness validation."
    ),
    (
        "TC_CART_039","Cart Page",
        "Verify XSS injection in quantity field is blocked and sanitised",
        "User is logged in; Cart has at least 1 item",
        NAV_BACK +
        "2. Click the quantity input field of any cart item.\n"
        "3. Clear and enter: <script>alert('XSS')</script>\n"
        "4. Press Tab or Enter.\n"
        "5. Observe page behaviour and check for any alert dialog.",
        "No JavaScript alert popup appears.\n"
        "The injected script tag is rejected or sanitised.\n"
        "Quantity field reverts to the last valid numeric value.\n"
        "No XSS vulnerability is exploitable.",
        "","Pass/Fail","Edge Case","Critical",
        "OWASP Top 10 — XSS input sanitisation."
    ),
    (
        "TC_CART_040","Cart Page",
        "Verify cart icon badge resets after all items are removed",
        PRE_CART,
        NAV_BACK +
        "2. Note the cart badge count in the header (e.g., 2).\n"
        "3. Remove all items from the cart one by one.\n"
        "4. Observe the cart icon badge in the top navigation.",
        "Cart icon badge updates to 0 or disappears entirely after all items are removed.\n"
        "Badge does NOT continue to show the old count.",
        "","Pass/Fail","Edge Case","High",
        "State synchronisation — cart badge after empty."
    ),
    (
        "TC_CART_041","Cart Page",
        "Verify browser Back navigation from Checkout returns to Cart with state intact",
        PRE_CART,
        NAV_BACK +
        "2. Click the 'CHECKOUT SECURELY' button.\n"
        "3. Wait for the checkout/payment page to load.\n"
        "4. Click the browser Back button.",
        "Browser navigates back to the Cart page.\n"
        "All cart items are still present after back navigation.\n"
        "Quantities and prices are unchanged.\n"
        "Cart is fully functional (can update qty, remove items, or checkout again).",
        "","Pass/Fail","Edge Case","Medium",
        "Browser history and cart state continuity."
    ),

    # ════════════════════════════════════════════════════════════ VALID LOGIN — LAST
    (
        "TC_CART_042","Cart Page",
        "Login with valid credentials and verify Cart page is fully accessible",
        PRE_CREDS,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
        "2. Enter Email: sreejith.s+4@webandcrafts.com\n"
        "3. Enter Password: Password\n"
        "4. Click Sign In\n"
        "5. After successful login, navigate to https://qa-sunnydiamonds.webc.in/cart\n"
        "6. Observe page access and content.",
        "Login succeeds without error.\n"
        "User is redirected to the home/dashboard page.\n"
        "Navigating to /cart loads the Cart page without redirect to Login.\n"
        "Cart badge, page title 'My Cart', and all UI elements are visible and functional.",
        "","Pass/Fail","Positive","Critical",
        "VALID LOGIN — Placed LAST as required. Verifies authenticated access to Cart page."
    ),
]

# ── Write rows ────────────────────────────────────────────────────────────────
TYPE_BG = {"Positive": POS_GREEN, "Negative": NEG_RED, "Edge Case": EDGE_YELLOW}

for idx, row_data in enumerate(TC):
    row  = 6 + idx
    tc_id, module, desc, pre, steps, exp, actual, status, ttype, priority, remarks = row_data
    bg   = TYPE_BG.get(ttype, WHITE)

    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal="center" if col in (1,2,8,9,10) else "left")
        c.border    = thin()
        # Font
        if col == 1:  # TC ID — bold
            c.font = Font(name="Arial", size=9, bold=True, color="000000")
        elif col == 10:  # Priority — coloured
            c.font = Font(name="Arial", size=9, bold=True,
                          color=("990000" if priority == "Critical" else "000000"))
            c.fill = PatternFill("solid", fgColor=PRI_COLORS.get(priority, WHITE))
        else:
            c.font = Font(name="Arial", size=9, bold=False, color="000000")

    ws.row_dimensions[row].height = 90

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14, 12, 44, 34, 62, 48, 18, 12, 14, 10, 40]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C6"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(_tmp)

try:
    if os.path.exists(FILE):
        os.replace(_tmp, FILE)
    else:
        import shutil as _sh; _sh.copy2(_tmp, FILE); os.remove(_tmp)
    _saved = FILE
except PermissionError:
    _saved = _tmp

pos  = sum(1 for t in TC if t[8] == "Positive")
neg  = sum(1 for t in TC if t[8] == "Negative")
edge = sum(1 for t in TC if t[8] == "Edge Case")
print(f"SUCCESS  ->  '{SHEET}' sheet written to: {_saved}")
print(f"Total TCs : {len(TC)}")
print(f"Positive  : {pos}")
print(f"Negative  : {neg}")
print(f"Edge Case : {edge}")
print(f"Last TC   : {TC[-1][0]} | {TC[-1][8]} | {TC[-1][9]}")
