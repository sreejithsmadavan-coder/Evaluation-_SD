import shutil, os, sys, io, re, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_guest.xlsx"

GREEN  = "E2EFDA"
RED    = "FCE4D6"
YELLOW = "FFF2CC"

thin   = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border    = border
    return c

def bg_color(tc_type):
    return GREEN if tc_type == "Positive" else (RED if tc_type == "Negative" else YELLOW)

LOGIN_URL    = "https://qa-sunnydiamonds.webc.in/login"
CHECKOUT_URL = "https://qa-sunnydiamonds.webc.in/checkout"
PLP_URL      = "https://qa-sunnydiamonds.webc.in/jewellery"
NAV          = f"Navigate back to {CHECKOUT_URL}"

# ─────────────────────────────────────────────────────────────────────────────
# Guest Checkout Flow (observed in video 0s–59s):
#   1. User logged out → browses PLP → adds Aminah Diamond Ring (₹30,094) to cart
#   2. Clicks CHECKOUT → redirected to Login page
#   3. Login page shows: Email, Password, Sign In, Sign up, OR "CONTINUE AS GUEST"
#   4. Clicks "CONTINUE AS GUEST" → Checkout page loads (no login required)
#   5. Fills Shipping Address form
#   6. Scrolls to Payment Method → COD + Pay Online both available (₹30,094 < ₹49K)
#   7. Selects COD → button shows "PLACE ORDER"
#   8. Clicks PLACE ORDER → "Processing your order..." modal
#   9. OTP modal: "Verify Your Mobile Number" → enter OTP → CONFIRM AND PLACE ORDER
# ─────────────────────────────────────────────────────────────────────────────

GUEST_TCS = [

    # ════════════════════════════════════════════════════════════════════
    # POSITIVE TEST CASES
    # ════════════════════════════════════════════════════════════════════

    (
        "Guest001",
        "Checkout - Guest Checkout",
        "[Positive] Guest user clicks CHECKOUT — verify redirect to Login page with 'CONTINUE AS GUEST' button",
        "User is NOT logged in (guest); at least 1 product is in the cart",
        f"1. Ensure user is logged out from {LOGIN_URL}\n"
        f"2. Navigate to {PLP_URL}\n"
        f"3. Add any product to cart (e.g. Aminah Diamond Ring ₹30,094)\n"
        f"4. Navigate to Cart page → click 'CHECKOUT SECURELY'\n"
        f"5. Observe the page the user lands on",
        f"User is redirected to the Login page ({LOGIN_URL}); "
        f"Login page displays: 'WELCOME BACK' heading, Email field, Password field, SIGN IN button, "
        f"'Don't have an account? Sign up' link, 'or' separator, and "
        f"'CONTINUE AS GUEST' button at the bottom",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (28s–frame14): Guest user clicking checkout is redirected to login page with 'CONTINUE AS GUEST' option. Core guest entry point."
    ),

    (
        "Guest002",
        "Checkout - Guest Checkout",
        "[Positive] Click 'CONTINUE AS GUEST' on Login page — verify navigation to Checkout page without login",
        "User is NOT logged in; redirected to Login page from cart/checkout",
        f"1. As a guest user, add a product to cart\n"
        f"2. Click CHECKOUT SECURELY → redirected to Login page\n"
        f"3. Click the 'CONTINUE AS GUEST' button\n"
        f"4. Observe page navigation",
        f"User is navigated to the Checkout page ({CHECKOUT_URL}) without requiring email/password; "
        f"Checkout page loads fully with Shipping Address form and Order Summary; "
        f"header does NOT show a logged-in username; URL is /checkout",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (30s–frame15): 'CONTINUE AS GUEST' click → checkout page loads. No credentials required for guest checkout."
    ),

    (
        "Guest003",
        "Checkout - Guest Checkout",
        "[Positive] Guest user can view complete Checkout page — Shipping Address, Order Summary, Coupon, Gift Card, Payment Method all displayed",
        "Guest user has clicked 'CONTINUE AS GUEST' and is on Checkout page",
        f"1. As guest, navigate to Checkout via 'CONTINUE AS GUEST'\n"
        f"2. Verify all sections present on the Checkout page",
        "Checkout page displays all sections for guest:\n"
        "• Shipping Address form with all 9 fields (First Name*, Last Name*, Email*, Phone*, Address*, Pin Code*, City*, State*, Country*)\n"
        "• 'Use this address as my billing address' checkbox\n"
        "• Coupon Code field + APPLY button\n"
        "• Gift Card Redeem field + APPLY button\n"
        "• Payment Method section (COD + Pay Online)\n"
        "• Order Summary with product details and Total\n"
        "• PAY NOW / PLACE ORDER button",
        "", "Not Tested", "Positive", "High",
        "Observed in video (32s–frame16): Full checkout page visible to guest user. All sections must render correctly without login."
    ),

    (
        "Guest004",
        "Checkout - Guest Checkout",
        "[Positive] Guest fills all mandatory Shipping Address fields with valid data — verify form accepts input",
        "Guest user is on Checkout page via 'CONTINUE AS GUEST'",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter valid First Name: 'John'\n"
        f"3. Enter valid Last Name: 'Doe'\n"
        f"4. Enter valid Email: 'johndoe@gmail.com'\n"
        f"5. Enter valid Phone: '9876543210'\n"
        f"6. Enter valid Address: '123 Main Street'\n"
        f"7. Enter valid Pin Code: '682001'\n"
        f"8. Enter valid City: 'Kochi'\n"
        f"9. Select State: 'Kerala'\n"
        f"10. Verify Country defaults to 'India'\n"
        f"11. Observe all fields accept input without errors",
        "All 9 mandatory fields accept valid input; no validation errors shown; form is ready for submission",
        "", "Not Tested", "Positive", "Critical",
        "Guest users must be able to fill the shipping address form. Email entered here is used for order confirmation — not tied to any account."
    ),

    (
        "Guest005",
        "Checkout - Guest Checkout",
        "[Positive] Guest selects 'Cash on Delivery' (cart < ₹49,000) — verify button changes to 'PLACE ORDER'",
        "Guest user on Checkout page; cart total < ₹49,000 (e.g. Aminah Diamond Ring ₹30,094)",
        f"1. As guest, navigate to Checkout (cart total < ₹49,000)\n"
        f"2. Fill all mandatory Shipping Address fields with valid data\n"
        f"3. Scroll to Payment Method section\n"
        f"4. Select 'Cash on Delivery' radio button\n"
        f"5. Observe submit button label",
        "Both 'Cash on Delivery' and 'Pay Online' options are visible; "
        "selecting COD changes submit button label to 'PLACE ORDER'; COD is available for guest when cart < ₹49,000",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (40s–frame20): Guest user selects COD → button shows 'PLACE ORDER'. COD availability rule (< ₹49,000) applies to guest checkout too."
    ),

    (
        "Guest006",
        "Checkout - Guest Checkout",
        "[Positive] Guest clicks 'PLACE ORDER' (COD) — verify 'Processing your order' modal displayed",
        "Guest user on Checkout page; all fields filled; COD selected",
        f"1. As guest, complete all mandatory fields\n"
        f"2. Select Cash on Delivery\n"
        f"3. Click 'PLACE ORDER' button\n"
        f"4. Observe immediate response",
        "'Please wait as we are processing your order, this will only take a while.' modal displayed with loading spinner; PLACE ORDER button is disabled during processing",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (46s–frame23): COD order processing modal shown for guest. Same processing flow as logged-in user."
    ),

    (
        "Guest007",
        "Checkout - Guest Checkout",
        "[Positive] Guest COD order — verify 'Verify Your Mobile Number' OTP modal appears",
        "Guest user on Checkout; PLACE ORDER clicked; processing modal complete",
        f"1. As guest, fill all fields; select COD; click PLACE ORDER\n"
        f"2. Wait for processing modal\n"
        f"3. Observe OTP verification modal",
        "'Verify Your Mobile Number' OTP modal appears; modal shows masked phone number from the Phone field entered by guest; 4-digit OTP input boxes shown; 'Resend a new OTP in X seconds' timer; 'CONFIRM AND PLACE ORDER' button",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (48s–frame24): OTP modal shown for guest user. OTP sent to phone number entered in shipping address form."
    ),

    (
        "Guest008",
        "Checkout - Guest Checkout",
        "[Positive] Guest enters valid OTP — verify order placed successfully and 'Thank You' page shown",
        "Guest user on OTP verification modal; valid OTP received on phone",
        f"1. As guest, complete checkout up to OTP modal\n"
        f"2. Enter valid 4-digit OTP received on phone\n"
        f"3. Click 'CONFIRM AND PLACE ORDER'\n"
        f"4. Observe result",
        "Order confirmed; 'Thank You For The Purchase!' success page displayed; 'Confirmation Mail Has Been Sent To Your Mail Id' message shown; 'CONTINUE SHOPPING' button visible",
        "", "Not Tested", "Positive", "Critical",
        "End-to-end guest COD checkout happy path. Confirmation email sent to guest email address entered in shipping form."
    ),

    (
        "Guest009",
        "Checkout - Guest Checkout",
        "[Positive] Guest selects 'Pay Online' (Razorpay) — verify PAY NOW button and Razorpay gateway loads",
        "Guest user on Checkout page; all mandatory fields filled",
        f"1. As guest, navigate to Checkout\n"
        f"2. Fill all mandatory fields\n"
        f"3. Scroll to Payment Method\n"
        f"4. Select 'Pay Online'\n"
        f"5. Click 'PAY NOW'\n"
        f"6. Observe Razorpay gateway behaviour",
        "'Pay Online' is selectable for guest; button label shows 'PAY NOW'; clicking PAY NOW opens Razorpay payment gateway popup; guest can complete online payment",
        "", "Not Tested", "Positive", "High",
        "Guest online payment flow. Razorpay must work for unauthenticated guest users."
    ),

    (
        "Guest010",
        "Checkout - Guest Checkout",
        "[Positive] Login page shows 'Sign up' link during guest checkout — verify navigation to Registration page",
        "Guest user redirected to Login page from Checkout",
        f"1. As guest, add product to cart → click CHECKOUT → land on Login page\n"
        f"2. Click 'Don't have an account? Sign up' link\n"
        f"3. Observe navigation",
        "User is navigated to the Registration/Create Account page; guest can choose to register before completing checkout",
        "", "Not Tested", "Positive", "Medium",
        "Observed in video (frame14): Login page shows 'Don't have an account? Sign up' link. Alternative flow — guest may choose to register instead."
    ),

    # ════════════════════════════════════════════════════════════════════
    # NEGATIVE TEST CASES
    # ════════════════════════════════════════════════════════════════════

    (
        "Guest011",
        "Checkout - Guest Checkout",
        "[Negative] Guest accesses /checkout URL directly without adding items to cart — verify redirect to Login page",
        "User is NOT logged in; cart is empty",
        f"1. Ensure user is logged out and cart is empty\n"
        f"2. Directly navigate to {CHECKOUT_URL} in browser address bar\n"
        f"3. Observe the page response",
        f"User is redirected to Login page ({LOGIN_URL}); checkout page does not load for unauthenticated guest with empty cart; 'CONTINUE AS GUEST' option shown on login page",
        "", "Not Tested", "Negative", "High",
        "Direct URL access to /checkout without cart items must redirect to login. Prevents empty checkout sessions."
    ),

    (
        "Guest012",
        "Checkout - Guest Checkout",
        "[Negative] Guest submits Checkout form with all mandatory fields empty — verify validation errors shown",
        "Guest user on Checkout page via 'CONTINUE AS GUEST'; no fields filled",
        f"1. As guest, navigate to Checkout via CONTINUE AS GUEST\n"
        f"2. Leave all Shipping Address fields empty\n"
        f"3. Select a payment method\n"
        f"4. Click PAY NOW / PLACE ORDER\n"
        f"5. Observe validation errors",
        "Validation errors appear near all mandatory fields: 'First Name is required', 'Last Name is required', 'Email is required', 'Phone is required', 'Address is required', 'Pin Code is required', 'City is required', 'State is required'; form does not submit",
        "", "Not Tested", "Negative", "Critical",
        "Mapped from checklist Mandatory Fields §1. Guest checkout must enforce same validations as logged-in checkout."
    ),

    (
        "Guest013",
        "Checkout - Guest Checkout",
        "[Negative] Guest enters invalid email format in Email field — verify error message shown",
        "Guest user on Checkout page",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter 'invalidemail' (no @ symbol) in Email field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error message displayed near Email field: 'Invalid email address' or 'Please enter a valid email'; form does not submit with invalid email",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Email §2. Guest email must be valid — it is used for order confirmation."
    ),

    (
        "Guest014",
        "Checkout - Guest Checkout",
        "[Negative] Guest enters invalid phone number (less than 10 digits) — verify error shown",
        "Guest user on Checkout page",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter '12345' (5 digits) in Phone Number field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error displayed near Phone field: 'Please enter a valid 10-digit phone number'; form does not submit",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Phone §11. OTP is sent to phone — invalid phone blocks order placement."
    ),

    (
        "Guest015",
        "Checkout - Guest Checkout",
        "[Negative] Guest enters invalid Pin Code (alphabets) — verify error message shown",
        "Guest user on Checkout page",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter 'ABCDEF' in Pin Code field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error displayed near Pin Code field: 'Please enter a valid 6-digit PIN code'; alphabetic input rejected",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Postal Code §12. Guest PIN code validation must match logged-in user validation."
    ),

    (
        "Guest016",
        "Checkout - Guest Checkout",
        "[Negative] Guest enters invalid OTP in OTP verification modal — verify error and order NOT placed",
        "Guest user on OTP modal after clicking PLACE ORDER (COD)",
        f"1. As guest, complete all checkout fields; select COD; click PLACE ORDER\n"
        f"2. Wait for OTP modal\n"
        f"3. Enter an incorrect OTP (e.g., '0000')\n"
        f"4. Click 'CONFIRM AND PLACE ORDER'\n"
        f"5. Observe error handling",
        "Error message displayed in OTP modal: 'Invalid OTP. Please try again.'; order is NOT placed; OTP fields reset; modal remains open for retry",
        "", "Not Tested", "Negative", "Critical",
        "Guest OTP security — invalid OTP must block order completion. Same validation as logged-in user OTP flow."
    ),

    (
        "Guest017",
        "Checkout - Guest Checkout",
        "[Negative] Guest leaves OTP fields empty and clicks 'CONFIRM AND PLACE ORDER' — verify validation error",
        "Guest user on OTP verification modal",
        f"1. As guest, reach OTP modal\n"
        f"2. Leave all 4 OTP input boxes empty\n"
        f"3. Click 'CONFIRM AND PLACE ORDER'\n"
        f"4. Observe validation",
        "Validation error shown: 'Please enter OTP' or OTP field highlighted in red; order is NOT placed; button does not proceed with empty OTP",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist mandatory field §1. Empty OTP must not allow order placement for guest."
    ),

    (
        "Guest018",
        "Checkout - Guest Checkout",
        "[Negative] Guest checkout — cart total above ₹49,000 — verify COD NOT available (Pay Online only)",
        "Guest user; cart total > ₹49,000",
        f"1. As guest, add product(s) to cart with total > ₹49,000\n"
        f"2. Click CHECKOUT → CONTINUE AS GUEST\n"
        f"3. Scroll to Payment Method section\n"
        f"4. Observe available payment options",
        "Only 'Pay Online' is shown in Payment Method section; 'Cash on Delivery' is hidden/absent; button shows 'PAY NOW'; COD restriction applies to guest checkout same as logged-in checkout",
        "", "Not Tested", "Negative", "Critical",
        "Business rule (₹49,000 COD threshold) must apply equally to guest users. Mapped from SD_CHECKOUT_004."
    ),

    (
        "Guest019",
        "Checkout - Guest Checkout",
        "[Negative] Guest enters First Name with special characters (e.g., 'John@#') — verify rejection",
        "Guest user on Checkout page",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter 'John@#' in First Name field\n"
        f"3. Tab out\n"
        f"4. Observe validation",
        "First Name field rejects special characters; error message displayed: 'First Name should contain alphabets only'; applies same alpha-only rule as logged-in checkout",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist Contact form §2. Alpha-only name validation applies equally to guest checkout form."
    ),

    # ════════════════════════════════════════════════════════════════════
    # EDGE CASE TEST CASES
    # ════════════════════════════════════════════════════════════════════

    (
        "Guest020",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest enters email of existing registered account — verify system behaviour",
        "Guest user on Checkout page; an account exists with email 'sreejith.s+4@webandcrafts.com'",
        f"1. As guest, navigate to Checkout\n"
        f"2. Enter email 'sreejith.s+4@webandcrafts.com' (existing registered account) in Email field\n"
        f"3. Fill remaining fields and attempt to place order\n"
        f"4. Observe system response",
        "System either: (a) allows guest order with existing email and sends confirmation to that email, OR (b) prompts user to sign in to existing account; system must NOT silently merge orders without user consent; clear message displayed about the email conflict",
        "", "Not Tested", "Edge Case", "High",
        "Edge case: guest using a registered email. System must handle gracefully — either allow as guest order or prompt to login. No silent account merge."
    ),

    (
        "Guest021",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest refreshes Checkout page after filling address fields — verify form data persistence",
        "Guest user on Checkout page with partially filled address form",
        f"1. As guest, navigate to Checkout\n"
        f"2. Fill First Name, Last Name, Email, Phone fields\n"
        f"3. Press F5 / browser refresh\n"
        f"4. Observe form state after reload",
        "After refresh: either form data is preserved (browser autofill or session storage) OR user is prompted that data will be lost; form does not throw errors; cart and Order Summary still intact",
        "", "Not Tested", "Edge Case", "Medium",
        "Guest session state after refresh. No account session means browser storage or session token must retain form state for good UX."
    ),

    (
        "Guest022",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest clicks browser back button from Checkout to Login page — verify 'CONTINUE AS GUEST' still available",
        "Guest user navigated to Checkout via 'CONTINUE AS GUEST'; now on Checkout page",
        f"1. As guest, proceed to Checkout via CONTINUE AS GUEST\n"
        f"2. On Checkout page, click browser Back button\n"
        f"3. Observe the page navigated to\n"
        f"4. Check if 'CONTINUE AS GUEST' button is still available",
        "Browser back navigates to Login page; 'CONTINUE AS GUEST' button is still present and functional; guest can click it again to return to Checkout; cart items are preserved",
        "", "Not Tested", "Edge Case", "Medium",
        "Browser back navigation from guest checkout. Login page must remain accessible with CONTINUE AS GUEST option intact."
    ),

    (
        "Guest023",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest removes last item from Order Summary on Checkout — verify redirect to Cart (same as logged-in user)",
        "Guest user on Checkout page with 1 item",
        f"1. As guest, navigate to Checkout via CONTINUE AS GUEST with 1 item\n"
        f"2. Click '×' remove button on item in Order Summary\n"
        f"3. Observe page behaviour",
        "Page redirects to Cart page showing 'Your Cart is Empty'; 'Item Removed from Cart' toast shown; 'EXPLORE PRODUCTS' button visible; guest redirect behaviour matches logged-in user behaviour (S001)",
        "", "Not Tested", "Edge Case", "High",
        "Mapped from S001 — empty cart redirect. Same redirect behaviour must apply to guest users as logged-in users."
    ),

    (
        "Guest024",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest session expires while filling Checkout form — verify graceful session handling",
        "Guest user on Checkout page; session token expires during form fill",
        f"1. As guest, navigate to Checkout\n"
        f"2. Begin filling address form\n"
        f"3. Leave page idle until session expires (or manually clear session cookies)\n"
        f"4. Attempt to submit the form\n"
        f"5. Observe system response",
        "System handles expired guest session gracefully; either redirects user back to Login page with message, OR re-establishes guest session automatically; form data not permanently lost; no unhandled error shown",
        "", "Not Tested", "Edge Case", "High",
        "Guest session timeout handling. Unlike logged-in users, guest has no re-authentication — system must handle this gracefully."
    ),

    (
        "Guest025",
        "Checkout - Guest Checkout",
        "[Edge Case] Guest resends OTP after countdown timer expires — verify new OTP sent successfully",
        "Guest user on OTP verification modal; countdown timer has reached 0",
        f"1. As guest, reach OTP modal (COD flow)\n"
        f"2. Wait for countdown timer to expire (e.g., 94 seconds)\n"
        f"3. Click 'Resend a new OTP' link\n"
        f"4. Observe whether new OTP is sent",
        "After timer expires, Resend OTP link becomes active; clicking sends new OTP to guest phone number; timer resets; confirmation shown: 'OTP has been resent'; previous OTP is invalidated",
        "", "Not Tested", "Edge Case", "High",
        "Resend OTP for guest checkout. Same OTP resend behaviour as logged-in users. Previous OTP must be invalidated when new one is requested."
    ),
]

# ── Load workbook ──────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded 'Checkout Page' — current rows: {ws.max_row}")

# ── Find TC_CHECKOUT_077 (Valid Login — must stay LAST) ────────────────────────
valid_login_row  = None
valid_login_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value and str(row[0].value).strip() == "TC_CHECKOUT_077":
        valid_login_row  = row[0].row
        valid_login_data = [c.value for c in row]
        break

if valid_login_row:
    start_row = valid_login_row
    print(f"TC_CHECKOUT_077 at row {valid_login_row} — will move to last")
else:
    start_row = ws.max_row + 1
    print(f"TC_CHECKOUT_077 not found — appending from row {start_row}")

# ── Write Guest TCs ────────────────────────────────────────────────────────────
for i, tc in enumerate(GUEST_TCS):
    r = start_row + i
    tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks = tc
    for col, val in enumerate([tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks], 1):
        write_cell(ws, r, col, val, bg_color(tc_type))
    print(f"  Written: {tc_id} ({tc_type}) at row {r}")

# ── Re-append Valid Login as absolute last ─────────────────────────────────────
final_row = start_row + len(GUEST_TCS)
if valid_login_data:
    for col, val in enumerate(valid_login_data, 1):
        write_cell(ws, final_row, col, val, bg_color("Positive"))
    print(f"\nTC_CHECKOUT_077 (Valid Login) re-appended at row {final_row} — LAST TC")

# ── Update metadata ────────────────────────────────────────────────────────────
total_tcs = final_row - 5
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            cell.value = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total_tcs}', cell.value)
            print(f"Metadata updated — Total TCs: {total_tcs}")

# ── Save & Replace ─────────────────────────────────────────────────────────────
wb.save(TMP)
print(f"\nSaved to: {TMP}")

for attempt in range(20):
    try:
        os.replace(TMP, FILE)
        print(f"\nSUCCESS: SunnyDiamonds_v2.xlsx updated!")
        print(f"Guest001–Guest025 added at rows {start_row}–{start_row + len(GUEST_TCS) - 1}")
        print(f"TC_CHECKOUT_077 (Valid Login) at row {final_row} — LAST")
        print(f"Total Checkout TCs: {total_tcs}")
        break
    except PermissionError:
        print(f"File locked ({attempt+1}/20) — please close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"\nFile still locked. Saved as: {TMP}")
    print("Close Excel then rename SunnyDiamonds_v2_guest.xlsx → SunnyDiamonds_v2.xlsx")
