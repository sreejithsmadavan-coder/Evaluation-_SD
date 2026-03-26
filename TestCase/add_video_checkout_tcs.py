import shutil, os, sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_work2.xlsx"

GREEN  = "E2EFDA"
RED    = "FCE4D6"
YELLOW = "FFF2CC"

thin   = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border    = border
    return c

def bg(tc_type):
    return GREEN if tc_type == "Positive" else (RED if tc_type == "Negative" else YELLOW)

NAV = "Navigate back to https://qa-sunnydiamonds.webc.in/checkout"

# ── New TCs observed directly from the screen-recording video ─────────────────
# (TC_ID, Module, Description, Preconditions, Steps, Expected, Actual, Status, Type, Priority, Remarks)

VIDEO_TCS = [
    # ─── A. Edit Cart from Checkout ──────────────────────────────────────────
    (
        "TC_CHECKOUT_VIDEO_01",
        "Checkout - Cart Edit",
        "Click 'Edit Cart' link in Order Summary on Checkout page — verify navigation to Cart page",
        "",
        "1. " + NAV + "\n"
        "2. Locate the Order Summary section (top-right)\n"
        "3. Click the 'Edit Cart' link next to the Order Summary heading\n"
        "4. Observe page navigation",
        "User is redirected to the Cart page (/cart); cart items, quantities, and prices are displayed correctly; breadcrumb updates to 'Home > Cart'",
        "", "Not Tested", "Positive", "High",
        "Observed in video (0s–4s): User clicks 'Edit Cart' in Order Summary and lands on Cart page. Mapped from checklist: URL/navigation validation."
    ),
    (
        "TC_CHECKOUT_VIDEO_02",
        "Checkout - Cart Edit",
        "Increase cart item quantity from Cart page — navigate back to Checkout — verify Order Summary total updates",
        "",
        "1. " + NAV + "\n"
        "2. Click 'Edit Cart' in Order Summary\n"
        "3. On Cart page, click '+' (increment) button on an item to increase quantity (e.g., qty 1 → 6)\n"
        "4. Observe 'Item Added to Cart' toast notification\n"
        "5. Click 'CHECKOUT SECURELY' button\n"
        "6. On Checkout page, verify Order Summary reflects new quantity and recalculated total",
        "'Item Added to Cart' toast displayed on cart page; on returning to Checkout, Order Summary shows updated quantity (e.g., 6x item); Subtotal and Total are recalculated correctly (qty × unit price)",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (4s–12s): Qty increased on Cart, 'Item Added to Cart' toast shown, Checkout total updated to new amount (e.g., 6 x 47,419 = 2,84,514). Financial accuracy test."
    ),
    (
        "TC_CHECKOUT_VIDEO_03",
        "Checkout - Cart Edit",
        "Remove item from Cart page — navigate back to Checkout — verify Order Summary updates",
        "",
        "1. " + NAV + "\n"
        "2. Click 'Edit Cart' in Order Summary\n"
        "3. On Cart page, click the delete/remove icon on a cart item\n"
        "4. Observe 'Item Removed from Cart' toast notification\n"
        "5. Click 'CHECKOUT SECURELY'\n"
        "6. On Checkout page, verify Order Summary no longer shows removed item and total is recalculated",
        "'Item Removed from Cart' toast displayed; Checkout Order Summary no longer shows the removed item; Subtotal and Total recalculate correctly based on remaining items",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (20s): 'Item Removed from Cart' toast visible. Order Summary on checkout updates after cart edit. Mapped from checklist §5 cart modification during checkout."
    ),
    (
        "TC_CHECKOUT_VIDEO_04",
        "Checkout - Cart Edit",
        "Decrease cart item quantity from Cart page — navigate back to Checkout — verify recalculated total",
        "",
        "1. " + NAV + "\n"
        "2. Click 'Edit Cart' in Order Summary\n"
        "3. On Cart page, click '-' (decrement) button to decrease quantity of an item\n"
        "4. Click 'CHECKOUT SECURELY'\n"
        "5. On Checkout page, verify Order Summary shows reduced quantity and lower total",
        "Quantity decremented on cart; Checkout Order Summary reflects reduced qty; Subtotal and Total recalculate correctly (new qty × unit price)",
        "", "Not Tested", "Positive", "High",
        "Cart quantity decrease test. Complements TC_CHECKOUT_VIDEO_02 (increase). Financial accuracy verification on checkout."
    ),
    # ─── B. Payment Method Button Label ──────────────────────────────────────
    (
        "TC_CHECKOUT_VIDEO_05",
        "Checkout - Payment Method",
        "Select 'Cash on Delivery' — verify PAY NOW button label changes to 'PLACE ORDER'",
        "",
        "1. " + NAV + "\n"
        "2. Scroll down to Payment Method section\n"
        "3. Select 'Cash on Delivery' radio button\n"
        "4. Observe the submit button label at the bottom of the form",
        "When 'Cash on Delivery' is selected, the submit button label changes from 'PAY NOW' to 'PLACE ORDER'; button is enabled and visible",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (28s–36s): COD selected, button clearly shows 'PLACE ORDER' instead of 'PAY NOW'. Distinct UX for COD vs online payment."
    ),
    (
        "TC_CHECKOUT_VIDEO_06",
        "Checkout - Payment Method",
        "Switch from Cash on Delivery to 'Pay Online' — verify button label changes back to 'PAY NOW'",
        "",
        "1. " + NAV + "\n"
        "2. Scroll to Payment Method section\n"
        "3. Select 'Cash on Delivery' radio button (button shows PLACE ORDER)\n"
        "4. Then select 'Pay Online' radio button\n"
        "5. Observe submit button label change",
        "Button label updates dynamically to 'PAY NOW' when Pay Online is selected; button label updates correctly on every payment method toggle",
        "", "Not Tested", "Positive", "High",
        "Observed in video (16s): Pay Online selected, button shows 'PAY NOW'. Toggle behavior between COD and Pay Online must update button label dynamically."
    ),
    # ─── C. COD — PLACE ORDER & Processing Modal ─────────────────────────────
    (
        "TC_CHECKOUT_VIDEO_07",
        "Checkout - COD Order",
        "Fill all mandatory fields, select Cash on Delivery, click 'PLACE ORDER' — verify processing modal appears",
        "",
        "1. " + NAV + "\n"
        "2. Fill all mandatory Shipping Address fields with valid data\n"
        "3. Scroll down to Payment Method\n"
        "4. Select 'Cash on Delivery'\n"
        "5. Click 'PLACE ORDER' button\n"
        "6. Observe immediate response",
        "Processing modal appears with message: 'Please wait as we are processing your order, this will only take a while'; modal includes a loading indicator; 'PLACE ORDER' button is disabled",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (40s): After clicking PLACE ORDER, processing/loading modal shown. COD order placement flow step 1."
    ),
    # ─── D. OTP Verification Modal ───────────────────────────────────────────
    (
        "TC_CHECKOUT_VIDEO_08",
        "Checkout - OTP Verification",
        "After clicking PLACE ORDER (COD) — verify 'Verify Your Mobile Number' OTP modal appears",
        "",
        "1. " + NAV + "\n"
        "2. Fill all mandatory fields with valid data\n"
        "3. Select Cash on Delivery\n"
        "4. Click PLACE ORDER\n"
        "5. Wait for processing modal to complete\n"
        "6. Observe OTP verification modal",
        "OTP verification modal titled 'Verify Your Mobile Number' appears; modal contains: instruction text mentioning masked phone number (e.g., +91XXXXXX48), 4-digit OTP input boxes, 'Resend a new OTP in X seconds' countdown timer, and 'CONFIRM AND PLACE ORDER' button",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (44s): 'Verify Your Mobile Number' modal with OTP input boxes, masked phone, resend timer (94s). COD requires phone OTP verification."
    ),
    (
        "TC_CHECKOUT_VIDEO_09",
        "Checkout - OTP Verification",
        "Verify OTP modal displays masked registered phone number correctly",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields with valid data including Phone Number\n"
        "3. Select Cash on Delivery and click PLACE ORDER\n"
        "4. Wait for OTP modal to appear\n"
        "5. Verify the phone number shown in modal matches registered number (masked format)",
        "OTP modal displays phone number in masked format (e.g., +91XXXXXX48 — only last 2 digits visible); masked number matches the phone entered in Shipping Address field",
        "", "Not Tested", "Positive", "High",
        "Observed in video (44s): Modal shows '+91XXXXXXXX48' masked format. Privacy requirement — full phone number must not be exposed in modal text."
    ),
    (
        "TC_CHECKOUT_VIDEO_10",
        "Checkout - OTP Verification",
        "Enter valid 4-digit OTP and click 'CONFIRM AND PLACE ORDER' — verify order placed successfully",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields, select COD, click PLACE ORDER\n"
        "3. Wait for OTP modal to appear\n"
        "4. Enter the valid 4-digit OTP received on registered phone\n"
        "5. Click 'CONFIRM AND PLACE ORDER' button\n"
        "6. Observe result",
        "Order is confirmed; OTP modal closes; user is redirected to success page displaying 'Thank You For The Purchase!' message; confirmation email notification shown",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (52s–56s): OTP '1234' entered, CONFIRM AND PLACE ORDER clicked, success page 'Thank You For The Purchase!' shown. Core COD happy path."
    ),
    (
        "TC_CHECKOUT_VIDEO_11",
        "Checkout - OTP Verification",
        "Enter invalid/wrong OTP in OTP verification modal — verify error message displayed",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields, select COD, click PLACE ORDER\n"
        "3. Wait for OTP modal\n"
        "4. Enter an incorrect 4-digit OTP (e.g., '0000' or any OTP not matching the sent OTP)\n"
        "5. Click 'CONFIRM AND PLACE ORDER'\n"
        "6. Observe error handling",
        "Error message displayed in OTP modal: 'Invalid OTP. Please try again.' or similar; OTP input fields cleared or highlighted in red; order is NOT placed; modal remains open",
        "", "Not Tested", "Negative", "Critical",
        "Mapped from checklist §16 error messages. Invalid OTP must not allow order placement. Security — prevent OTP bypass."
    ),
    (
        "TC_CHECKOUT_VIDEO_12",
        "Checkout - OTP Verification",
        "Leave OTP fields empty and click 'CONFIRM AND PLACE ORDER' — verify validation error",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields, select COD, click PLACE ORDER\n"
        "3. Wait for OTP modal\n"
        "4. Do NOT enter any OTP — leave all 4 input boxes empty\n"
        "5. Click 'CONFIRM AND PLACE ORDER'\n"
        "6. Observe validation response",
        "Validation error displayed: 'Please enter OTP' or 'OTP is required'; CONFIRM button does not proceed; order is NOT placed",
        "", "Not Tested", "Negative", "High",
        "Mapped from checklist §1 mandatory field validation. OTP is required to proceed. Must not allow empty OTP submission."
    ),
    (
        "TC_CHECKOUT_VIDEO_13",
        "Checkout - OTP Verification",
        "Verify 'Resend OTP' countdown timer is displayed in OTP modal",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields, select COD, click PLACE ORDER\n"
        "3. Wait for OTP modal to appear\n"
        "4. Observe the resend OTP timer text\n"
        "5. Wait and observe timer countdown",
        "OTP modal displays 'Resend a new OTP in X seconds' countdown (starts at ~94 seconds per video); timer counts down each second; 'Resend' option is disabled/greyed out while timer is active",
        "", "Not Tested", "Positive", "High",
        "Observed in video (44s): 'Resend a new OTP in 94 seconds' timer visible. Resend OTP timer prevents spam — industry standard for OTP modals."
    ),
    (
        "TC_CHECKOUT_VIDEO_14",
        "Checkout - OTP Verification",
        "Click 'Resend OTP' after countdown expires — verify new OTP is sent",
        "",
        "1. " + NAV + "\n"
        "2. Fill all fields, select COD, click PLACE ORDER\n"
        "3. Wait for OTP modal\n"
        "4. Wait for countdown timer to expire (reach 0 seconds)\n"
        "5. Click 'Resend a new OTP' link/button\n"
        "6. Observe toast or confirmation that new OTP was sent",
        "After timer expires, 'Resend OTP' becomes clickable/active; clicking it sends a new OTP to the registered phone; timer resets; confirmation message: 'OTP has been resent to +91XXXXXX48'",
        "", "Not Tested", "Positive", "High",
        "Observed in video: 'Resend a new OTP in 94 seconds' countdown. After expiry, resend must work. Mapped from checklist §11 phone number functionality."
    ),
    # ─── E. Order Success Page ────────────────────────────────────────────────
    (
        "TC_CHECKOUT_VIDEO_15",
        "Checkout - Order Success",
        "Verify 'Thank You For The Purchase!' success page displayed after COD order with valid OTP",
        "",
        "1. " + NAV + "\n"
        "2. Fill all mandatory fields with valid data\n"
        "3. Select Cash on Delivery\n"
        "4. Click PLACE ORDER → process → enter valid OTP → click CONFIRM AND PLACE ORDER\n"
        "5. Observe success page",
        "Success page displays: 'Thank You For The Purchase!'; sub-message: 'Confirmation Mail Has Been Sent To Your Mail Id'; 'We Hope You Enjoy Your Purchase'; green checkmark icon; 'CONTINUE SHOPPING' button is visible",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (56s): Full success page 'Thank You For The Purchase!' with confirmation email sent message. Validates end-to-end COD order flow."
    ),
    (
        "TC_CHECKOUT_VIDEO_16",
        "Checkout - Order Success",
        "Click 'CONTINUE SHOPPING' on success page — verify navigation to home or PLP page",
        "",
        "1. Complete a successful COD order (as per TC_CHECKOUT_VIDEO_15)\n"
        "2. On success page, click 'CONTINUE SHOPPING' button\n"
        "3. Observe navigation destination",
        "User is redirected to the Home page or All Jewellery PLP (/jewellery); cart icon shows 0 items; previous order is no longer in cart",
        "", "Not Tested", "Positive", "Medium",
        "Observed in video (56s): 'CONTINUE SHOPPING' button visible on success page. Post-order navigation must work correctly. Cart should be cleared after successful order."
    ),
]

# ── Load workbook ──────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded 'Checkout Page' — current rows: {ws.max_row}")

# ── Find Valid Login TC (must remain LAST) ────────────────────────────────────
valid_login_row  = None
valid_login_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value and "TC_CHECKOUT_076" in str(row[0].value):
        valid_login_row  = row[0].row
        valid_login_data = [c.value for c in row]
        break

if valid_login_row is None:
    # fallback: use last row
    valid_login_row = ws.max_row
    valid_login_data = None
    print("WARNING: TC_CHECKOUT_076 not found; appending after last row")
else:
    print(f"Valid Login TC found at row {valid_login_row} — will overwrite and re-append last")

# ── Write new video TCs starting at Valid Login row ───────────────────────────
start_row = valid_login_row

for i, tc in enumerate(VIDEO_TCS):
    r = start_row + i
    tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks = tc
    for col, val in enumerate([tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks], 1):
        write_cell(ws, r, col, val, bg(tc_type))

print(f"Added {len(VIDEO_TCS)} video TCs at rows {start_row}–{start_row + len(VIDEO_TCS) - 1}")

# ── Re-append Valid Login as the absolute last TC ────────────────────────────
final_row = start_row + len(VIDEO_TCS)
if valid_login_data:
    for col, val in enumerate(valid_login_data, 1):
        write_cell(ws, final_row, col, val, bg("Positive"))
    ws.cell(row=final_row, column=1).value = "TC_CHECKOUT_077"   # renumber
    print(f"Valid Login TC re-appended at row {final_row} as TC_CHECKOUT_077 (last)")
else:
    # write default valid login TC
    vl = [
        "TC_CHECKOUT_077",
        "Checkout - Authentication",
        "Valid Login — Verify authenticated user (sreejith.s+4@webandcrafts.com) can access Checkout",
        "",
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n2. Login with sreejith.s+4@webandcrafts.com / Password\n3. Add product to cart\n4. Navigate to Checkout\n5. Verify full checkout page loads",
        "User authenticated; Checkout page loads with all sections; order can be placed successfully",
        "", "Not Tested", "Positive", "Critical",
        "Valid login TC — ALWAYS LAST per QA standard. sreejith.s+4@webandcrafts.com / Password."
    ]
    for col, val in enumerate(vl, 1):
        write_cell(ws, final_row, col, val, bg("Positive"))
    print(f"Valid Login TC written at row {final_row} as TC_CHECKOUT_077")

# ── Update metadata total TC count ────────────────────────────────────────────
total_tcs = final_row - 5
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            cell.value = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total_tcs}', cell.value)
            print(f"Metadata updated — Total TCs: {total_tcs}")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(TMP)
try:
    os.replace(TMP, FILE)
    print(f"\nSUCCESS: SunnyDiamonds_v2.xlsx saved — Checkout Page now has {total_tcs} TCs")
    print(f"Last TC: TC_CHECKOUT_077 (Valid Login) at row {final_row}")
except PermissionError:
    print(f"\nFile locked. Saved as: {TMP}\nClose SunnyDiamonds_v2.xlsx in Excel and rename manually.")
