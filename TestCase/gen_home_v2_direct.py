# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET = "Home Page"

wb = load_workbook(FILE)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# ── Colours ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GOLD_LIGHT  = "FFF3CD"
POS_GREEN   = "E2EFDA"
NEG_RED     = "FCE4D6"
EDGE_YELLOW = "FFF2CC"
WHITE       = "FFFFFF"
GREY_MID    = "D9D9D9"
GREY_LIGHT  = "F2F2F2"

def S(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

thin  = Border(left=S(), right=S(), top=S(), bottom=S())
thick = Border(left=S(), right=S(), top=S(), bottom=S("medium","1F3864"))

# ── Row 1 Banner ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:L1")
c = ws.cell(row=1, column=1, value="SUNNY DIAMONDS — HOME PAGE QA TEST CASES")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=16)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Row 2 Metadata ────────────────────────────────────────────────────────────
meta = [
    ("A2","Project URL:","B2","https://qa-sunnydiamonds.webc.in/"),
    ("D2","Module:","E2","Home Page"),
    ("G2","Created By:","H2","Sreejith S Madavan"),
    ("J2","Created Date:","K2","25-Mar-2026"),
]
for lc,lv,vc,vv in meta:
    l=ws[lc]; l.value=lv
    l.font=Font(name="Arial",bold=True,color=DARK_BLUE,size=10)
    l.fill=PatternFill("solid",fgColor=GREY_MID)
    l.alignment=Alignment(horizontal="left",vertical="center"); l.border=thin
    v=ws[vc]; v.value=vv
    v.font=Font(name="Arial",color=DARK_BLUE,size=10)
    v.fill=PatternFill("solid",fgColor=GREY_LIGHT)
    v.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    v.border=thin
ws.merge_cells("B2:C2"); ws.merge_cells("E2:F2")
ws.merge_cells("H2:I2"); ws.merge_cells("K2:L2")
ws.row_dimensions[2].height = 22

for col in range(1,13):
    ws.cell(row=3,column=col).fill = PatternFill("solid",fgColor=DARK_BLUE)
ws.row_dimensions[3].height = 6

# ── Row 4 Headers ─────────────────────────────────────────────────────────────
HEADERS = ["Test Case ID","Module Name","Test Case Description","Preconditions",
           "Test Steps","Expected Result","Actual Result","Status",
           "Test Type","Priority","Playwright Navigation","Remarks"]
for i,h in enumerate(HEADERS,1):
    c=ws.cell(row=4,column=i,value=h)
    c.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=MID_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=thick
ws.row_dimensions[4].height = 30

# ── Shared content ────────────────────────────────────────────────────────────
URL = "https://qa-sunnydiamonds.webc.in/"
PRE_BASE = (f"1. Application is accessible at {URL}\n"
            "2. Browser is open\n"
            "3. User is NOT logged in (guest state)")
PRE_COOKIE = (f"1. Application is accessible at {URL}\n"
              "2. Browser is open (fresh session or cookies cleared)")
PRE_LOGGED = (f"1. Application accessible at {URL}\n"
              "2. Valid account exists: sreejith.s+4@webandcrafts.com / Password\n"
              "3. User is NOT currently logged in")

NAV_FIRST  = "page.goto('https://qa-sunnydiamonds.webc.in/')\n[FIRST TEST — direct navigation]"
NAV_BACK   = "await page.goBack()\n[Returns to Home Page from previous state]"
NAV_BACK_F = "await page.goBack()  // fallback: page.goto(URL)"

# ── Test Cases ─────────────────────────────────────────────────────────────────
# Tuple: (ID, Module, Desc, Pre, Steps, Expected, ActualResult, Status, Type, Priority, PlaywrightNav, Remarks)
TC = [

    # ═══════════════════════ POSITIVE ════════════════════════════════════════
    (
        "TC_HOME_001","Home Page",
        "Verify Home Page loads correctly with all major sections visible",
        PRE_BASE,
        "1. Navigate to https://qa-sunnydiamonds.webc.in/\n"
        "2. Wait for page to fully load\n"
        "3. Verify the following sections are visible:\n"
        "   - Announcement banner\n"
        "   - Primary navigation menu\n"
        "   - Search bar\n"
        "   - Hero/banner section with 'Shop Now' CTA\n"
        "   - 'Shop by Category' section\n"
        "   - Featured banners (Earrings, Pendants, Rings)\n"
        "   - Trending Products section\n"
        "   - Gift Ideas section\n"
        "   - Product carousel\n"
        "   - Testimonials section\n"
        "   - 'Our Promise to You' trust badges\n"
        "   - Newsletter subscription\n"
        "   - Footer with all columns",
        "Home page loads successfully within acceptable time\n"
        "All sections listed are visible without scrolling issues\n"
        "No broken layout or overlapping elements\n"
        "No console errors (JS/CSS)\n"
        "Page title is 'Sunny Diamonds' (or similar brand title)",
        "","Pass/Fail","Positive","High",
        NAV_FIRST,
        "TC_HOME_001 — FIRST TC: uses direct page.goto(). All subsequent TCs use goBack()"
    ),
    (
        "TC_HOME_002","Home Page",
        "Verify Announcement Banner displays the promotional message correctly",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the very top of the page\n"
        "3. Locate the announcement banner\n"
        "4. Read and verify the banner text",
        "Announcement banner is visible at the top of the page\n"
        "Banner displays: 'Sunny Diamonds Celebrates The \"Season of Sparkle\" from October 15th to November 15th.!'\n"
        "Text is fully readable without truncation",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Announcement banner content verification"
    ),
    (
        "TC_HOME_003","Home Page",
        "Verify Sunny Diamonds logo/brand name navigates back to Home Page when clicked",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click on any internal page link (e.g., About Us)\n"
        "3. Once on the new page, click the 'Sunny Diamonds' logo/brand name in the header\n"
        "4. Verify the destination URL",
        "User is redirected back to https://qa-sunnydiamonds.webc.in/\n"
        "Home page reloads correctly\n"
        "All home page sections are visible",
        "","Pass/Fail","Positive","High",
        "await page.goBack()  // navigate away first, then click logo",
        "Logo navigation — brand identity anchor"
    ),
    (
        "TC_HOME_004","Home Page",
        "Verify ALL JEWELLERY mega menu opens and displays all sub-category options",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Hover over or click 'ALL JEWELLERY' in the main navigation\n"
        "3. Observe the mega menu dropdown\n"
        "4. Verify all listed sub-categories are present:\n"
        "   - Category: All, Earrings, Nosepins, Rings, Pendants, Necklace, Bracelets, Bangles\n"
        "   - Metal Color: Yellow Gold, Rose Gold, White Gold\n"
        "   - Metal Purity: 22K, 18K\n"
        "   - Price: Below 10K, 10K-30K, 30K-50K, 50K-80K, Above 80K\n"
        "   - Explore, Solitaire, Trending links",
        "Mega menu opens on hover/click\n"
        "All sub-categories listed are displayed correctly\n"
        "Menu items are clickable\n"
        "Menu closes when clicking elsewhere",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Mega menu — ALL JEWELLERY dropdown validation"
    ),
    (
        "TC_HOME_005","Home Page",
        "Verify EARRINGS navigation menu item navigates to Earrings listing page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click 'EARRINGS' in the main navigation menu\n"
        "3. Verify the destination URL and page content",
        "User navigated to https://qa-sunnydiamonds.webc.in/jewellery/earrings\n"
        "Earrings listing page loads with product grid\n"
        "Page title/header shows 'Diamond Earrings'\n"
        "Filter and sort options are visible",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Nav menu — EARRINGS link"
    ),
    (
        "TC_HOME_006","Home Page",
        "Verify TRENDING navigation item navigates to Trending products page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click 'TRENDING' in the main navigation menu\n"
        "3. Verify the destination",
        "User navigated to https://qa-sunnydiamonds.webc.in/trending\n"
        "Trending products page loads with product grid\n"
        "1000+ products visible with filter/sort options",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Nav menu — TRENDING link"
    ),
    (
        "TC_HOME_007","Home Page",
        "Verify Hero Section 'Shop Now' CTA button navigates to jewellery listing",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Locate the Hero/Banner section at the top\n"
        "3. Click the 'Shop Now' button",
        "User is navigated to a jewellery listing or category page\n"
        "Products are displayed\n"
        "URL changes from home page URL",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Hero CTA — 'Shop Now' button"
    ),
    (
        "TC_HOME_008","Home Page",
        "Verify 'Shop by Category' — Rings tile navigates to Rings listing page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Shop by Category' section\n"
        "3. Click the 'Rings' category tile",
        "User navigated to https://qa-sunnydiamonds.webc.in/jewellery/rings\n"
        "Rings listing page loads with product grid\n"
        "Page shows '635 Products' with filters",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Shop by Category — Rings tile"
    ),
    (
        "TC_HOME_009","Home Page",
        "Verify 'Shop by Category' — SHOP ALL tile navigates to All Jewellery listing",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Shop by Category' section\n"
        "3. Click the 'SHOP ALL' tile",
        "User navigated to jewellery/all-products page or /jewellery\n"
        "All jewellery products are displayed with filter and sort",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Shop by Category — SHOP ALL tile"
    ),
    (
        "TC_HOME_010","Home Page",
        "Verify Featured Banner 'EXPLORE EARRINGS' navigates to Earrings page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the featured editorial banners section\n"
        "3. Click 'EXPLORE EARRINGS' banner/button",
        "User navigated to Earrings listing page\n"
        "URL: https://qa-sunnydiamonds.webc.in/jewellery/earrings\n"
        "Page loads with earrings products grid",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Featured banner — EXPLORE EARRINGS CTA"
    ),
    (
        "TC_HOME_011","Home Page",
        "Verify Trending Products section shows 10 products with correct names and prices",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Trending products' section\n"
        "3. Count visible products\n"
        "4. Verify each product shows: name, price (in INR), 'ADD TO CART' button",
        "Exactly 10 trending products are displayed\n"
        "Each product shows: product name, price in ₹ format, ADD TO CART button\n"
        "First product: 'Isabelette Diamond Earring — ₹48,970'\n"
        "Last product: 'Quinlan Diamond Earring — ₹26,714'",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Trending Products — count and content validation"
    ),
    (
        "TC_HOME_012","Home Page",
        "Verify 'View All' link in Trending Products section navigates to Trending page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Trending products' section\n"
        "3. Click the 'View All' link",
        "User navigated to https://qa-sunnydiamonds.webc.in/trending\n"
        "Trending page shows 1000+ products",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Trending Products — View All link"
    ),
    (
        "TC_HOME_013","Home Page",
        "Verify Gift Ideas 'Find Your Gift' CTA navigates to Gifts page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Make Every Moment Special' section\n"
        "3. Click 'Find Your Gift' button",
        "User navigated to the Gifts/Gift Ideas page\n"
        "URL changes to gifts page\n"
        "Gift category products are displayed",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Gift Ideas — Find Your Gift CTA"
    ),
    (
        "TC_HOME_014","Home Page",
        "Verify Gift Ideas 'Gifts Under ₹10,000' tile navigates to filtered products",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the Gift Ideas section\n"
        "3. Click the 'Gifts Under ₹10,000' tile",
        "User navigated to jewellery listing filtered by price range 0-10000\n"
        "Products displayed are priced at or below ₹10,000",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Gift Ideas — price-range tile navigation"
    ),
    (
        "TC_HOME_015","Home Page",
        "Verify product carousel 'SHOP NOW' navigates to correct product detail page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the product carousel section\n"
        "3. Click 'SHOP NOW' on 'ISABELETTE DIAMOND EARRING'",
        "User navigated to the product detail page for Isabelette Diamond Earring\n"
        "Product name, price (₹48,970), images, and Add to Cart are visible\n"
        "URL contains the product slug",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Product Carousel — SHOP NOW navigation"
    ),
    (
        "TC_HOME_016","Home Page",
        "Verify Testimonials section displays customer reviews with names and dates",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Testimonials — Real Stories, Real Smiles' section\n"
        "3. Verify customer review cards are visible\n"
        "4. Check for customer name, review date, and review text",
        "Testimonials section is visible\n"
        "Multiple customer reviews are displayed\n"
        "Each card shows: customer name, date, review text\n"
        "Carousel/scroll is functional (if applicable)",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Testimonials section — content and carousel validation"
    ),
    (
        "TC_HOME_017","Home Page",
        "Verify 'Our Promise to You' section displays all 8 trust badges",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to 'Our Promise to You — Pure. Certified. Trusted.' section\n"
        "3. Count and verify each trust badge",
        "All 8 trust badges are visible:\n"
        "1. Internally Flawless Diamonds\n"
        "2. 100% Money Back on Diamond Value\n"
        "3. Certifications of Diamonds\n"
        "4. BIS Hall Mark for Jewellery\n"
        "5. Brand Assured Quality\n"
        "6. 15 Days Return Policy\n"
        "7. Cash On Delivery\n"
        "8. Pan India Free Shipping\n"
        "Each badge has an icon and label",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Trust badges — all 8 present and visible"
    ),
    (
        "TC_HOME_018","Home Page",
        "Verify Newsletter subscription with a valid email address",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the Newsletter section at the bottom ('Keep in touch')\n"
        "3. Enter a valid email: 'test.newsletter@example.com'\n"
        "4. Click the 'Subscribe' button",
        "Subscription is accepted\n"
        "Success message displayed (e.g., 'Thank you for subscribing!')\n"
        "No error message shown\n"
        "Email field is cleared after successful subscription",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Newsletter — valid email subscription"
    ),
    (
        "TC_HOME_019","Home Page",
        "Verify Footer — 'About Us' link navigates to About Us page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to footer\n"
        "3. Under 'COMPANY' section, click 'About Us'",
        "User navigated to https://qa-sunnydiamonds.webc.in/about-us\n"
        "About Us page loads correctly\n"
        "No 404 error",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Footer navigation — About Us"
    ),
    (
        "TC_HOME_020","Home Page",
        "Verify Footer — 'FAQs' link navigates to FAQ page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to footer\n"
        "3. Under 'SUPPORT' section, click 'FAQs'",
        "User navigated to https://qa-sunnydiamonds.webc.in/faq\n"
        "FAQ page loads with topics and answers",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Footer navigation — FAQs"
    ),
    (
        "TC_HOME_021","Home Page",
        "Verify Cookie Consent — 'Accept All' button works and banner dismisses",
        PRE_COOKIE,
        "1. [goBack() to Home Page] (or fresh session)\n"
        "2. Observe the cookie consent banner at the bottom\n"
        "3. Click 'Accept All' button",
        "Cookie consent banner is dismissed\n"
        "Page remains on home page\n"
        "Cookie preference is saved (banner does not reappear on refresh)\n"
        "No errors in console",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Cookie consent — Accept All"
    ),
    (
        "TC_HOME_022","Home Page",
        "Verify Cookie Consent — 'Decline' button dismisses banner without saving cookies",
        PRE_COOKIE,
        "1. [goBack() to Home Page] (or fresh session)\n"
        "2. Observe the cookie consent banner\n"
        "3. Click 'Decline' button",
        "Cookie consent banner is dismissed\n"
        "Non-essential cookies are NOT set\n"
        "Page continues to function normally",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Cookie consent — Decline"
    ),
    (
        "TC_HOME_023","Home Page",
        "Verify header 'Log In' link navigates to Login page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Locate 'Log In' link in the header account panel\n"
        "3. Click 'Log In'",
        "User navigated to https://qa-sunnydiamonds.webc.in/login\n"
        "Login page loads with Email and Password fields",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Header — Log In link navigation"
    ),
    (
        "TC_HOME_024","Home Page",
        "Verify header 'Sign Up' link navigates to Registration page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Locate 'Sign Up' link in the header account panel\n"
        "3. Click 'Sign Up'",
        "User navigated to https://qa-sunnydiamonds.webc.in/create\n"
        "Registration page loads with all form fields",
        "","Pass/Fail","Positive","High",
        NAV_BACK,
        "Header — Sign Up link navigation"
    ),
    (
        "TC_HOME_025","Home Page",
        "Verify 'Store Locator' top bar link navigates to Store Locator page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click 'Store Locator' in the top utility bar",
        "User navigated to the Store Locator page\n"
        "Store location information is displayed\n"
        "No 404 error",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Top bar — Store Locator link"
    ),
    (
        "TC_HOME_026","Home Page",
        "Verify 'Order Tracking' top bar link navigates to Order Tracking form",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click 'Order Tracking' in the top utility bar",
        "User navigated to https://qa-sunnydiamonds.webc.in/sales/guest/form\n"
        "Order Tracking form displays: Order ID, Email, Phone Number fields",
        "","Pass/Fail","Positive","Medium",
        NAV_BACK,
        "Top bar — Order Tracking link"
    ),

    # ═══════════════════════ NEGATIVE ════════════════════════════════════════
    (
        "TC_HOME_027","Home Page",
        "Verify Newsletter subscription with an invalid email format",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to Newsletter section ('Keep in touch')\n"
        "3. Enter invalid email: 'invalidemail'\n"
        "4. Click 'Subscribe' button",
        "Subscription is NOT processed\n"
        "Validation error displayed: 'Please enter a valid email address'\n"
        "User remains on home page",
        "","Pass/Fail","Negative","High",
        NAV_BACK,
        "Newsletter — invalid email validation"
    ),
    (
        "TC_HOME_028","Home Page",
        "Verify Newsletter subscription with empty email field",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to Newsletter section\n"
        "3. Leave the email field empty\n"
        "4. Click 'Subscribe' button",
        "Subscription is NOT processed\n"
        "Validation error: 'Email is required' or similar\n"
        "User remains on home page",
        "","Pass/Fail","Negative","High",
        NAV_BACK,
        "Newsletter — empty email field"
    ),
    (
        "TC_HOME_029","Home Page",
        "Verify accessing an invalid/non-existent URL displays a 404 error page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. In the address bar, type: https://qa-sunnydiamonds.webc.in/this-page-does-not-exist\n"
        "3. Press Enter and observe the result",
        "A 404 error page is displayed\n"
        "Page shows '404 — Page Not Found' message\n"
        "'Back to Home' button is visible and functional\n"
        "Header and footer are still rendered correctly",
        "","Pass/Fail","Negative","High",
        "page.goto('https://qa-sunnydiamonds.webc.in/this-page-does-not-exist')\n// Direct nav for 404 test",
        "404 error page handling — invalid URL"
    ),
    (
        "TC_HOME_030","Home Page",
        "Verify Search bar — clicking search icon with empty query does not crash or navigate incorrectly",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Click the Search icon in the header\n"
        "3. Leave the search input empty\n"
        "4. Press Enter or click Search",
        "No crash or error\n"
        "Either: no navigation occurs, OR user is redirected to listing with all products\n"
        "No blank/error page is shown",
        "","Pass/Fail","Negative","Medium",
        NAV_BACK,
        "Search — empty query handling"
    ),

    # ═══════════════════════ EDGE CASES ══════════════════════════════════════
    (
        "TC_HOME_031","Home Page",
        "Verify Home Page layout is responsive on mobile viewport (375x812 — iPhone SE)",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. In Playwright: await page.setViewportSize({width:375, height:812})\n"
        "3. Reload home page\n"
        "4. Verify:\n"
        "   - Navigation collapses to hamburger or mobile menu\n"
        "   - Bottom navigation (Home, Profile, Cart, Search) is visible\n"
        "   - All sections stack vertically\n"
        "   - Buttons are tappable size (min 44x44px)\n"
        "   - No horizontal scroll",
        "Home page renders correctly at 375px width\n"
        "Mobile navigation (bottom bar) is visible\n"
        "No horizontal scrollbar\n"
        "All content is readable without zooming\n"
        "CTAs (Shop Now, Add to Cart) are accessible",
        "","Pass/Fail","Edge Case","High",
        "await page.setViewportSize({width:375, height:812})\nawait page.goBack()",
        "Responsive — mobile viewport 375px"
    ),
    (
        "TC_HOME_032","Home Page",
        "Verify Home Page layout is responsive on tablet viewport (768x1024 — iPad)",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. In Playwright: await page.setViewportSize({width:768, height:1024})\n"
        "3. Reload home page\n"
        "4. Inspect layout — navigation, grid, product cards",
        "Page renders correctly at 768px\n"
        "Navigation adapts appropriately\n"
        "Product cards maintain proper grid layout\n"
        "No overlapping elements",
        "","Pass/Fail","Edge Case","High",
        "await page.setViewportSize({width:768, height:1024})\nawait page.goBack()",
        "Responsive — tablet viewport 768px"
    ),
    (
        "TC_HOME_033","Home Page",
        "Verify Home Page is served over HTTPS with valid SSL certificate",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Check the browser address bar for HTTPS and padlock icon\n"
        "3. Open DevTools → Security tab\n"
        "4. Verify SSL certificate validity and no mixed-content warnings",
        "Page loads over HTTPS\n"
        "Padlock icon visible in address bar\n"
        "No 'Not Secure' warning\n"
        "No mixed-content errors in DevTools console\n"
        "SSL certificate is valid and not expired",
        "","Pass/Fail","Edge Case","High",
        NAV_BACK,
        "Security — HTTPS enforcement and SSL certificate"
    ),
    (
        "TC_HOME_034","Home Page",
        "Verify Home Page browser tab title and page meta title are correct",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Check the browser tab title\n"
        "3. In DevTools, inspect the <title> tag in <head>\n"
        "4. Check meta description tag",
        "Browser tab displays brand name (e.g., 'Sunny Diamonds')\n"
        "Page <title> tag contains 'Sunny Diamonds'\n"
        "Meta description is present and relevant to the brand",
        "","Pass/Fail","Edge Case","Medium",
        NAV_BACK,
        "SEO/meta — page title and description check"
    ),
    (
        "TC_HOME_035","Home Page",
        "Verify Home Page loads without JavaScript console errors",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Open DevTools → Console tab\n"
        "3. Reload the home page\n"
        "4. Monitor for any red error messages",
        "No JavaScript errors (red) in the console\n"
        "Warnings (yellow) may be acceptable\n"
        "No 404 errors for assets (images, scripts, fonts)\n"
        "Page renders correctly",
        "","Pass/Fail","Edge Case","High",
        "await page.goBack()\n// Monitor: page.on('console', msg => console.log(msg.type(), msg.text()))",
        "Frontend quality — no JS console errors"
    ),
    (
        "TC_HOME_036","Home Page",
        "Verify image lazy loading works — product images load as user scrolls",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Open DevTools → Network tab, filter by 'Img'\n"
        "3. Load the home page\n"
        "4. Observe which images load initially (above fold)\n"
        "5. Scroll down slowly\n"
        "6. Observe images in Trending Products and carousels loading",
        "Images above the fold load immediately on page load\n"
        "Images below the fold are lazy-loaded as user scrolls\n"
        "No broken image icons (placeholder shows until loaded)\n"
        "All images eventually load without 404 errors",
        "","Pass/Fail","Edge Case","Medium",
        NAV_BACK,
        "Performance — lazy loading of product images"
    ),
    (
        "TC_HOME_037","Home Page",
        "Verify Home Page performance — page load time is within acceptable threshold",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Open DevTools → Network tab, clear cache\n"
        "3. Hard reload (Ctrl+Shift+R)\n"
        "4. Check: DOMContentLoaded time, Load event time, LCP (Largest Contentful Paint)",
        "DOMContentLoaded <= 3 seconds\n"
        "Full page Load <= 5 seconds (on standard network)\n"
        "No single resource takes > 2 seconds to load\n"
        "No 5xx server errors in network tab",
        "","Pass/Fail","Edge Case","High",
        NAV_BACK,
        "Performance — page load time threshold"
    ),
    (
        "TC_HOME_038","Home Page",
        "Verify Footer copyright year displays current year (2026)",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the very bottom of the page\n"
        "3. Read the copyright text",
        "Copyright text reads: '© 2026 Sunny Diamonds. All rights reserved.'\n"
        "Year is current (2026)\n"
        "Text is fully visible and not truncated",
        "","Pass/Fail","Edge Case","Low",
        NAV_BACK,
        "Footer — copyright year validation"
    ),
    (
        "TC_HOME_039","Home Page",
        "Verify Terms and Conditions footer link navigates to T&C page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the footer bottom bar\n"
        "3. Click 'Terms and Conditions' link",
        "User navigated to https://qa-sunnydiamonds.webc.in/terms-and-conditions\n"
        "Terms and Conditions page loads correctly\n"
        "No 404 error",
        "","Pass/Fail","Edge Case","Medium",
        NAV_BACK,
        "Footer bottom — Terms & Conditions link"
    ),
    (
        "TC_HOME_040","Home Page",
        "Verify Privacy Policy footer link navigates to Privacy Policy page",
        PRE_BASE,
        "1. [goBack() to Home Page]\n"
        "2. Scroll to the footer bottom bar\n"
        "3. Click 'Privacy Policy' link",
        "User navigated to https://qa-sunnydiamonds.webc.in/privacy-policy\n"
        "Privacy Policy page loads correctly\n"
        "No 404 error",
        "","Pass/Fail","Edge Case","Medium",
        NAV_BACK,
        "Footer bottom — Privacy Policy link"
    ),

    # ════════════════ POSITIVE — VALID LOGIN LAST ═════════════════════════════
    (
        "TC_HOME_041","Home Page",
        "Verify Home Page state after successful login — user profile and cart are accessible",
        PRE_LOGGED,
        "1. [goBack() to Home Page]\n"
        "2. Click 'Log In' in the header\n"
        "3. Enter Email: 'sreejith.s+4@webandcrafts.com'\n"
        "4. Enter Password: 'Password'\n"
        "5. Click 'Sign In'\n"
        "6. After redirect, navigate back to https://qa-sunnydiamonds.webc.in/\n"
        "7. Verify the header and account section\n"
        "   ---\n"
        "   POST-TEST RESET:\n"
        "   - Click Profile icon → Personal Details\n"
        "   - Click 'Log Out' → Confirm 'LOG OUT'",
        "Login is successful\n"
        "User is redirected to home page\n"
        "Header now shows user's account/profile (not 'Log In / Sign Up')\n"
        "Cart icon is accessible\n"
        "Wishlist is accessible\n"
        "All home page sections remain intact\n"
        "Welcome/personalized message may be displayed",
        "","Pass/Fail","Positive","High",
        "page.goto('https://qa-sunnydiamonds.webc.in/login')\n// Login → then page.goto(URL) to verify home state\n// VALID LOGIN — PLACED LAST",
        "✅ VALID LOGIN — LAST TC. Email: sreejith.s+4@webandcrafts.com | Password: Password. Verifies home page in authenticated state"
    ),
]

# ── Write rows ─────────────────────────────────────────────────────────────────
TYPE_COLORS = {"Positive":POS_GREEN,"Negative":NEG_RED,"Edge Case":EDGE_YELLOW}
PRI_COLORS  = {"High":"FFD7D7","Critical":"FFB3B3","Medium":"FFF3CD","Low":"E8F5E9"}

for r_idx, tc in enumerate(TC):
    row = 5 + r_idx
    ws.row_dimensions[row].height = 130
    (tc_id,mod,desc,pre,steps,exp,act,status,ttype,pri,pwnav,rem) = tc
    bg = TYPE_COLORS.get(ttype, WHITE)
    for c_idx,val in enumerate([tc_id,mod,desc,pre,steps,exp,act,status,ttype,pri,pwnav,rem],1):
        c = ws.cell(row=row,column=c_idx,value=val)
        c.font=Font(name="Arial",size=9,color="000000",bold=(c_idx==1))
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        c.border=thin
    ws.cell(row=row,column=9).font=Font(name="Arial",size=9,bold=True,
        color=("1F6B2E" if ttype=="Positive" else
               "7B1E0C" if ttype=="Negative" else "7B5B00"))
    p_bg = PRI_COLORS.get(pri,"FFFFFF")
    ws.cell(row=row,column=10).fill=PatternFill("solid",fgColor=p_bg)
    ws.cell(row=row,column=10).font=Font(name="Arial",size=9,bold=True,
        color=("990000" if pri=="Critical" else "000000"))
    # Playwright nav column — light blue background
    ws.cell(row=row,column=11).fill=PatternFill("solid",fgColor=LIGHT_BLUE)
    ws.cell(row=row,column=11).font=Font(name="Arial",size=8,color="1F3864")

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14,14,40,34,58,42,18,12,14,10,32,32]
for i,w in enumerate(COL_WIDTHS,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C5"

wb.save(FILE)
pos  = sum(1 for t in TC if t[8]=="Positive")
neg  = sum(1 for t in TC if t[8]=="Negative")
edge = sum(1 for t in TC if t[8]=="Edge Case")
print(f"SUCCESS — '{SHEET}' sheet added to: {FILE}")
print(f"Total: {len(TC)} | Positive: {pos} | Negative: {neg} | Edge Case: {edge}")
print(f"Last TC: {TC[-1][0]} — {TC[-1][2][:70]}")
print(f"Sheets: {wb.sheetnames}")
