"""
gen_checkout_tc.py
──────────────────────────────────────────────────────────────────────────────
Generates "Checkout Page" sheet in SunnyDiamonds_v2.xlsx
Target : https://qa-sunnydiamonds.webc.in/checkout
Author : Sreejith S Madavan
Date   : 2026-03-26
TCs    : TC_CHECKOUT_001 → TC_CHECKOUT_066 (65 test cases)
──────────────────────────────────────────────────────────────────────────────
"""

import shutil, os
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
FILE  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET = "Checkout Page"
URL   = "https://qa-sunnydiamonds.webc.in/checkout"

# ── Colours ────────────────────────────────────────────────────────────────
C_BANNER   = "1F3864"   # dark navy — banner row
C_HEADER   = "2E75B6"   # mid blue — column headers
C_META     = "D6E4F0"   # light blue — metadata row
C_LEGEND   = "EBF3FB"   # pale blue — legend row
C_PRECON   = "FFF9E6"   # pale yellow — common preconditions
C_POS      = "E2EFDA"   # green — Positive
C_NEG      = "FCE4D6"   # pink/red — Negative
C_EDGE     = "FFF2CC"   # yellow — Edge Case
C_CRIT_TXT = "990000"   # dark red — Critical priority text
C_HIGH_TXT = "7F3F00"   # brown — High priority text

def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

thin  = Side(style="thin",   color="B0B0B0")
thick = Side(style="medium", color="4472C4")
border_all  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
border_thick= Border(left=thick, right=thick, top=thick, bottom=thick)

# ── Column widths (11 columns) ──────────────────────────────────────────────
COL_WIDTHS = [18, 16, 40, 36, 60, 45, 20, 10, 14, 12, 25]

# ── Merge & write a banner / label row ─────────────────────────────────────
def set_banner(ws, row, text, bg, fg="FFFFFF", bold=True, size=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = fill(bg)
    c.font      = font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border_thick
    ws.row_dimensions[row].height = 24

# ── Apply row colour + borders ──────────────────────────────────────────────
def style_tc_row(ws, row_num, bg):
    for col in range(1, 12):
        c = ws.cell(row=row_num, column=col)
        c.fill      = fill(bg)
        c.border    = border_all
        c.alignment = wrap_align()
        c.font      = font()

def apply_priority_color(ws, row_num, priority):
    col = 10  # Priority column
    c = ws.cell(row=row_num, column=col)
    if priority == "Critical":
        c.font = font(bold=True, color=C_CRIT_TXT)
    elif priority == "High":
        c.font = font(bold=True, color=C_HIGH_TXT)

# ═══════════════════════════════════════════════════════════════════════════
#  TEST CASE DATA
# ═══════════════════════════════════════════════════════════════════════════
# Fields: (tc_id, module, description, precondition, steps, expected, tc_type, priority, remarks)

TC_DATA = [

# ─── TC_CHECKOUT_001 — Full Setup Flow ─────────────────────────────────────
(
  "TC_CHECKOUT_001",
  "Checkout Page",
  "Full setup flow — Login, add 2 products to cart, navigate to Checkout page",
  "User has valid account: sreejith.s+4@webandcrafts.com / Password",
  "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
  "2. Enter email: sreejith.s+4@webandcrafts.com\n"
  "3. Enter password: Password\n"
  "4. Click 'Login' button\n"
  "5. Navigate to 'ALL JEWELLERY' (PLP) via header menu\n"
  "6. Click on 'Aminah Diamond Ring'\n"
  "7. On PDP, select a variant and click 'ADD TO CART'\n"
  "8. Navigate back to PLP\n"
  "9. Click on '18K Rose Gold Eden Diamond Ring'\n"
  "10. On PDP, select a variant and click 'ADD TO CART'\n"
  "11. Click cart icon in header\n"
  "12. On Cart page, click 'CHECKOUT SECURELY'\n"
  "13. Observe Checkout page loads",
  "User is logged in. Two products appear in cart. Checkout page loads at /checkout with Shipping Address form, Order Summary, and Payment section visible.",
  "Positive", "Critical",
  "Full prerequisite setup for all subsequent TCs. Validates end-to-end flow."
),

# ─── TC_CHECKOUT_002 — Valid First Name ────────────────────────────────────
(
  "TC_CHECKOUT_002",
  "Checkout Page",
  "Shipping Address — Enter valid First Name (alphabets only)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'First Name' field\n"
  "3. Enter: 'Sreejith'\n"
  "4. Click outside the field (tab-out)",
  "First Name field accepts 'Sreejith'. No validation error shown.",
  "Positive", "High",
  "BVA: 8-char valid name; alpha-only. Mapped from checklist: Text field validation."
),

# ─── TC_CHECKOUT_003 — Valid Last Name ─────────────────────────────────────
(
  "TC_CHECKOUT_003",
  "Checkout Page",
  "Shipping Address — Enter valid Last Name",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Last Name' field\n"
  "3. Enter: 'Madavan'\n"
  "4. Tab out of field",
  "Last Name accepts 'Madavan' without error.",
  "Positive", "Medium",
  ""
),

# ─── TC_CHECKOUT_004 — Valid Email ─────────────────────────────────────────
(
  "TC_CHECKOUT_004",
  "Checkout Page",
  "Shipping Address — Enter valid Email Address",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Email Address' field\n"
  "3. Enter: 'sreejith.s+4@webandcrafts.com'\n"
  "4. Tab out of field",
  "Email accepted. No error message displayed.",
  "Positive", "High",
  "Mapped from checklist: Valid email format."
),

# ─── TC_CHECKOUT_005 — Valid Phone ─────────────────────────────────────────
(
  "TC_CHECKOUT_005",
  "Checkout Page",
  "Shipping Address — Enter valid 10-digit Phone Number",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Phone Number' field\n"
  "3. Enter: '9876543210'\n"
  "4. Tab out of field",
  "Phone accepted without error. 10-digit value stored.",
  "Positive", "High",
  "Mapped from checklist: Valid 10-digit phone."
),

# ─── TC_CHECKOUT_006 — Valid Address ───────────────────────────────────────
(
  "TC_CHECKOUT_006",
  "Checkout Page",
  "Shipping Address — Enter valid street Address",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Address' field\n"
  "3. Enter: '42 MG Road, Kakkanad'\n"
  "4. Tab out of field",
  "Address field accepts the input without error.",
  "Positive", "Medium",
  ""
),

# ─── TC_CHECKOUT_007 — Valid Pin Code ──────────────────────────────────────
(
  "TC_CHECKOUT_007",
  "Checkout Page",
  "Shipping Address — Enter valid 6-digit Pin Code",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Pin Code' field\n"
  "3. Enter: '682021'\n"
  "4. Tab out of field",
  "Pin Code accepted. No error. City/State may auto-populate.",
  "Positive", "High",
  "Mapped from checklist: Valid postal code / ZIP."
),

# ─── TC_CHECKOUT_008 — Valid City ──────────────────────────────────────────
(
  "TC_CHECKOUT_008",
  "Checkout Page",
  "Shipping Address — Enter valid City name",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'City' field\n"
  "3. Enter: 'Kochi'\n"
  "4. Tab out",
  "City accepted without error.",
  "Positive", "Medium",
  ""
),

# ─── TC_CHECKOUT_009 — Valid State ─────────────────────────────────────────
(
  "TC_CHECKOUT_009",
  "Checkout Page",
  "Shipping Address — Select valid State from dropdown",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click 'State' dropdown\n"
  "3. Select 'Kerala'\n"
  "4. Observe field updates",
  "State dropdown updates to 'Kerala'. No error.",
  "Positive", "Medium",
  "Mapped from checklist: Dropdown validation."
),

# ─── TC_CHECKOUT_010 — Default Country ─────────────────────────────────────
(
  "TC_CHECKOUT_010",
  "Checkout Page",
  "Shipping Address — Verify Country defaults to India",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Country' dropdown/field\n"
  "3. Observe default value",
  "Country field defaults to 'India'. User can change if needed.",
  "Positive", "Low",
  "Mapped from checklist: Dropdown default selection."
),

# ─── TC_CHECKOUT_011 — Billing same as Shipping checkbox ───────────────────
(
  "TC_CHECKOUT_011",
  "Checkout Page",
  "Verify 'Use this address as billing' checkbox is checked by default",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Scroll to 'Use this address as my billing address' checkbox\n"
  "3. Observe its checked state",
  "Checkbox is checked by default. Separate billing address form is hidden.",
  "Positive", "Medium",
  "Mapped from checklist: Checkbox/radio validation."
),

# ─── TC_CHECKOUT_012 — Uncheck billing checkbox → billing form appears ──────
(
  "TC_CHECKOUT_012",
  "Checkout Page",
  "Uncheck 'Use same billing address' — verify separate billing form appears",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Uncheck 'Use this address as my billing address' checkbox\n"
  "3. Observe UI change",
  "Separate Billing Address form becomes visible with its own required fields.",
  "Positive", "High",
  ""
),

# ─── TC_CHECKOUT_013 — Apply valid coupon code ──────────────────────────────
(
  "TC_CHECKOUT_013",
  "Checkout Page",
  "Apply a valid coupon code — verify discount applied to order total",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Coupon Code' input field\n"
  "3. Enter a valid coupon code\n"
  "4. Click 'Apply' button\n"
  "5. Observe Order Summary",
  "Coupon applied successfully. Coupon Discount line shows deduction. Order total is recalculated.",
  "Positive", "High",
  ""
),

# ─── TC_CHECKOUT_014 — Apply valid gift card ────────────────────────────────
(
  "TC_CHECKOUT_014",
  "Checkout Page",
  "Apply a valid Gift Card — verify redemption amount deducted from total",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Locate 'Gift Card' input field\n"
  "3. Enter a valid gift card number\n"
  "4. Click 'Apply' button\n"
  "5. Observe Order Summary",
  "Gift card applied. 'Redeemed Amount' line shown in summary. Total reduced accordingly.",
  "Positive", "High",
  ""
),

# ─── TC_CHECKOUT_015 — Select Cash on Delivery ──────────────────────────────
(
  "TC_CHECKOUT_015",
  "Checkout Page",
  "Select 'Cash on Delivery' payment method",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Scroll to Payment section\n"
  "3. Select 'Cash on Delivery' radio button\n"
  "4. Observe UI",
  "COD radio selected. No card detail fields appear. PAY NOW button is enabled.",
  "Positive", "Critical",
  "Mapped from checklist: Payment method selection."
),

# ─── TC_CHECKOUT_016 — Select Pay Online (Razorpay) ─────────────────────────
(
  "TC_CHECKOUT_016",
  "Checkout Page",
  "Select 'Pay Online' payment method — Razorpay gateway loads",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Select 'Pay Online' radio button\n"
  "3. Fill all shipping address fields with valid data\n"
  "4. Click 'PAY NOW'\n"
  "5. Observe Razorpay modal",
  "Razorpay payment modal/popup opens with correct order amount. Card/UPI/wallet options visible.",
  "Positive", "Critical",
  "Mapped from checklist: Payment gateway integration."
),

# ─── TC_CHECKOUT_017 — Successful COD order placement ───────────────────────
(
  "TC_CHECKOUT_017",
  "Checkout Page",
  "Complete order successfully with Cash on Delivery",
  "User is logged in; Checkout page has 2 items; all address fields valid",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Fill all required address fields with valid data\n"
  "3. Select 'Cash on Delivery'\n"
  "4. Click 'PAY NOW'\n"
  "5. Observe confirmation",
  "Order placed successfully. Confirmation page/message shown with order ID. Cart is cleared.",
  "Positive", "Critical",
  "Happy path end-to-end order placement."
),

# ─── TC_CHECKOUT_018 — Order summary product details ────────────────────────
(
  "TC_CHECKOUT_018",
  "Checkout Page",
  "Verify Order Summary displays correct product names, SKU, quantity and price",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Scroll to Order Summary section\n"
  "3. Verify each item: product name, SKU, quantity, unit price",
  "All products in cart appear in Order Summary with correct name, SKU, qty, and price matching cart.",
  "Positive", "High",
  "Mapped from checklist: Order summary accuracy."
),

# ─── TC_CHECKOUT_019 — Subtotal calculation ─────────────────────────────────
(
  "TC_CHECKOUT_019",
  "Checkout Page",
  "Verify Subtotal = sum of (price × qty) for all items in Order Summary",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Note each item's price and quantity in Order Summary\n"
  "3. Calculate expected subtotal manually\n"
  "4. Compare with displayed Subtotal",
  "Displayed Subtotal equals sum of (price × qty) for all items. No rounding errors.",
  "Positive", "High",
  "Financial accuracy test."
),

# ─── TC_CHECKOUT_020 — Total calculation with service charges ────────────────
(
  "TC_CHECKOUT_020",
  "Checkout Page",
  "Verify Total = Subtotal + Service Charges − Coupon Discount − Redeemed Amount",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Note Subtotal, Service Charges (if any), discounts in Order Summary\n"
  "3. Calculate expected Total\n"
  "4. Compare with displayed Total",
  "Displayed Total matches the calculated value. No discrepancy.",
  "Positive", "Critical",
  "Financial accuracy — payment total must be exact."
),

# ─── TC_CHECKOUT_021 — First Name BVA max (56 chars) ────────────────────────
(
  "TC_CHECKOUT_021",
  "Checkout Page",
  "First Name BVA — Enter exactly 56 characters (maximum allowed)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. In 'First Name' field, enter 56 alphabetic characters: 'Abcdefghijklmnopqrstuvwxyzabcdefghijklmnopqrstuvwxyzabcd'\n"
  "3. Tab out",
  "Field accepts 56-char value. No error shown. Field does not truncate input.",
  "Positive", "Medium",
  "BVA max boundary. Mapped from checklist: Text field max 56 chars."
),

# ─────────── NEGATIVE TEST CASES ────────────────────────────────────────────

# ─── TC_CHECKOUT_022 — Submit empty form ────────────────────────────────────
(
  "TC_CHECKOUT_022",
  "Checkout Page",
  "Submit Checkout with all Shipping Address fields empty",
  "User is logged in and on Checkout page with 2 items in cart; all fields are blank",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Ensure all address fields are empty\n"
  "3. Click 'PAY NOW' button",
  "Validation errors appear for all 9 required fields (First Name, Last Name, Email, Phone, Address, Pin Code, City, State, Country). Form is NOT submitted.",
  "Negative", "Critical",
  "Mapped from checklist: Mandatory field validation."
),

# ─── TC_CHECKOUT_023 — Required field: First Name empty ─────────────────────
(
  "TC_CHECKOUT_023",
  "Checkout Page",
  "Leave First Name empty — verify field-level error on tab-out",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click into 'First Name' field, do not enter any value\n"
  "3. Tab out to next field",
  "Inline error message appears near First Name field (e.g., 'First name is required'). Error class 'errorText' becomes visible.",
  "Negative", "High",
  "Mapped from checklist: Mandatory field validation + error message near field."
),

# ─── TC_CHECKOUT_024 — Invalid Email format ─────────────────────────────────
(
  "TC_CHECKOUT_024",
  "Checkout Page",
  "Enter invalid Email format (missing @ symbol)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'invalidemail.com' in Email Address field\n"
  "3. Tab out",
  "Validation error shown: 'Invalid email address' or similar. Field highlighted. Form not submittable.",
  "Negative", "High",
  "Mapped from checklist: Email invalid format."
),

# ─── TC_CHECKOUT_025 — Email with leading/trailing spaces ───────────────────
(
  "TC_CHECKOUT_025",
  "Checkout Page",
  "Enter Email with leading and trailing spaces",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '  test@example.com  ' (with spaces) in Email field\n"
  "3. Tab out",
  "Email is trimmed and accepted, OR validation error is shown. Spaces should not be submitted as part of email.",
  "Negative", "Medium",
  "Mapped from checklist: Email whitespace handling."
),

# ─── TC_CHECKOUT_026 — Phone: less than 10 digits ───────────────────────────
(
  "TC_CHECKOUT_026",
  "Checkout Page",
  "Enter Phone Number with fewer than 10 digits",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '98765' in Phone Number field\n"
  "3. Tab out",
  "Validation error: 'Please enter a valid 10-digit phone number' or similar. Form not submittable.",
  "Negative", "High",
  "Mapped from checklist: Phone number invalid length. BVA: min-1."
),

# ─── TC_CHECKOUT_027 — Phone: more than 10 digits ───────────────────────────
(
  "TC_CHECKOUT_027",
  "Checkout Page",
  "Enter Phone Number with more than 10 digits",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '98765432101234' in Phone Number field\n"
  "3. Tab out",
  "Validation error shown for invalid phone length OR field restricts to 10 digits. Form not submittable.",
  "Negative", "High",
  "Mapped from checklist: Phone number invalid length. BVA: max+1."
),

# ─── TC_CHECKOUT_028 — Phone with special characters ────────────────────────
(
  "TC_CHECKOUT_028",
  "Checkout Page",
  "Enter Phone Number with special characters",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '98765@#$!0' in Phone Number field\n"
  "3. Tab out",
  "Validation error: phone must contain digits only. Special chars rejected or field shows error.",
  "Negative", "High",
  "Mapped from checklist: Phone field special characters."
),

# ─── TC_CHECKOUT_029 — Invalid Pin Code (alphabets) ─────────────────────────
(
  "TC_CHECKOUT_029",
  "Checkout Page",
  "Enter Pin Code with alphabetic characters",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'ABCDEF' in Pin Code field\n"
  "3. Tab out",
  "Validation error: 'Enter a valid 6-digit pin code' or similar. Alphabets not accepted.",
  "Negative", "High",
  "Mapped from checklist: Postal code / ZIP invalid format."
),

# ─── TC_CHECKOUT_030 — First Name with numbers ──────────────────────────────
(
  "TC_CHECKOUT_030",
  "Checkout Page",
  "Enter First Name containing numeric digits",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'Sree123' in First Name field\n"
  "3. Tab out",
  "Validation error: 'Name must contain alphabets only' or similar. Numbers not accepted.",
  "Negative", "Medium",
  "Mapped from checklist: Alpha-only text field."
),

# ─── TC_CHECKOUT_031 — Apply invalid coupon code ────────────────────────────
(
  "TC_CHECKOUT_031",
  "Checkout Page",
  "Apply an invalid/expired Coupon Code",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'INVALID999' in Coupon Code field\n"
  "3. Click 'Apply'",
  "Error message displayed: 'Invalid coupon code' or 'Coupon expired'. Discount NOT applied to total.",
  "Negative", "High",
  "Mapped from checklist: Error message for invalid input."
),

# ─── TC_CHECKOUT_032 — Apply invalid gift card ──────────────────────────────
(
  "TC_CHECKOUT_032",
  "Checkout Page",
  "Apply an invalid Gift Card number",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'FAKEGIFTCARD123' in Gift Card field\n"
  "3. Click 'Apply'",
  "Error message: 'Invalid gift card' or 'Gift card not found'. No deduction from total.",
  "Negative", "High",
  ""
),

# ─── TC_CHECKOUT_033 — Empty coupon code submit ──────────────────────────────
(
  "TC_CHECKOUT_033",
  "Checkout Page",
  "Click 'Apply' for Coupon Code with empty input",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Leave Coupon Code field empty\n"
  "3. Click 'Apply'",
  "Error or prompt: 'Please enter a coupon code'. Apply action does not proceed with empty field.",
  "Negative", "Medium",
  ""
),

# ─── TC_CHECKOUT_034 — Unauthenticated direct checkout URL access ─────────────
(
  "TC_CHECKOUT_034",
  "Checkout Page",
  "Access Checkout page directly without being logged in",
  "User is NOT logged in (fresh browser / guest session)",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Observe page response",
  "User is redirected to Login page or an 'Unauthorized' message is shown. Checkout page is NOT accessible without authentication.",
  "Negative", "Critical",
  "Mapped from checklist: OWASP — Unauthenticated direct URL access. Security test."
),

# ─────────── EDGE CASE TEST CASES ───────────────────────────────────────────

# ─── TC_CHECKOUT_035 — First Name BVA: min-1 (1 char) ───────────────────────
(
  "TC_CHECKOUT_035",
  "Checkout Page",
  "First Name BVA — Enter 1 character (below minimum of 2)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'A' in First Name field\n"
  "3. Tab out",
  "Validation error shown: 'First name must be at least 2 characters'. Single char rejected.",
  "Edge Case", "Medium",
  "BVA: min-1. Mapped from checklist: Min 2 chars for First Name."
),

# ─── TC_CHECKOUT_036 — First Name BVA: min (2 chars) ────────────────────────
(
  "TC_CHECKOUT_036",
  "Checkout Page",
  "First Name BVA — Enter exactly 2 characters (minimum allowed)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'Jo' in First Name field\n"
  "3. Tab out",
  "Field accepts 2-character name without error. BVA min boundary is valid.",
  "Edge Case", "Medium",
  "BVA: min boundary. Mapped from checklist: min 2 chars."
),

# ─── TC_CHECKOUT_037 — First Name BVA: max+1 (57 chars) ─────────────────────
(
  "TC_CHECKOUT_037",
  "Checkout Page",
  "First Name BVA — Enter 57 characters (one above maximum of 56)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter a 57-character string in First Name field\n"
  "3. Tab out",
  "Validation error shown OR field truncates to 56 chars. 57-char input is rejected or clipped.",
  "Edge Case", "Medium",
  "BVA: max+1. Mapped from checklist: max 56 chars text field."
),

# ─── TC_CHECKOUT_038 — Phone BVA: 9 digits (min-1) ──────────────────────────
(
  "TC_CHECKOUT_038",
  "Checkout Page",
  "Phone Number BVA — Enter 9 digits (one below minimum valid length)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '987654321' (9 digits) in Phone Number field\n"
  "3. Tab out",
  "Validation error: invalid phone number. 9-digit input rejected.",
  "Edge Case", "High",
  "BVA: min-1 for phone. Mapped from checklist: phone length validation."
),

# ─── TC_CHECKOUT_039 — Phone with country code ──────────────────────────────
(
  "TC_CHECKOUT_039",
  "Checkout Page",
  "Enter Phone Number with +91 country code prefix",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '+919876543210' in Phone Number field\n"
  "3. Tab out",
  "Field either accepts +91 prefix (13 chars total) or shows validation error if only 10 digits expected.",
  "Edge Case", "Medium",
  "Mapped from checklist: Phone with country code."
),

# ─── TC_CHECKOUT_040 — Pin Code: 5 digits (below min) ───────────────────────
(
  "TC_CHECKOUT_040",
  "Checkout Page",
  "Pin Code BVA — Enter 5 digits (below 6-digit minimum)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '68202' in Pin Code field\n"
  "3. Tab out",
  "Validation error: 'Enter valid 6-digit pin code'. 5-digit input rejected.",
  "Edge Case", "High",
  "BVA: min-1 for PIN code."
),

# ─── TC_CHECKOUT_041 — XSS injection in First Name ──────────────────────────
(
  "TC_CHECKOUT_041",
  "Checkout Page",
  "OWASP — XSS injection in First Name field",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '<script>alert(\"XSS\")</script>' in First Name field\n"
  "3. Tab out\n"
  "4. Attempt to submit form",
  "Script tag is NOT executed. Input is sanitized or rejected. No alert popup appears. Error shown or input stripped.",
  "Edge Case", "Critical",
  "OWASP Top 10 — XSS (A03:2021). All text inputs must be tested."
),

# ─── TC_CHECKOUT_042 — SQL Injection in Address field ───────────────────────
(
  "TC_CHECKOUT_042",
  "Checkout Page",
  "OWASP — SQL Injection in Address field",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter \"' OR '1'='1'; DROP TABLE orders; --\" in Address field\n"
  "3. Tab out\n"
  "4. Attempt to submit form",
  "SQL injection payload is NOT executed. Input is sanitized. No database error shown. Form rejects or escapes the input.",
  "Edge Case", "Critical",
  "OWASP Top 10 — SQL Injection (A03:2021). Critical security TC."
),

# ─── TC_CHECKOUT_043 — XSS in Email field ───────────────────────────────────
(
  "TC_CHECKOUT_043",
  "Checkout Page",
  "OWASP — XSS injection attempt in Email field",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '<img src=x onerror=alert(1)>@test.com' in Email field\n"
  "3. Tab out",
  "Script/image tag is NOT executed. Email validation rejects the input. No browser alert fires.",
  "Edge Case", "Critical",
  "OWASP Top 10 — XSS via email field."
),

# ─── TC_CHECKOUT_044 — XSS in Coupon Code field ─────────────────────────────
(
  "TC_CHECKOUT_044",
  "Checkout Page",
  "OWASP — XSS injection in Coupon Code field",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter '<script>alert(\"xss\")</script>' in Coupon Code field\n"
  "3. Click 'Apply'",
  "XSS payload is rejected/sanitized. No script executes. Error message shown for invalid coupon.",
  "Edge Case", "Critical",
  "OWASP Top 10 — XSS. Input sanitization for coupon field."
),

# ─── TC_CHECKOUT_045 — Page refresh during checkout ─────────────────────────
(
  "TC_CHECKOUT_045",
  "Checkout Page",
  "Refresh browser during checkout — verify form data persistence",
  "User is logged in and on Checkout page with partially filled address form",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Fill First Name, Last Name, Email fields\n"
  "3. Press F5 / browser refresh\n"
  "4. Observe form state",
  "After refresh: user remains on checkout page. Cart items persist. Form fields may reset (acceptable) or retain values (better UX).",
  "Edge Case", "High",
  "Mapped from checklist: Page refresh state handling."
),

# ─── TC_CHECKOUT_046 — Browser back from checkout → cart ────────────────────
(
  "TC_CHECKOUT_046",
  "Checkout Page",
  "Navigate back from Checkout to Cart using browser back button",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click browser back button\n"
  "3. Observe destination page and cart state",
  "User is navigated back to Cart page. Cart items remain intact with correct quantities and prices.",
  "Edge Case", "High",
  "Mapped from checklist: Back navigation handling."
),

# ─── TC_CHECKOUT_047 — Multiple browser tabs open ───────────────────────────
(
  "TC_CHECKOUT_047",
  "Checkout Page",
  "Open Checkout page in two browser tabs simultaneously",
  "User is logged in; cart has 2 items",
  "1. Navigate to https://qa-sunnydiamonds.webc.in/checkout in Tab 1\n"
  "2. Open a new tab\n"
  "3. Navigate to https://qa-sunnydiamonds.webc.in/checkout in Tab 2\n"
  "4. Fill form in Tab 1, click PAY NOW\n"
  "5. Switch to Tab 2 and attempt payment",
  "Only one order is placed. Second tab shows 'Order already processed' or cart is empty. No duplicate orders.",
  "Edge Case", "High",
  "Mapped from checklist: Multiple tabs open."
),

# ─── TC_CHECKOUT_048 — Cart modification during checkout ────────────────────
(
  "TC_CHECKOUT_048",
  "Checkout Page",
  "Modify cart in another tab while checkout is in progress",
  "User is logged in; Tab 1 on Checkout page; Tab 2 on Cart page",
  "1. Open Checkout page in Tab 1\n"
  "2. Open Cart page in Tab 2\n"
  "3. Remove one item from cart in Tab 2\n"
  "4. Return to Tab 1 (Checkout)\n"
  "5. Attempt to complete checkout",
  "System detects cart change. Order Summary in Tab 1 updates or shows stale data warning. Order total reflects current cart state.",
  "Edge Case", "High",
  "Mapped from checklist: Cart modification during checkout session."
),

# ─── TC_CHECKOUT_049 — Session timeout during payment ───────────────────────
(
  "TC_CHECKOUT_049",
  "Checkout Page",
  "Session expires while user is on Checkout page — verify graceful handling",
  "User is logged in and on Checkout page; session will be expired (simulate)",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Wait for session to expire (or clear session cookie manually)\n"
  "3. Attempt to click 'PAY NOW'",
  "User is redirected to Login page with a message 'Your session has expired. Please log in again.' Order is NOT partially submitted.",
  "Edge Case", "Critical",
  "Mapped from checklist: Session timeout during payment. Security critical."
),

# ─── TC_CHECKOUT_050 — Browser back after payment completion ────────────────
(
  "TC_CHECKOUT_050",
  "Checkout Page",
  "Press browser back button after successful order — verify no duplicate order",
  "User has just completed an order successfully; confirmation page is shown",
  "1. Complete a COD order (see TC_CHECKOUT_017)\n"
  "2. On order confirmation page, click browser back button\n"
  "3. Observe destination\n"
  "4. If returned to checkout, attempt to click PAY NOW again",
  "Browser back either goes to cart (now empty) or checkout page shows 'Order already placed'. PAY NOW cannot submit duplicate order.",
  "Edge Case", "Critical",
  "Prevents double-order placement. Financial integrity test."
),

# ─── TC_CHECKOUT_051 — CSRF protection ──────────────────────────────────────
(
  "TC_CHECKOUT_051",
  "Checkout Page",
  "OWASP — Verify CSRF token present in checkout form submission",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Open browser DevTools → Network tab\n"
  "3. Fill all required fields\n"
  "4. Click PAY NOW\n"
  "5. Inspect the POST request payload in Network tab",
  "POST request includes a CSRF token (e.g., X-CSRF-Token header or hidden _token field). Token is unique per session.",
  "Edge Case", "Critical",
  "OWASP Top 10 — A01:2021 Broken Access Control / CSRF. Security verification."
),

# ─── TC_CHECKOUT_053 — CAPTCHA validation ────────────────────────────────────
(
  "TC_CHECKOUT_053",
  "Checkout Page",
  "Verify reCAPTCHA is present and required before checkout form submission",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Fill all required shipping address fields with valid data\n"
  "3. Select 'Cash on Delivery' payment method\n"
  "4. Without completing reCAPTCHA (if visible), click 'PAY NOW'\n"
  "5. Observe form submission response",
  "reCAPTCHA challenge is displayed on the page. Form submission is blocked if reCAPTCHA is not completed. Error or warning shown: 'Please complete the CAPTCHA verification'.",
  "Negative", "Critical",
  "Mapped from checklist: Section 13 — CAPTCHA and Bot Protection. reCAPTCHA (g-recaptcha-response) field present on checkout page."
),

# ─── TC_CHECKOUT_054 — Mandatory field asterisk (*) in red ───────────────────
(
  "TC_CHECKOUT_054",
  "Checkout Page",
  "Verify all mandatory Shipping Address fields are marked with red asterisk (*)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Observe labels for all 9 required fields: First Name, Last Name, Email Address, Phone Number, Address, Pin Code, City, State, Country\n"
  "3. Verify asterisk (*) symbol is visible next to each label\n"
  "4. Verify asterisk is displayed in red colour",
  "All 9 mandatory fields display a red asterisk (*) next to their labels. Asterisk is clearly visible and in red colour as per design standard.",
  "Positive", "Medium",
  "Mapped from checklist: Section 1 — Mandatory fields labeled with red asterisk (*)."
),

# ─── TC_CHECKOUT_055 — Email with special characters in username ──────────────
(
  "TC_CHECKOUT_055",
  "Checkout Page",
  "Enter Email Address with special characters in username (e.g., test#$%@example.com)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'test#$%@example.com' in Email Address field\n"
  "3. Tab out of field",
  "Validation error displayed: 'Invalid email address'. Special characters (#, $, %) in email username are rejected. Field is highlighted with error.",
  "Negative", "High",
  "Mapped from checklist: Section 2 — Special characters in email should be blocked. Error should indicate what is wrong."
),

# ─── TC_CHECKOUT_056 — First Name with leading spaces ────────────────────────
(
  "TC_CHECKOUT_056",
  "Checkout Page",
  "Enter First Name with leading spaces — verify leading spaces are blocked",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click into 'First Name' field\n"
  "3. Enter '   Sreejith' (3 leading spaces before name)\n"
  "4. Tab out of field",
  "Leading spaces are trimmed automatically, OR validation error shown: 'Leading spaces are not allowed'. Field should not accept or store value with leading spaces.",
  "Negative", "Medium",
  "Mapped from checklist: Section 5 (Text Fields — whitespace) + Contact us form Section 2 — Leading space should not be allowed."
),

# ─── TC_CHECKOUT_057 — Last Name: no minimum character limit ─────────────────
(
  "TC_CHECKOUT_057",
  "Checkout Page",
  "Enter Last Name with single character — verify no minimum limit enforced",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'S' in Last Name field (single character)\n"
  "3. Tab out of field",
  "Last Name field accepts a single character without error. No minimum character limit is enforced for Last Name (unlike First Name which requires min 2 chars).",
  "Edge Case", "Medium",
  "Mapped from checklist: Contact us form Section 2 — No minimum limit for Last Name."
),

# ─── TC_CHECKOUT_058 — Last Name BVA: exactly 56 characters ──────────────────
(
  "TC_CHECKOUT_058",
  "Checkout Page",
  "Last Name BVA — Enter exactly 56 characters (maximum allowed)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter a 56-character alphabetic string in Last Name field: 'Abcdefghijklmnopqrstuvwxyzabcdefghijklmnopqrstuvwxyzabcd'\n"
  "3. Tab out of field",
  "Last Name field accepts exactly 56 characters without error. BVA max boundary is valid.",
  "Positive", "Medium",
  "Mapped from checklist: Contact us form Section 2 — Maximum characters for Last Name should be 56. BVA max boundary."
),

# ─── TC_CHECKOUT_059 — Last Name BVA: 57 characters (max+1) ─────────────────
(
  "TC_CHECKOUT_059",
  "Checkout Page",
  "Last Name BVA — Enter 57 characters (one above maximum of 56)",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter a 57-character string in Last Name field\n"
  "3. Tab out of field",
  "Validation error shown OR field truncates input to 56 chars. 57-character Last Name is rejected or clipped.",
  "Edge Case", "Medium",
  "Mapped from checklist: Contact us form Section 2 — max 56 chars. BVA max+1 boundary."
),

# ─── TC_CHECKOUT_060 — PAY NOW button disabled after first click ─────────────
(
  "TC_CHECKOUT_060",
  "Checkout Page",
  "Verify PAY NOW button is disabled / shows loader after first click (prevent double submission)",
  "User is logged in; all required fields filled with valid data; COD selected",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Fill all required address fields with valid data\n"
  "3. Select 'Cash on Delivery'\n"
  "4. Click 'PAY NOW' button\n"
  "5. Immediately click 'PAY NOW' button again",
  "After first click, PAY NOW button becomes disabled or shows a loading spinner. Second click does NOT submit another order. Only one order is placed.",
  "Negative", "Critical",
  "Mapped from checklist: Contact us form Section 10 — Remove multiple clicks on submit button by disabling or adding loader. Prevents duplicate orders."
),

# ─── TC_CHECKOUT_061 — Network error handling during payment ─────────────────
(
  "TC_CHECKOUT_061",
  "Checkout Page",
  "Simulate network error during PAY NOW — verify error message displayed",
  "User is logged in; all address fields filled; COD payment selected",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Fill all required fields with valid data\n"
  "3. Select Cash on Delivery\n"
  "4. Throttle network to offline in browser DevTools (Network → Offline)\n"
  "5. Click 'PAY NOW'\n"
  "6. Observe error handling",
  "Error message displayed: 'Network error. Please check your internet connection and try again.' Order is NOT partially placed. PAY NOW button becomes re-enabled for retry.",
  "Negative", "High",
  "Mapped from checklist: Form Validations sheet — Network Issues error message."
),

# ─── TC_CHECKOUT_062 — Phone with leading/trailing whitespace ────────────────
(
  "TC_CHECKOUT_062",
  "Checkout Page",
  "Enter Phone Number with leading and trailing whitespace",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click into 'Phone Number' field\n"
  "3. Enter '  9876543210  ' (with leading and trailing spaces)\n"
  "4. Tab out of field",
  "Phone field trims whitespace and accepts '9876543210', OR validation error is shown. Whitespace is not stored as part of the phone number.",
  "Negative", "Medium",
  "Mapped from checklist: Section 11 — Phone Numbers — validate how system handles leading and trailing whitespaces."
),

# ─── TC_CHECKOUT_063 — Placeholder text visible and disappears on typing ─────
(
  "TC_CHECKOUT_063",
  "Checkout Page",
  "Verify placeholder text is visible in empty fields and disappears when user starts typing",
  "User is logged in and on Checkout page with 2 items in cart; all fields are empty",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Observe each input field (First Name, Last Name, Email, Phone, Address, Pin Code, City) in its empty state\n"
  "3. Note placeholder text in each field\n"
  "4. Click into 'First Name' field and start typing 'S'\n"
  "5. Observe placeholder behaviour",
  "Each input field displays appropriate placeholder text when empty (e.g., 'Enter first name'). Placeholder text disappears immediately when user starts typing.",
  "Positive", "Low",
  "Mapped from checklist: Section 5 — Text Fields — Placeholder text displayed correctly; disappears when user starts typing."
),

# ─── TC_CHECKOUT_064 — State dropdown sorted alphabetically ──────────────────
(
  "TC_CHECKOUT_064",
  "Checkout Page",
  "Verify State dropdown options are sorted alphabetically or in logical order",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Click the 'State' dropdown field\n"
  "3. Scroll through the list of state options\n"
  "4. Verify the order of states",
  "State dropdown contains all Indian states. Options are sorted alphabetically (A→Z) or in a consistent logical order. Disabled/default option (e.g., 'Select State') appears at the top.",
  "Positive", "Low",
  "Mapped from checklist: Section 6 — Dropdowns — Options sorted alphabetically or in logical order."
),

# ─── TC_CHECKOUT_065 — Emoji and Unicode in text fields ──────────────────────
(
  "TC_CHECKOUT_065",
  "Checkout Page",
  "Edge Case — Enter emoji and Unicode characters in First Name field",
  "User is logged in and on Checkout page with 2 items in cart",
  "1. Navigate back to https://qa-sunnydiamonds.webc.in/checkout\n"
  "2. Enter 'Sree\U0001f600th' (name with emoji) in First Name field\n"
  "3. Tab out of field\n"
  "4. Also test: Enter '\u0D38\u0D4D\u0D30\u0D40\u0D1C\u0D3F\u0D24\u0D4D\u0D24\u0D4D' (Malayalam Unicode) in First Name field",
  "Emoji characters are rejected with a validation error (alpha-only field). Unicode/non-ASCII characters in names may be accepted or rejected depending on implementation — system should handle gracefully without crash.",
  "Edge Case", "High",
  "Mapped from checklist: Section 5 — Text Fields — test for emoji and Unicode characters; system should handle correctly."
),

# ─── TC_CHECKOUT_066 — Valid Login (ALWAYS LAST) ─────────────────────────────
(
  "TC_CHECKOUT_066",
  "Checkout Page",
  "Valid Login — Verify authenticated user can access Checkout page",
  "User is NOT logged in; fresh browser context; cart is empty",
  "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n"
  "2. Enter email: sreejith.s+4@webandcrafts.com\n"
  "3. Enter password: Password\n"
  "4. Click 'Login' button\n"
  "5. Add at least 1 product to cart\n"
  "6. Navigate to Cart → click 'CHECKOUT SECURELY'\n"
  "7. Verify Checkout page loads",
  "User is authenticated. Checkout page is accessible. Shipping Address form, Order Summary, and Payment section are displayed correctly.",
  "Positive", "Critical",
  "Valid login TC — ALWAYS LAST per QA standard. Confirms full auth → checkout flow."
),

]  # end TC_DATA

# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════════════

_tmp = FILE.replace(".xlsx", "_tmp.xlsx")
shutil.copy2(FILE, _tmp)
wb = load_workbook(_tmp)

# Remove existing sheet if present
if SHEET in wb.sheetnames:
    del wb[SHEET]

ws = wb.create_sheet(SHEET)

# ── Column widths ───────────────────────────────────────────────────────────
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── ROW 1: Banner ───────────────────────────────────────────────────────────
set_banner(ws, 1,
    f"SUNNY DIAMONDS — QA TEST CASES  |  CHECKOUT PAGE  |  {URL}",
    C_BANNER, fg="FFFFFF", size=13)
ws.row_dimensions[1].height = 28

# ── ROW 2: Metadata ─────────────────────────────────────────────────────────
ws.merge_cells("A2:K2")
meta_text = (
    "Created By: Sreejith S Madavan   |   Module: Checkout Page   |   "
    "Created Date: 2026-03-26   |   Environment: QA — https://qa-sunnydiamonds.webc.in   |   "
    f"Total TCs: {len(TC_DATA)}"
)
c = ws.cell(row=2, column=1, value=meta_text)
c.fill      = fill(C_META)
c.font      = font(bold=True, color="1F3864", size=10)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
c.border    = border_thick
ws.row_dimensions[2].height = 20

# ── ROW 3: Legend ───────────────────────────────────────────────────────────
ws.merge_cells("A3:K3")
legend_text = (
    "LEGEND:  "
    "🟢 Green (E2EFDA) = Positive   |   "
    "🔴 Pink (FCE4D6) = Negative   |   "
    "🟡 Yellow (FFF2CC) = Edge Case   |   "
    "Critical priority = Dark Red text   |   "
    "TC_002 onwards: Step 1 = 'Navigate back to Checkout URL'"
)
c = ws.cell(row=3, column=1, value=legend_text)
c.fill      = fill(C_LEGEND)
c.font      = font(bold=False, color="2E75B6", size=9)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border    = border_all
ws.row_dimensions[3].height = 18

# ── ROW 4: Common Preconditions ─────────────────────────────────────────────
ws.merge_cells("A4:K4")
precon_text = (
    "COMMON PRECONDITIONS (apply to all TCs):  "
    "Application is accessible at https://qa-sunnydiamonds.webc.in  |  "
    "Browser is open and running (Chrome recommended)  |  "
    "Valid test account available: sreejith.s+4@webandcrafts.com / Password"
)
c = ws.cell(row=4, column=1, value=precon_text)
c.fill      = fill(C_PRECON)
c.font      = font(bold=True, color="7F3F00", size=9)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border    = border_all
ws.row_dimensions[4].height = 22

# ── ROW 5: Column Headers ────────────────────────────────────────────────────
HEADERS = [
    "Test Case ID", "Module Name", "Test Case Description", "Preconditions",
    "Test Steps", "Expected Result", "Actual Result", "Status",
    "Test Type", "Priority", "Remarks"
]
for col_idx, hdr in enumerate(HEADERS, start=1):
    c = ws.cell(row=5, column=col_idx, value=hdr)
    c.fill      = fill(C_HEADER)
    c.font      = font(bold=True, color="FFFFFF", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border_thick
ws.row_dimensions[5].height = 22

# ── ROW 6+ : Test Cases ──────────────────────────────────────────────────────
for row_offset, tc in enumerate(TC_DATA):
    row_num = 6 + row_offset
    tc_id, module, desc, precon, steps, expected, tc_type, priority, remarks = tc

    # Determine row background
    if tc_type == "Positive":
        bg = C_POS
    elif tc_type == "Negative":
        bg = C_NEG
    else:
        bg = C_EDGE

    style_tc_row(ws, row_num, bg)

    values = [tc_id, module, desc, precon, steps, expected, "", "", tc_type, priority, remarks]
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=col_idx, value=val)
        c.fill      = fill(bg)
        c.border    = border_all
        c.alignment = wrap_align()
        c.font      = font()

    # Priority colouring
    apply_priority_color(ws, row_num, priority)

    # Auto row height hint
    ws.row_dimensions[row_num].height = 80

# ── Freeze panes at row 6 ────────────────────────────────────────────────────
ws.freeze_panes = "A6"

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(_tmp)
saved_to = None
try:
    os.replace(_tmp, FILE)
    saved_to = FILE
except PermissionError:
    try:
        shutil.copy2(_tmp, FILE)
        os.remove(_tmp)
        saved_to = FILE
    except PermissionError:
        saved_to = _tmp

if saved_to == FILE:
    print("SUCCESS: Saved to " + FILE)
else:
    print("WARNING: Original file locked. Saved as: " + _tmp)
    print("  Close SunnyDiamonds_v2.xlsx in Excel, then rename _tmp.xlsx to overwrite it.")

print("Sheet  : " + SHEET)
print("Total  : " + str(len(TC_DATA)) + " test cases")
print("First  : TC_CHECKOUT_001 (Full Setup Flow)")
print("Last   : TC_CHECKOUT_052 (Valid Login)")
