# -*- coding: utf-8 -*-
"""
FRIDAY — Senior QA Agent
PDP Test Case Generator  v2  (Manual Testing Format — Refactored)
Target : SunnyDiamonds_v2.xlsx  →  sheet "PDP Page"  (replaces existing)
URL    : https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=45

Improvements over v1:
  • Common preconditions removed from every TC — shown once in a dedicated header row
  • No Playwright / automation code in steps or preconditions
  • Navigation: TC_PDP_001 uses full URL; all others use "Navigate back to PDP page"
  • Steps are clean, concise, human-readable — executable by manual testers
  • Preconditions per TC only capture the specific session/state requirement
"""

import shutil, os, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── File / Sheet ───────────────────────────────────────────────────────────────
FILE    = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
SHEET   = "PDP Page"
PDP_URL = "https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=45"
TODAY   = datetime.date(2026, 3, 25).strftime("%d-%B-%Y")

_tmp = FILE.replace(".xlsx", "_pdpv2_tmp.xlsx")
shutil.copy2(FILE, _tmp)
wb = load_workbook(_tmp)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# ── Palette ────────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
AMBER        = "FFF3CD"
POS_GREEN    = "E2EFDA"
NEG_RED      = "FCE4D6"
EDGE_YELLOW  = "FFF2CC"
WHITE        = "FFFFFF"
DEEP_NAVY    = "0D2B5E"
COMMON_PRE   = "E8F4FD"   # light blue tint for common-preconditions row

PRI_COLORS = {
    "Critical" : "FFD7D7",
    "High"     : "FFE5CC",
    "Medium"   : "FFFACC",
    "Low"      : "E2EFDA",
}
TYPE_FILL = {
    "Positive"  : PatternFill("solid", fgColor=POS_GREEN),
    "Negative"  : PatternFill("solid", fgColor=NEG_RED),
    "Edge Case" : PatternFill("solid", fgColor=EDGE_YELLOW),
}

thin  = Side(style="thin",   color="B0B0B0")
thick = Side(style="medium", color="1F3864")
THIN_BORDER  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

TOTAL_COLS = 11

# ── Navigation helpers (manual, human-readable) ─────────────────────────────
NAV_001  = f"1. Open the browser and navigate to:\n   {PDP_URL}\n2. Wait for the page to fully load.\n"
NAV_BACK = f"1. Navigate back to the PDP page:\n   {PDP_URL}\n"

# For TCs that navigate to a DIFFERENT URL, the first step is defined inline.

# ── Specific preconditions (common ones are in the global header row) ────────
PRE_GUEST = "User is NOT logged in (guest session)"
PRE_FRESH = "Fresh browser session — clear cookies and cache before running this test"
PRE_LOGIN = (
    "User is NOT logged in\n"
    "Valid test credentials:\n"
    "  Email   : sreejith.s+4@webandcrafts.com\n"
    "  Password: Password"
)

# ── Test Cases  TC_PDP_001 … TC_PDP_044 ────────────────────────────────────────
# Columns: (id, module, description, preconditions, steps, expected,
#            actual, status, test_type, priority, remarks)

TC = [

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Page Load & Core Display
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_001", "PDP Page",
     "Verify PDP page loads successfully for a valid product URL",
     PRE_GUEST,
     NAV_001 +
     "3. Observe page title, product name, price, images, and CTAs.",
     "• Page loads without errors (HTTP 200)\n"
     "• Product title '18 K ROSE GOLD MIA DIAMOND PENDANT' is displayed\n"
     "• Product price ₹47,419 is visible\n"
     "• Product images load without broken links\n"
     "• 'ADD TO CART' and 'BUY NOW' buttons are visible and enabled",
     "", "Pass/Fail", "Positive", "Critical",
     "Smoke TC — baseline for all PDP tests"),

    ("TC_PDP_002", "PDP Page",
     "Verify product title and SKU are displayed correctly",
     PRE_GUEST,
     NAV_BACK +
     "2. Check the product heading at the top of the page.\n"
     "3. Locate the SKU value displayed below the title.",
     "• Title reads: '18 K ROSE GOLD MIA DIAMOND PENDANT'\n"
     "• SKU: '1120821883027' is clearly visible\n"
     "• Title is prominent and legible",
     "", "Pass/Fail", "Positive", "High",
     "Product identity fields"),

    ("TC_PDP_003", "PDP Page",
     "Verify product MRP is displayed with inclusive-tax label",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the price section on the PDP.\n"
     "3. Check the price value and the accompanying tax label.",
     "• Price ₹47,419 is prominently displayed\n"
     "• Label '(MRP Inclusive of all taxes)' appears alongside the price\n"
     "• Price uses ₹ symbol with correct Indian number formatting",
     "", "Pass/Fail", "Positive", "High",
     "Price transparency — MRP label validation"),

    ("TC_PDP_004", "PDP Page",
     "Verify product rating (4.5) is displayed near the product title",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the star rating or numeric rating near the product title/SKU.\n"
     "3. Note the displayed rating value.",
     "• Rating value '(4.5)' is visible\n"
     "• Rating is positioned adjacent to the product title or SKU\n"
     "• Star icons or numeric value renders correctly",
     "", "Pass/Fail", "Positive", "Medium",
     "Social proof display"),

    ("TC_PDP_005", "PDP Page",
     "Verify breadcrumb shows 'Home > Product Name' and Home link is clickable",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the breadcrumb trail at the top of the page.\n"
     "3. Verify the breadcrumb text and links.\n"
     "4. Click the 'Home' link in the breadcrumb.",
     "• Breadcrumb reads: Home > 18 K Rose Gold Mia Diamond Pendant\n"
     "• 'Home' is a clickable hyperlink\n"
     "• Clicking 'Home' navigates to https://qa-sunnydiamonds.webc.in/",
     "", "Pass/Fail", "Positive", "Medium",
     "Breadcrumb navigation hierarchy"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Product Image Gallery
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_006", "PDP Page",
     "Verify product image gallery loads with multiple images",
     PRE_GUEST,
     NAV_BACK +
     "2. Observe the product image gallery / slider section.\n"
     "3. Count the number of available product images.\n"
     "4. Check that all images load without broken icons.",
     "• Multiple product images are loaded (31 images expected)\n"
     "• All images are clear and product-specific (no broken icons)\n"
     "• Gallery navigation controls (arrows / thumbnails) are visible",
     "", "Pass/Fail", "Positive", "High",
     "Gallery has 31 product images per analysis"),

    ("TC_PDP_007", "PDP Page",
     "Verify clicking a thumbnail updates the main product image",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the thumbnail strip in the product gallery.\n"
     "3. Click on any thumbnail other than the currently active one.\n"
     "4. Observe the main image area.",
     "• Main product image updates to match the clicked thumbnail\n"
     "• The selected thumbnail is visually highlighted (border or overlay)\n"
     "• Image transition is smooth with no errors",
     "", "Pass/Fail", "Positive", "Medium",
     "Thumbnail-to-main image sync"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Price Breakup & Fair Pricing
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_008", "PDP Page",
     "Verify 'Price Breakup' toggle expands to show itemised cost details",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'Price Breakup' button/toggle on the PDP.\n"
     "3. Click the 'Price Breakup' toggle.\n"
     "4. Observe the expanded content.",
     "• Price Breakup section expands on click\n"
     "• The following line items are visible:\n"
     "    – Metal Price       : ₹9,372\n"
     "    – Diamond Price     : ₹33,450\n"
     "    – Making Charge     : ₹3,215.46\n"
     "    – GST (3%)          : ₹1,381.124\n"
     "    – Total             : ₹47,419\n"
     "• Disclaimer about weight variation is shown",
     "", "Pass/Fail", "Positive", "High",
     "Price transparency — detailed breakup"),

    ("TC_PDP_009", "PDP Page",
     "Verify 'Price Breakup' section collapses on second click",
     PRE_GUEST,
     NAV_BACK +
     "2. Click 'Price Breakup' toggle to expand the section.\n"
     "3. Click the 'Price Breakup' toggle again to collapse.\n"
     "4. Observe the section state.",
     "• Price Breakup section collapses on the second click\n"
     "• Itemised details are hidden\n"
     "• Only the main MRP (₹47,419) remains visible",
     "", "Pass/Fail", "Positive", "Medium",
     "Toggle collapse behavior"),

    ("TC_PDP_010", "PDP Page",
     "Verify Fair Pricing section shows Sunny Diamonds price vs Estimated Retail",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'Fair Pricing' section on the PDP.\n"
     "3. Observe the two price values displayed.",
     "• 'Fair Pricing' label or 'Learn More' link is visible\n"
     "• Sunny Diamonds price: ₹47,419 is shown\n"
     "• Estimated Retail price: ₹52,160 is shown\n"
     "• Savings comparison is clearly presented",
     "", "Pass/Fail", "Positive", "Medium",
     "Competitive pricing section"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Quantity Input
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_011", "PDP Page",
     "Verify default quantity is 1 and the '+' button increments correctly",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the Quantity section on the PDP.\n"
     "3. Note the default value in the quantity field.\n"
     "4. Click the '+' (increment) button once.\n"
     "5. Note the updated quantity value.",
     "• Default quantity is 1\n"
     "• Clicking '+' changes quantity to 2\n"
     "• Quantity field updates instantly without page reload",
     "", "Pass/Fail", "Positive", "High",
     "Quantity stepper — increment"),

    ("TC_PDP_012", "PDP Page",
     "Verify the '–' button decrements quantity correctly",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the '+' button twice to set quantity to 3.\n"
     "3. Click the '–' (decrement) button once.\n"
     "4. Note the quantity value.",
     "• Quantity decreases from 3 to 2\n"
     "• Quantity field reflects the updated value\n"
     "• No page reload occurs",
     "", "Pass/Fail", "Positive", "Medium",
     "Quantity stepper — decrement"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Check Availability (Pincode)
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_013", "PDP Page",
     "Verify 'Check Availability' returns delivery info for a valid 6-digit pincode",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'Check availability' section on the PDP.\n"
     "3. Click the pincode input field.\n"
     "4. Enter a valid 6-digit pincode: 682035\n"
     "5. Click the check / submit button.\n"
     "6. Observe the availability result.",
     "• Pincode 682035 is accepted without validation error\n"
     "• Delivery availability result is displayed\n"
     "  (e.g., estimated delivery date or in-stock confirmation)\n"
     "• No server or application error occurs",
     "", "Pass/Fail", "Positive", "High",
     "Valid pincode: 682035 — Ernakulam, Kerala"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Accordions (Metal / Diamond / Manufactured By)
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_014", "PDP Page",
     "Verify 'Metal Details' accordion expands and displays correct specifications",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll to the product specifications area.\n"
     "3. Click the 'METAL DETAILS' accordion button.\n"
     "4. Read the displayed specifications.",
     "• Metal Details section expands\n"
     "• The following fields are visible:\n"
     "    – Metal Weight  : 1 gm\n"
     "    – Setting Type  : OPEN\n"
     "    – Metal Purity  : 18K\n"
     "    – Metal Type    : GOLD\n"
     "    – Metal Color   : ROSE GOLD",
     "", "Pass/Fail", "Positive", "High",
     "Metal spec accordion"),

    ("TC_PDP_015", "PDP Page",
     "Verify 'Diamond Details' accordion expands and displays diamond specifications",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the 'DIAMOND DETAILS' accordion button.\n"
     "3. Read the displayed diamond information.",
     "• Diamond Details section expands\n"
     "• Diamond specifications (cut, weight, clarity, colour) are displayed\n"
     "• Content is readable and correctly formatted\n"
     "• Section collapses when clicked again",
     "", "Pass/Fail", "Positive", "High",
     "Diamond spec accordion"),

    ("TC_PDP_016", "PDP Page",
     "Verify 'Manufactured By' accordion expands and displays manufacturer info",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the 'MANUFACTURED BY' accordion button.\n"
     "3. Read the displayed manufacturer details.",
     "• Manufactured By section expands\n"
     "• Manufacturer details are displayed\n"
     "• Section collapses when clicked again",
     "", "Pass/Fail", "Positive", "Medium",
     "Manufacturer info accordion"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Add to Cart & Buy Now
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_017", "PDP Page",
     "Verify 'ADD TO CART' button is visible and enabled on the PDP",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'ADD TO CART' button in the product action section.\n"
     "3. Check button state and visual appearance.",
     "• 'ADD TO CART' button is prominently visible\n"
     "• Button is enabled (not disabled / greyed out)\n"
     "• Button is visible in both the product section and the sticky bar\n"
     "• Visual style matches the site's design",
     "", "Pass/Fail", "Positive", "Critical",
     "Primary CTA visibility — guest mode"),

    ("TC_PDP_018", "PDP Page",
     "Verify clicking 'ADD TO CART' as a guest triggers appropriate action",
     PRE_GUEST,
     NAV_BACK +
     "2. Ensure the user is NOT logged in.\n"
     "3. Click the 'ADD TO CART' button.\n"
     "4. Observe the resulting behavior.",
     "• The system either:\n"
     "  (a) Redirects to the login page with a return URL back to PDP, OR\n"
     "  (b) Adds the item to a guest cart and shows a success notification\n"
     "• No unhandled error, blank screen, or console error occurs\n"
     "• Feedback is clearly visible to the user",
     "", "Pass/Fail", "Positive", "Critical",
     "Guest Add to Cart — login redirect or guest cart"),

    ("TC_PDP_019", "PDP Page",
     "Verify 'BUY NOW' button is visible and enabled on the PDP",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'BUY NOW' button in the product action section.\n"
     "3. Check button state and placement relative to ADD TO CART.",
     "• 'BUY NOW' button is visible and enabled\n"
     "• Button is clearly differentiated from ADD TO CART\n"
     "• Both buttons are accessible without scrolling (above the fold)",
     "", "Pass/Fail", "Positive", "Critical",
     "Secondary CTA — BUY NOW visibility"),

    ("TC_PDP_020", "PDP Page",
     "Verify clicking 'BUY NOW' as a guest triggers appropriate action",
     PRE_GUEST,
     NAV_BACK +
     "2. Ensure the user is NOT logged in.\n"
     "3. Click the 'BUY NOW' button.\n"
     "4. Observe the resulting action.",
     "• The system either:\n"
     "  (a) Redirects to the login page, OR\n"
     "  (b) Initiates the checkout flow\n"
     "• No JavaScript error or blank screen occurs\n"
     "• The user is guided clearly through the next step",
     "", "Pass/Fail", "Positive", "High",
     "Guest Buy Now flow"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — You May Also Like
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_021", "PDP Page",
     "Verify 'You May Also Like' section displays related product recommendations",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll down to the 'YOU MAY ALSO LIKE' section.\n"
     "3. Count the products displayed and check each product card.",
     "• 'YOU MAY ALSO LIKE' heading is visible\n"
     "• At least 10 related products are shown\n"
     "• Each product card shows: image, product name, price\n"
     "• Each card has an 'ADD TO CART' button",
     "", "Pass/Fail", "Positive", "Medium",
     "Related products — 10 items expected"),

    ("TC_PDP_022", "PDP Page",
     "Verify clicking a related product navigates to its PDP",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll to the 'YOU MAY ALSO LIKE' section.\n"
     "3. Click on the product name or image of 'Amaury Diamond Earring'.\n"
     "4. Observe page navigation.",
     "• Browser navigates to the PDP of 'Amaury Diamond Earring'\n"
     "• PDP loads with price ₹33,726 and correct product details\n"
     "• URL updates to the new product slug",
     "", "Pass/Fail", "Positive", "Medium",
     "Cross-sell navigation"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Social Share
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_023", "PDP Page",
     "Verify Social Share section is visible and share options are accessible",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'Social Share' / 'Share this link via' section on the PDP.\n"
     "3. Check the available sharing options.",
     "• 'Social Share' / 'Share this link via' text is visible\n"
     "• At least one share option is shown (e.g., WhatsApp, Facebook, copy link)\n"
     "• Clicking a share option opens the corresponding share dialog",
     "", "Pass/Fail", "Positive", "Low",
     "Social share widget"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Trust Badges & Policies
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_024", "PDP Page",
     "Verify all 8 'Our Promise' trust badges are displayed",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll below the 'You May Also Like' section.\n"
     "3. Locate the 'PURE. CERTIFIED. TRUSTED.' section.\n"
     "4. Count and verify each trust badge.",
     "• 'PURE. CERTIFIED. TRUSTED.' heading is visible\n"
     "• All 8 badges are displayed:\n"
     "  1. Internally Flawless Diamonds\n"
     "  2. 100% Money Back on Diamond Value\n"
     "  3. Certifications of Diamonds\n"
     "  4. BIS Hall Mark for Jewellery\n"
     "  5. Brand Assured Quality\n"
     "  6. 15 Days Return Policy\n"
     "  7. Cash On Delivery\n"
     "  8. Pan India Free Shipping",
     "", "Pass/Fail", "Positive", "Medium",
     "All 8 trust badges verified"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Newsletter
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_025", "PDP Page",
     "Verify newsletter subscription with a valid email address",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll to the footer newsletter section.\n"
     "3. Click the email input field (placeholder: 'Your email *').\n"
     "4. Enter a valid email: testuser@example.com\n"
     "5. Click the 'SUBSCRIBE' button.\n"
     "6. Observe the response.",
     "• Email is accepted without validation error\n"
     "• A success message is displayed (e.g., 'Thank you for subscribing!')\n"
     "• No server error occurs",
     "", "Pass/Fail", "Positive", "Low",
     "Newsletter happy path"),

    # ══════════════════════════════════════════════════════════════════════════
    # POSITIVE — Cookie Consent
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_026", "PDP Page",
     "Verify Cookie Consent banner appears and 'Accept All' dismisses it",
     PRE_FRESH,
     f"1. Open the browser and navigate to:\n   {PDP_URL}\n"
     "2. Observe the cookie consent banner at the bottom of the page.\n"
     "3. Click the 'Accept All' button.\n"
     "4. Observe whether the banner disappears.",
     "• Cookie banner 'We Value Your Privacy' appears on first visit\n"
     "• 'Accept All' and 'Decline' buttons are present\n"
     "• Clicking 'Accept All' dismisses the banner\n"
     "• Preference is persisted (banner does not reappear on refresh)",
     "", "Pass/Fail", "Positive", "Low",
     "GDPR/cookie consent — Accept All flow"),

    ("TC_PDP_027", "PDP Page",
     "Verify clicking 'Decline' on cookie banner dismisses it without breaking the page",
     PRE_FRESH,
     f"1. Open the browser and navigate to:\n   {PDP_URL}\n"
     "2. When the cookie banner appears, click 'Decline'.\n"
     "3. Observe page state after dismissal.",
     "• Banner is dismissed after clicking 'Decline'\n"
     "• Non-essential cookies are not set\n"
     "• Page remains fully functional (images, CTAs, navigation all work)\n"
     "• No layout shift or page error after declining",
     "", "Pass/Fail", "Positive", "Low",
     "Cookie decline — page must remain functional"),

    # ══════════════════════════════════════════════════════════════════════════
    # NEGATIVE — Pincode Validation
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_028", "PDP Page",
     "Verify error when pincode with fewer than 6 digits is submitted",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the 'Check availability' pincode input.\n"
     "3. Enter an incomplete 4-digit pincode: 6820\n"
     "4. Click the check button.\n"
     "5. Observe the error feedback.",
     "• System shows a validation error: 'Please enter a valid 6-digit pincode'\n"
     "  OR: the check button remains disabled until 6 digits are entered\n"
     "• No server request is made for an incomplete pincode\n"
     "• Pincode field is highlighted in red or shows inline error",
     "", "Pass/Fail", "Negative", "High",
     "Invalid pincode — less than 6 digits"),

    ("TC_PDP_029", "PDP Page",
     "Verify pincode field rejects non-numeric (alphabetic) input",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the pincode input field.\n"
     "3. Type alphabetic characters: ABCDEF\n"
     "4. Observe the input field behavior and any error displayed.",
     "• Field rejects or ignores non-numeric characters\n"
     "  OR: Validation error 'Please enter a numeric pincode' is displayed\n"
     "• Check button does not process the invalid input\n"
     "• Application does not crash",
     "", "Pass/Fail", "Negative", "Medium",
     "Non-numeric pincode input"),

    ("TC_PDP_030", "PDP Page",
     "Verify appropriate message when an unserviceable pincode is submitted",
     PRE_GUEST,
     NAV_BACK +
     "2. Enter a valid-format but unserviceable pincode: 110001 (New Delhi)\n"
     "3. Click the check availability button.\n"
     "4. Observe the response message.",
     "• System shows a clear 'Delivery not available' message\n"
     "  (e.g., 'Sorry, we do not deliver to this pincode')\n"
     "• No JavaScript error is thrown\n"
     "• The page does not show a blank or loading spinner indefinitely",
     "", "Pass/Fail", "Negative", "High",
     "Unserviceable area message"),

    # ══════════════════════════════════════════════════════════════════════════
    # NEGATIVE — Quantity Validation
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_031", "PDP Page",
     "Verify quantity field rejects 0 and does not allow adding 0 items to cart",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the quantity input field.\n"
     "3. Clear the field and manually type: 0\n"
     "4. Click the 'ADD TO CART' button.\n"
     "5. Observe the system response.",
     "• Quantity 0 is rejected:\n"
     "  – Field resets to 1, OR\n"
     "  – An error message 'Quantity must be at least 1' is displayed, OR\n"
     "  – ADD TO CART button is disabled when quantity is 0\n"
     "• No item with quantity 0 is added to the cart",
     "", "Pass/Fail", "Negative", "High",
     "Quantity minimum boundary — min = 1"),

    # ══════════════════════════════════════════════════════════════════════════
    # NEGATIVE — Newsletter Invalid Email
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_032", "PDP Page",
     "Verify newsletter subscription rejects an invalid email format",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll to the footer newsletter section.\n"
     "3. Enter an invalid email: notanemail\n"
     "4. Click the 'SUBSCRIBE' button.\n"
     "5. Observe the validation response.",
     "• System shows: 'Please enter a valid email address'\n"
     "  OR: HTML5 email validation prevents form submission\n"
     "• Subscription is NOT processed\n"
     "• Error message is clearly visible",
     "", "Pass/Fail", "Negative", "Medium",
     "Newsletter — invalid email format"),

    # ══════════════════════════════════════════════════════════════════════════
    # NEGATIVE — Invalid Variant / URL
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_033", "PDP Page",
     "Verify PDP handles an invalid variant_id gracefully without crashing",
     PRE_GUEST,
     "1. Navigate to the PDP with a non-existent variant_id:\n"
     "   https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant?variant_id=99999\n"
     "2. Wait for the page response.\n"
     "3. Observe the displayed content or error.",
     "• Page handles the invalid variant gracefully:\n"
     "  – Default variant is loaded, OR\n"
     "  – A 'Variant not found' message is shown\n"
     "• Application does NOT display a 500 / unhandled error\n"
     "• User is not left on a blank page",
     "", "Pass/Fail", "Negative", "Medium",
     "Invalid variant_id — graceful fallback"),

    # ══════════════════════════════════════════════════════════════════════════
    # NEGATIVE — Empty Pincode Submission
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_034", "PDP Page",
     "Verify error when check availability is clicked with an empty pincode field",
     PRE_GUEST,
     NAV_BACK +
     "2. Locate the pincode input field — ensure it is empty.\n"
     "3. Click the check / submit button without entering any value.\n"
     "4. Observe the validation response.",
     "• System shows an error: 'Please enter a pincode'\n"
     "  OR: The check button is disabled when the field is empty\n"
     "• No server call is made for an empty input\n"
     "• Error message is visible near the pincode field",
     "", "Pass/Fail", "Negative", "Medium",
     "Empty pincode field submission"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASES — Quantity BVA
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_035", "PDP Page",
     "Verify quantity field handles an extremely large value (upper boundary — BVA)",
     PRE_GUEST,
     NAV_BACK +
     "2. Clear the quantity field and type: 9999999\n"
     "3. Click 'ADD TO CART'.\n"
     "4. Observe the system response.",
     "• System handles the extreme value without crashing\n"
     "• Either: A max-quantity limit message is shown (e.g., 'Max order qty: X')\n"
     "  OR: The item is added but total price calculation remains correct\n"
     "• No JavaScript runtime error occurs",
     "", "Pass/Fail", "Edge Case", "Medium",
     "BVA: Quantity upper boundary — no max defined in HTML"),

    ("TC_PDP_036", "PDP Page",
     "Verify quantity of exactly 1 is accepted (minimum boundary value — BVA)",
     PRE_GUEST,
     NAV_BACK +
     "2. Ensure the quantity field shows 1 (default).\n"
     "3. Click 'ADD TO CART' without changing the quantity.\n"
     "4. Observe the result.",
     "• Quantity 1 is accepted without validation error\n"
     "• ADD TO CART proceeds successfully\n"
     "• Item is added to cart with quantity 1",
     "", "Pass/Fail", "Edge Case", "Medium",
     "BVA: Quantity min boundary = 1"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASES — Responsive / Viewport
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_037", "PDP Page",
     "Verify PDP layout is fully responsive on a mobile viewport (375×812)",
     PRE_GUEST,
     "1. Resize the browser window to 375 × 812 pixels (iPhone SE).\n"
     f"2. Navigate to {PDP_URL}\n"
     "3. Scroll through the full page and check each section.",
     "• Page is fully responsive — no horizontal scroll required\n"
     "• Product title, price, and CTAs are readable\n"
     "• Images resize correctly to fit the screen\n"
     "• Mobile sticky navbar (Home, Profile, Cart, Search) is visible\n"
     "• ADD TO CART and BUY NOW buttons are accessible and tappable",
     "", "Pass/Fail", "Edge Case", "High",
     "Responsive — mobile 375px viewport"),

    ("TC_PDP_038", "PDP Page",
     "Verify PDP layout is correct on a tablet viewport (768×1024)",
     PRE_GUEST,
     "1. Resize the browser window to 768 × 1024 pixels (iPad).\n"
     f"2. Navigate to {PDP_URL}\n"
     "3. Check the layout of the product image and details sections.",
     "• Page adapts to tablet layout without overflow or cut-off\n"
     "• Product image and details are correctly positioned\n"
     "• All CTAs, accordions, and sections remain accessible\n"
     "• No overlapping elements",
     "", "Pass/Fail", "Edge Case", "Medium",
     "Responsive — tablet 768px viewport"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASE — URL without variant_id
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_039", "PDP Page",
     "Verify PDP loads the default variant when no variant_id is in the URL",
     PRE_GUEST,
     "1. Navigate to the PDP URL without any query parameter:\n"
     "   https://qa-sunnydiamonds.webc.in/18-k-rose-gold-mia-diamond-pendant\n"
     "2. Wait for the page to load.\n"
     "3. Observe the product displayed.",
     "• Page loads without error\n"
     "• Default product variant is displayed\n"
     "• Product title and price are visible\n"
     "• No 404 or error page is shown",
     "", "Pass/Fail", "Edge Case", "Medium",
     "URL edge — no variant_id parameter"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASE — Pincode BVA (6 vs 7 digits)
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_040", "PDP Page",
     "Verify pincode accepts exactly 6 digits and rejects 7 digits (BVA)",
     PRE_GUEST,
     NAV_BACK +
     "2. Enter exactly 6 digits in the pincode field: 682035\n"
     "3. Click check — note the result.\n"
     "4. Clear the field and enter 7 digits: 6820350\n"
     "5. Click check — note the result.",
     "• 6-digit pincode (682035): Accepted — availability check proceeds\n"
     "• 7-digit pincode (6820350): Rejected with validation error\n"
     "  OR: Input is restricted to a maximum of 6 characters",
     "", "Pass/Fail", "Edge Case", "High",
     "BVA: Pincode — 6 digits valid, 7 digits invalid"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASE — Security (OWASP XSS)
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_041", "PDP Page",
     "Verify XSS injection in pincode field is sanitised — Security (OWASP A03)",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the pincode input field.\n"
     "3. Enter the XSS payload: <script>alert(1)</script>\n"
     "4. Click the check availability button.\n"
     "5. Observe page behavior carefully.",
     "• The XSS script is NOT executed in the browser\n"
     "• No JavaScript alert popup appears\n"
     "• Input is sanitised or escaped by the application\n"
     "• A validation error for invalid pincode format is shown",
     "", "Pass/Fail", "Edge Case", "Critical",
     "OWASP A03 — XSS in pincode input"),

    ("TC_PDP_042", "PDP Page",
     "Verify XSS payload in newsletter email field is sanitised — Security (OWASP A03)",
     PRE_GUEST,
     NAV_BACK +
     "2. Scroll to the footer newsletter section.\n"
     "3. Enter the XSS payload in the email field: <img src=x onerror=alert(1)>\n"
     "4. Click the 'SUBSCRIBE' button.\n"
     "5. Observe page behavior.",
     "• XSS script is NOT executed\n"
     "• No JavaScript alert or DOM manipulation occurs\n"
     "• Input is sanitised/stripped of HTML tags\n"
     "• Validation error for invalid email format is shown",
     "", "Pass/Fail", "Edge Case", "Critical",
     "OWASP A03 — XSS in newsletter email"),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASE — Browser Back/Forward State
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_043", "PDP Page",
     "Verify browser back navigation from a related product returns to PDP correctly",
     PRE_GUEST,
     NAV_BACK +
     "2. Click the 'Price Breakup' toggle to expand it.\n"
     "3. Click on a related product in 'You May Also Like' to navigate to its PDP.\n"
     "4. Press the browser Back button.\n"
     "5. Observe the returned page.",
     "• Browser navigates back to the original PDP\n"
     "• Product title '18 K ROSE GOLD MIA DIAMOND PENDANT' and price ₹47,419 are correct\n"
     "• Page loads without error\n"
     "• No infinite redirect or navigation loop occurs",
     "", "Pass/Fail", "Edge Case", "High",
     "Browser back — PDP state recovery"),

    # ══════════════════════════════════════════════════════════════════════════
    # VALID LOGIN — MUST BE LAST
    # ══════════════════════════════════════════════════════════════════════════

    ("TC_PDP_044", "PDP Page",
     "Login with valid credentials and verify Add to Cart works in an authenticated session",
     PRE_LOGIN,
     "1. Navigate to: https://qa-sunnydiamonds.webc.in/login\n"
     "2. Enter Email: sreejith.s+4@webandcrafts.com\n"
     "3. Enter Password: Password\n"
     "4. Click the Sign In / Login button.\n"
     "5. Verify successful login (URL changes away from /login).\n"
     f"6. Navigate to the PDP:\n   {PDP_URL}\n"
     "7. Click the 'ADD TO CART' button.\n"
     "8. Observe the cart update.",
     "• Login is successful — user is redirected to homepage or dashboard\n"
     "• PDP loads correctly for the authenticated user\n"
     "• Clicking 'ADD TO CART' adds the product to the cart\n"
     "• Cart icon/count updates to reflect the added item\n"
     "• Success notification (e.g., 'Item added to cart') is displayed\n"
     "• No login prompt or authentication error occurs",
     "", "Pass/Fail", "Positive", "Critical",
     "VALID LOGIN — PLACED LAST | Email: sreejith.s+4@webandcrafts.com"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITING
# ═══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def merged_row(ws, row_num, text, bg, font_color="FFFFFF", size=13,
               bold=True, height=26, halign="center"):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=TOTAL_COLS)
    c = ws.cell(row=row_num, column=1)
    c.value = text
    c.fill  = _fill(bg)
    c.font  = Font(name="Arial", size=size, bold=bold, color=font_color)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    ws.row_dimensions[row_num].height = height

# ── Row 1 — Project banner ────────────────────────────────────────────────────
merged_row(ws, 1,
           "🔶  SUNNY DIAMONDS — QA TEST SUITE  🔶",
           DARK_BLUE, size=14, height=30)

# ── Row 2 — Module banner ─────────────────────────────────────────────────────
merged_row(ws, 2,
           f"MODULE: PDP PAGE (Product Detail Page)  |  URL: {PDP_URL}",
           MID_BLUE, size=11, height=22)

# ── Row 3 — Metadata ──────────────────────────────────────────────────────────
merged_row(ws, 3,
           (f"Test Cases Created By: Sreejith S Madavan     |     "
            f"Created Date: {TODAY}     |     "
            f"Total TCs: {len(TC)}     |     "
            f"Positive: {sum(1 for t in TC if t[8]=='Positive')}  "
            f"Negative: {sum(1 for t in TC if t[8]=='Negative')}  "
            f"Edge Case: {sum(1 for t in TC if t[8]=='Edge Case')}     |     "
            f"Environment: QA"),
           MID_BLUE, size=9, bold=False, height=18, halign="left")

# ── Row 4 — Common Preconditions notice ───────────────────────────────────────
merged_row(ws, 4,
           ("⚠️  COMMON PRECONDITIONS (Apply to ALL test cases — not repeated per TC):   "
            "1. Application is accessible at https://qa-sunnydiamonds.webc.in     "
            "2. Browser is open and running     "
            f"3. PDP URL: {PDP_URL}"),
           COMMON_PRE, font_color="1F3864",
           size=9, bold=True, height=20, halign="left")

# ── Row 5 — Column headers ────────────────────────────────────────────────────
HEADERS = ["Test Case ID", "Module Name", "Test Case Description",
           "Preconditions", "Test Steps", "Expected Result",
           "Actual Result", "Status", "Test Type", "Priority", "Remarks"]

for col_idx, hdr in enumerate(HEADERS, 1):
    c = ws.cell(row=5, column=col_idx)
    c.value = hdr
    c.fill  = _fill(DEEP_NAVY)
    c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
ws.row_dimensions[5].height = 22

# ── Rows 6+ — Test case data ──────────────────────────────────────────────────
for offset, tc in enumerate(TC):
    row = 6 + offset
    (tc_id, module, desc, pre, steps, expected,
     actual, status, tc_type, priority, remarks) = tc

    row_fill = TYPE_FILL.get(tc_type, _fill(WHITE))

    for col_idx, val in enumerate(
            [tc_id, module, desc, pre, steps, expected,
             actual, status, tc_type, priority, remarks], 1):
        c = ws.cell(row=row, column=col_idx)
        c.value = val
        c.fill  = row_fill
        c.font  = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = THIN_BORDER

    ws.row_dimensions[row].height = 95

    # TC ID — bold
    ws.cell(row=row, column=1).font = Font(name="Arial", size=9, bold=True)

    # Module — centred
    ws.cell(row=row, column=2).alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True)

    # Status — centred
    ws.cell(row=row, column=8).alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True)

    # Test Type — centred + bold
    ws.cell(row=row, column=9).font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=row, column=9).alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True)

    # Priority — colour-coded
    ws.cell(row=row, column=10).fill = _fill(PRI_COLORS.get(priority, WHITE))
    ws.cell(row=row, column=10).font = Font(
        name="Arial", size=9, bold=True,
        color=("990000" if priority == "Critical" else "000000"))
    ws.cell(row=row, column=10).alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True)

# ── Column widths ─────────────────────────────────────────────────────────────
COL_WIDTHS = [14, 14, 44, 34, 62, 50, 18, 12, 14, 10, 38]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "C6"   # freeze below the common-preconditions row

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(_tmp)

try:
    if os.path.exists(FILE):
        os.replace(_tmp, FILE)
    else:
        shutil.copy2(_tmp, FILE)
        os.remove(_tmp)
    _saved_to = FILE
except PermissionError:
    _saved_to = _tmp

# ── Console summary ───────────────────────────────────────────────────────────
pos  = sum(1 for t in TC if t[8] == "Positive")
neg  = sum(1 for t in TC if t[8] == "Negative")
edge = sum(1 for t in TC if t[8] == "Edge Case")
last = TC[-1]

print(f"SUCCESS — '{SHEET}' sheet written to: {_saved_to}")
print(f"Total TCs  : {len(TC)}")
print(f"Positive   : {pos}")
print(f"Negative   : {neg}")
print(f"Edge Case  : {edge}")
print(f"Last TC    : {last[0]} | Type: {last[8]}")
print(f"Freeze row : C6  (below Common Preconditions notice)")
print(f"All sheets : {wb.sheetnames}")
