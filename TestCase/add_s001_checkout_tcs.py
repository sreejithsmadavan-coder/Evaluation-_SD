import shutil, os, sys, io, re, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_s001.xlsx"

GREEN  = "E2EFDA"
RED    = "FCE4D6"
YELLOW = "FFF2CC"

thin   = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border    = border
    return c

def bg_color(tc_type):
    return GREEN if tc_type == "Positive" else (RED if tc_type == "Negative" else YELLOW)

CHECKOUT_URL = "https://qa-sunnydiamonds.webc.in/checkout"
CART_URL     = "https://qa-sunnydiamonds.webc.in/cart"
NAV = f"Navigate back to {CHECKOUT_URL}"

# ─────────────────────────────────────────────────────────────────────────────
# Scenario observed in video (20.6s):
#   - Checkout page with 1 item (18K Rose Gold Eden Diamond Ring ₹24,430)
#   - User clicks '×' remove button on item in Order Summary
#   - Page redirects to Cart page (/cart)
#   - Cart page shows: "Your Cart is Empty" + empty bag icon
#   - "Item Removed from Cart" toast notification at bottom
#   - "EXPLORE PRODUCTS" button displayed
#   - Cart icon in header shows 0 items
# ─────────────────────────────────────────────────────────────────────────────

S_TCS = [
    # ── S001: Core redirect scenario ─────────────────────────────────────────
    (
        "S001",
        "Checkout - Empty Cart Redirect",
        "Remove the only item from Order Summary on Checkout page — verify page redirects to Cart page",
        "User is logged in; Checkout page is open with exactly 1 item in Order Summary (18K Rose Gold Eden Diamond Ring ₹24,430)",
        f"1. Login to https://qa-sunnydiamonds.webc.in\n"
        f"2. Add 1 product to cart and navigate to Checkout page ({CHECKOUT_URL})\n"
        f"3. Locate the item in the Order Summary section (top-right)\n"
        f"4. Click the '×' (remove/delete) button on the item in Order Summary\n"
        f"5. Observe the page behaviour after item removal",
        f"Page automatically redirects to the Cart page ({CART_URL}); URL changes to /cart; "
        f"Cart page displays 'Your Cart is Empty' with an empty bag icon; "
        f"user does not remain on Checkout page after removing the last item",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (12s–14s): Clicking '×' on last item in Order Summary triggers redirect to /cart with 'Your Cart is Empty'. Core empty-cart redirect behaviour."
    ),

    # ── S002: Toast notification ──────────────────────────────────────────────
    (
        "S002",
        "Checkout - Empty Cart Redirect",
        "Remove last item from Checkout Order Summary — verify 'Item Removed from Cart' toast notification is displayed on Cart page",
        "User is logged in; Checkout page is open with exactly 1 item in Order Summary",
        f"1. {NAV} with 1 item in cart\n"
        f"2. Click the '×' remove button on the item in Order Summary\n"
        f"3. Observe the toast notification after redirect to Cart page",
        "After redirect to Cart page, a toast notification appears at the bottom of the page with message 'Item Removed from Cart'; toast is visible for a few seconds then auto-dismisses",
        "", "Not Tested", "Positive", "High",
        "Observed in video (14s–18s): 'Item Removed from Cart' toast clearly visible at bottom of empty cart page after redirect. Mapped from checklist: error/success message validation."
    ),

    # ── S003: Empty cart UI elements ──────────────────────────────────────────
    (
        "S003",
        "Checkout - Empty Cart Redirect",
        "After redirect to empty Cart page — verify 'Your Cart is Empty' message and empty bag icon are displayed",
        "User has been redirected to Cart page after removing last item from Checkout",
        f"1. {NAV} with 1 item in cart\n"
        f"2. Click '×' on item in Order Summary\n"
        f"3. After redirect to Cart page, observe all UI elements",
        "Cart page displays:\n"
        "• Empty bag/cart icon (visual illustration)\n"
        "• Heading: 'Your Cart is Empty'\n"
        "• Sub-text: 'Looks like you haven't added anything yet. Browse our collection and find something you love.'\n"
        "• 'EXPLORE PRODUCTS' button is visible and clickable",
        "", "Not Tested", "Positive", "High",
        "Observed in video (14s–18s): Empty cart page shows bag icon, 'Your Cart is Empty' text, sub-message, and 'EXPLORE PRODUCTS' button. UI completeness check."
    ),

    # ── S004: EXPLORE PRODUCTS button navigation ──────────────────────────────
    (
        "S004",
        "Checkout - Empty Cart Redirect",
        "Click 'EXPLORE PRODUCTS' button on empty Cart page — verify navigation to Jewellery PLP",
        "User is on empty Cart page after removing last item from Checkout",
        f"1. {NAV} with 1 item in cart\n"
        f"2. Click '×' on item in Order Summary → redirected to empty Cart page\n"
        f"3. Click 'EXPLORE PRODUCTS' button\n"
        f"4. Observe navigation destination",
        "User is navigated to the Jewellery PLP page (/jewellery or /all-jewellery); product listings are displayed; URL changes to the PLP URL",
        "", "Not Tested", "Positive", "Medium",
        "Observed in video (14s): 'EXPLORE PRODUCTS' button visible on empty cart page. CTA must navigate user to product listing so they can continue shopping."
    ),

    # ── S005: Cart icon updates ───────────────────────────────────────────────
    (
        "S005",
        "Checkout - Empty Cart Redirect",
        "Remove last item from Checkout — verify cart icon in header updates to show 0 items",
        "User is logged in; Checkout page open with 1 item; cart icon shows count = 1",
        f"1. {NAV} with 1 item in cart\n"
        f"2. Note cart icon count in header (should show 1)\n"
        f"3. Click '×' remove button on item in Order Summary\n"
        f"4. After redirect to Cart page, observe cart icon in header",
        "Cart icon in the header updates to show 0 items (badge removed or shows 0); cart count reflects the empty state immediately after item removal",
        "", "Not Tested", "Positive", "High",
        "Standard cart state consistency test. Cart icon badge must update in real-time when item is removed from checkout Order Summary."
    ),

    # ── S006: Multiple items — remove one → stay on checkout ─────────────────
    (
        "S006",
        "Checkout - Empty Cart Redirect",
        "Remove one item from Order Summary when multiple items exist — verify page stays on Checkout (no redirect)",
        "User is logged in; Checkout page open with 2 or more items in Order Summary",
        f"1. Login; add 2 different products to cart\n"
        f"2. Navigate to Checkout page\n"
        f"3. In Order Summary, click '×' to remove only 1 of the 2 items\n"
        f"4. Observe page behaviour",
        "Page does NOT redirect to Cart page; user remains on Checkout page; Order Summary updates to show only the remaining item; Subtotal and Total recalculate correctly; redirect only happens when ALL items are removed",
        "", "Not Tested", "Positive", "Critical",
        "Negative test for redirect: redirect should only trigger when cart becomes empty. Removing 1 of 2 items must keep user on checkout."
    ),

    # ── S007: Remove all items one by one ────────────────────────────────────
    (
        "S007",
        "Checkout - Empty Cart Redirect",
        "Remove all items one by one from Order Summary (2 items) — verify redirect happens only after last item removed",
        "User is logged in; Checkout page open with 2 items in Order Summary",
        f"1. Login; add 2 products to cart → navigate to Checkout\n"
        f"2. Click '×' to remove the first item from Order Summary\n"
        f"3. Verify page stays on Checkout with 1 item remaining\n"
        f"4. Click '×' to remove the second (last) item\n"
        f"5. Observe page behaviour",
        "After removing first item: page stays on Checkout, Order Summary shows 1 item;\n"
        "After removing second (last) item: page immediately redirects to Cart page; 'Your Cart is Empty' displayed; 'Item Removed from Cart' toast shown",
        "", "Not Tested", "Positive", "Critical",
        "End-to-end test of sequential item removal. Redirect must trigger precisely when cart becomes empty — not before."
    ),

    # ── S008: Browser back after redirect ────────────────────────────────────
    (
        "S008",
        "Checkout - Empty Cart Redirect",
        "After redirect to empty Cart page — press browser back button — verify Checkout page does not reload with removed item",
        "User has been redirected to empty Cart page after removing last item from Checkout",
        f"1. {NAV} with 1 item in cart\n"
        f"2. Click '×' to remove item → redirected to empty Cart page\n"
        f"3. Press browser Back button\n"
        f"4. Observe what page loads and whether removed item reappears",
        "Browser Back navigates to previous page (e.g., PDP or PLP); if Checkout page loads, Order Summary must be empty (item should NOT reappear); cart remains empty — no ghost data from removed item",
        "", "Not Tested", "Edge Case", "High",
        "Edge case: browser history navigation after cart becomes empty. Removed item must not reappear on back navigation. Prevents data inconsistency."
    ),

    # ── S009: Page refresh on empty cart after redirect ───────────────────────
    (
        "S009",
        "Checkout - Empty Cart Redirect",
        "After redirect to empty Cart page — refresh browser — verify empty cart state is maintained",
        "User is on Cart page showing 'Your Cart is Empty' after redirect from Checkout",
        f"1. {NAV} with 1 item → remove item → redirected to empty Cart page\n"
        f"2. Press F5 / browser refresh\n"
        f"3. Observe page after reload",
        "Page reloads showing 'Your Cart is Empty'; cart remains empty after refresh; removed item does not reappear; 'EXPLORE PRODUCTS' button still present",
        "", "Not Tested", "Edge Case", "Medium",
        "State persistence test. Cart must remain empty after page refresh — no ghost item recovery from cache or session."
    ),
]

# ── Load workbook ──────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded 'Checkout Page' sheet — current rows: {ws.max_row}")

# ── Find TC_CHECKOUT_077 (Valid Login — must stay LAST) ───────────────────────
valid_login_row  = None
valid_login_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value and str(row[0].value).strip() == "TC_CHECKOUT_077":
        valid_login_row  = row[0].row
        valid_login_data = [c.value for c in row]
        break

if valid_login_row:
    start_row = valid_login_row
    print(f"TC_CHECKOUT_077 (Valid Login) at row {valid_login_row} — will move to last after new TCs")
else:
    start_row = ws.max_row + 1
    print(f"TC_CHECKOUT_077 not found — appending from row {start_row}")

# ── Write S TCs ───────────────────────────────────────────────────────────────
for i, tc in enumerate(S_TCS):
    r = start_row + i
    tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks = tc
    for col, val in enumerate([tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks], 1):
        write_cell(ws, r, col, val, bg_color(tc_type))
    print(f"  Written: {tc_id} at row {r}")

# ── Re-append Valid Login as absolute last ────────────────────────────────────
final_row = start_row + len(S_TCS)
if valid_login_data:
    for col, val in enumerate(valid_login_data, 1):
        write_cell(ws, final_row, col, val, bg_color("Positive"))
    print(f"\nValid Login (TC_CHECKOUT_077) re-appended at row {final_row} — LAST TC")

# ── Update metadata total TC count ────────────────────────────────────────────
total_tcs = final_row - 5
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            cell.value = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total_tcs}', cell.value)
            print(f"Metadata updated — Total TCs: {total_tcs}")

# ── Save & Replace ─────────────────────────────────────────────────────────────
wb.save(TMP)
print(f"\nSaved to: {TMP}")

for attempt in range(20):
    try:
        os.replace(TMP, FILE)
        print(f"\nSUCCESS: SunnyDiamonds_v2.xlsx updated!")
        print(f"S001–S009 added at rows {start_row}–{start_row + len(S_TCS) - 1}")
        print(f"TC_CHECKOUT_077 (Valid Login) at row {final_row} — LAST")
        print(f"Total Checkout TCs: {total_tcs}")
        break
    except PermissionError:
        print(f"File locked ({attempt+1}/20) — please close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"\nFile still locked. All TCs saved in: {TMP}")
    print("Close Excel then rename SunnyDiamonds_v2_s001.xlsx → SunnyDiamonds_v2.xlsx")
