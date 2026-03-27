import shutil, os, sys, io, re, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_g3.xlsx"

GREEN = "E2EFDA"; RED = "FCE4D6"; YELLOW = "FFF2CC"
thin  = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def wc(ws, row, col, value, fill):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=fill)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border    = border

def bg(t):
    return GREEN if t == "Positive" else (RED if t == "Negative" else YELLOW)

L   = "https://qa-sunnydiamonds.webc.in/login"
C   = "https://qa-sunnydiamonds.webc.in/checkout"
PLP = "https://qa-sunnydiamonds.webc.in/jewellery"
NAV = f"Navigate back to {C}"

# ─────────────────────────────────────────────────────────────────────────────
# Guest Checkout TCs  (ID, Module, Desc, Precond, Steps, Expected, Actual,
#                      Status, Type, Priority, Remarks)
# ─────────────────────────────────────────────────────────────────────────────
TCS = [

# ── POSITIVE ─────────────────────────────────────────────────────────────────
(
    "Guest001",
    "Checkout - Guest Checkout",
    "[Positive] Guest user clicks CHECKOUT SECURELY — verify redirect to Login page with CONTINUE AS GUEST button",
    "User is NOT logged in (guest/incognito); at least 1 product is in the cart",
    f"1. Ensure user is logged out (clear session or use incognito)\n"
    f"2. Navigate to {PLP}\n"
    f"3. Add any product to cart (e.g. Aminah Diamond Ring)\n"
    f"4. Navigate to Cart page\n"
    f"5. Click CHECKOUT SECURELY button\n"
    f"6. Observe page navigation",
    f"User is redirected to Login page ({L}); page shows: Email, Password fields, SIGN IN button, "
    f"separator, and CONTINUE AS GUEST button at the bottom",
    "", "Not Tested", "Positive", "Critical",
    "Observed in video: Guest clicking CHECKOUT SECURELY redirects to login page. "
    "CONTINUE AS GUEST button must be present for unauthenticated users."
),
(
    "Guest002",
    "Checkout - Guest Checkout",
    "[Positive] Click CONTINUE AS GUEST — verify navigation to Checkout page without login",
    "Guest user is on Login page (redirected from cart); not logged in",
    f"1. As guest, add product to cart and click CHECKOUT SECURELY\n"
    f"2. Land on Login page\n"
    f"3. Click CONTINUE AS GUEST button\n"
    f"4. Observe page navigation and URL",
    f"User navigates directly to Checkout page ({C}) without entering credentials; "
    f"full Checkout page loads with Shipping Address form, Order Summary, Payment Method; "
    f"URL is /checkout; header shows no logged-in username",
    "", "Not Tested", "Positive", "Critical",
    "Observed in video: CONTINUE AS GUEST click lands on checkout page. No credentials required for guest checkout."
),
(
    "Guest003",
    "Checkout - Guest Checkout",
    "[Positive] Guest Checkout page — verify all sections visible (Shipping Address, Order Summary, Coupon, Payment)",
    "Guest user has clicked CONTINUE AS GUEST and is on Checkout page",
    f"1. As guest, navigate to {C} via CONTINUE AS GUEST\n"
    f"2. Verify all sections on the page:\n"
    f"   a. Shipping Address form (9 mandatory fields)\n"
    f"   b. Use this address as billing address checkbox\n"
    f"   c. Coupon Code section with APPLY button\n"
    f"   d. Gift Card Redeem section with APPLY button\n"
    f"   e. Payment Method section (COD + Pay Online)\n"
    f"   f. Order Summary (product name, qty, price, total)\n"
    f"   g. PAY NOW / PLACE ORDER button",
    "All sections render correctly for guest; no login-required errors; "
    "Shipping Address form shows 9 fields all marked mandatory (*); "
    "Order Summary reflects cart items; Payment Method shows appropriate options based on cart total",
    "", "Not Tested", "Positive", "High",
    "Observed in video: Full checkout page visible to guest user. All sections must render without login session."
),
(
    "Guest004",
    "Checkout - Guest Checkout",
    "[Positive] Guest fills all mandatory Shipping Address fields with valid data — verify no validation errors",
    "Guest user is on Checkout page",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter: First Name=John, Last Name=Doe, Email=johndoe@gmail.com\n"
    f"3. Enter: Phone=9876543210, Address=123 Main Street\n"
    f"4. Enter: Pin Code=682001, City=Kochi\n"
    f"5. Select: State=Kerala (Country defaults to India)\n"
    f"6. Observe all fields accept input without errors",
    "All 9 mandatory fields accept valid input; no validation errors; "
    "form is ready for order submission; email entered is used for order confirmation",
    "", "Not Tested", "Positive", "Critical",
    "Guest must be able to fill shipping address. Email is used for confirmation — not tied to any account."
),
(
    "Guest005",
    "Checkout - Guest Checkout",
    "[Positive] Guest selects Cash on Delivery (cart < Rs.49,000) — verify PLACE ORDER button appears",
    "Guest on Checkout page; cart total < Rs.49,000",
    f"1. As guest, navigate to {C} with cart total < Rs.49,000\n"
    f"2. Fill all mandatory Shipping Address fields\n"
    f"3. Scroll to Payment Method section\n"
    f"4. Select Cash on Delivery radio button\n"
    f"5. Observe submit button label",
    "Cash on Delivery option is visible and selectable; "
    "selecting COD changes submit button label from PAY NOW to PLACE ORDER; "
    "COD is available for guest when cart < Rs.49,000 (same rule as logged-in users)",
    "", "Not Tested", "Positive", "Critical",
    "Observed in video: Guest COD selection shows PLACE ORDER. "
    "Rs.49,000 COD threshold rule applies equally to guest checkout."
),
(
    "Guest006",
    "Checkout - Guest Checkout",
    "[Positive] Guest clicks PLACE ORDER (COD) — verify Processing order modal displayed",
    "Guest on Checkout page; all mandatory fields filled; COD selected",
    f"1. As guest, complete all mandatory Shipping Address fields with valid data\n"
    f"2. Select Cash on Delivery\n"
    f"3. Click PLACE ORDER button\n"
    f"4. Observe immediate system response",
    "Processing modal appears with message: "
    "Please wait as we are processing your order, this will only take a while; "
    "loading spinner displayed; PLACE ORDER button is disabled during processing",
    "", "Not Tested", "Positive", "Critical",
    "Observed in video: Processing modal shown after guest PLACE ORDER click. Same COD flow as logged-in users."
),
(
    "Guest007",
    "Checkout - Guest Checkout",
    "[Positive] Guest COD order — verify Verify Your Mobile Number OTP modal appears",
    "Guest clicked PLACE ORDER; processing modal completed",
    f"1. As guest, complete checkout with COD and click PLACE ORDER\n"
    f"2. Wait for processing modal to finish\n"
    f"3. Observe OTP verification modal",
    "Verify Your Mobile Number OTP modal appears; "
    "modal contains: masked phone number (e.g. +91XXXXXX10) from guest-entered Phone field, "
    "4-digit OTP input boxes, Resend a new OTP in X seconds countdown timer, "
    "CONFIRM AND PLACE ORDER button",
    "", "Not Tested", "Positive", "Critical",
    "Observed in video: OTP modal triggered for guest. OTP sent to phone number entered in guest shipping form."
),
(
    "Guest008",
    "Checkout - Guest Checkout",
    "[Positive] Guest enters valid OTP and clicks CONFIRM AND PLACE ORDER — verify Thank You page shown",
    "Guest on OTP verification modal; valid OTP received on registered phone",
    f"1. As guest, complete checkout up to OTP modal\n"
    f"2. Enter valid 4-digit OTP received on the phone\n"
    f"3. Click CONFIRM AND PLACE ORDER\n"
    f"4. Observe result",
    "Order confirmed; Thank You For The Purchase! success page displayed; "
    "Confirmation Mail Has Been Sent To Your Mail Id shown; "
    "CONTINUE SHOPPING button visible; confirmation email sent to guest-provided email address",
    "", "Not Tested", "Positive", "Critical",
    "End-to-end guest COD checkout happy path. Confirmation email goes to guest email — not an account email."
),
(
    "Guest009",
    "Checkout - Guest Checkout",
    "[Positive] Guest selects Pay Online — verify PAY NOW button appears and Razorpay loads",
    "Guest on Checkout page; all mandatory fields filled",
    f"1. As guest, navigate to {C}\n"
    f"2. Fill all mandatory Shipping Address fields with valid data\n"
    f"3. Scroll to Payment Method\n"
    f"4. Select Pay Online radio button\n"
    f"5. Click PAY NOW\n"
    f"6. Observe Razorpay gateway behaviour",
    "Pay Online is selectable for guest; button shows PAY NOW; "
    "clicking PAY NOW opens Razorpay payment gateway popup; "
    "guest can complete online payment without an account",
    "", "Not Tested", "Positive", "High",
    "Guest Pay Online flow. Razorpay must function for unauthenticated/guest users."
),
(
    "Guest010",
    "Checkout - Guest Checkout",
    "[Positive] Login page Sign Up link visible during guest checkout — verify navigation to Registration page",
    "Guest is on Login page (redirected from cart/checkout)",
    f"1. As guest, add product to cart and click CHECKOUT SECURELY\n"
    f"2. Land on Login page\n"
    f"3. Click Don't have an account? Sign up link\n"
    f"4. Observe navigation",
    "User navigated to Registration/Create Account page; "
    "guest can choose to register before completing checkout; "
    "cart items are preserved after registration",
    "", "Not Tested", "Positive", "Medium",
    "Observed in video: Login page shows Sign up link as alternative flow. Guest may register instead of continuing as guest."
),

# ── NEGATIVE ─────────────────────────────────────────────────────────────────
(
    "Guest011",
    "Checkout - Guest Checkout",
    "[Negative] Guest accesses /checkout URL directly with empty cart — verify redirect to Login page",
    "User is NOT logged in; cart is empty",
    f"1. Ensure user is logged out and cart is empty\n"
    f"2. Directly type {C} in browser address bar and press Enter\n"
    f"3. Observe page response",
    f"User is redirected to Login page; checkout does not load for guest with empty cart; "
    f"CONTINUE AS GUEST option shown; no unhandled error or blank page",
    "", "Not Tested", "Negative", "High",
    "Direct URL access to /checkout without cart items must redirect to login. Prevents empty checkout sessions."
),
(
    "Guest011",
    "Checkout - Guest Checkout",
    "[Negative] Guest submits Checkout form with all mandatory fields empty — verify field-level errors shown",
    "Guest on Checkout page via CONTINUE AS GUEST; no fields filled",
    f"1. As guest, navigate to {C} via CONTINUE AS GUEST\n"
    f"2. Leave all Shipping Address fields empty\n"
    f"3. Select any payment method\n"
    f"4. Click PAY NOW / PLACE ORDER\n"
    f"5. Observe validation errors",
    "Validation errors shown near each mandatory field: "
    "First Name is required, Last Name is required, Email is required, "
    "Phone is required, Address is required, Pin Code is required, "
    "City is required, State is required; form does not submit",
    "", "Not Tested", "Negative", "Critical",
    "Mandatory field validation must apply equally to guest checkout. Same error rules as logged-in checkout."
),
(
    "Guest012",
    "Checkout - Guest Checkout",
    "[Negative] Guest enters invalid email format (missing @) — verify error message shown",
    "Guest on Checkout page",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter invalidemail (no @ symbol) in Email field\n"
    f"3. Tab out or click PAY NOW\n"
    f"4. Observe validation",
    "Error near Email field: Invalid email address or Please enter a valid email; "
    "form does not submit; guest email must be valid as it is used for order confirmation",
    "", "Not Tested", "Negative", "High",
    "Mapped from checklist Email section. Guest email is critical — used for order confirmation. Must be valid."
),
(
    "Guest013",
    "Checkout - Guest Checkout",
    "[Negative] Guest enters phone number with fewer than 10 digits — verify error shown",
    "Guest on Checkout page",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter 12345 (5 digits) in Phone Number field\n"
    f"3. Tab out or click PAY NOW\n"
    f"4. Observe validation",
    "Error near Phone field: Please enter a valid 10-digit phone number; "
    "form does not submit; OTP cannot be sent to invalid phone number",
    "", "Not Tested", "Negative", "High",
    "OTP is sent to this phone — invalid phone blocks order. Mapped from checklist Phone section."
),
(
    "Guest014",
    "Checkout - Guest Checkout",
    "[Negative] Guest enters alphabetic characters in Pin Code — verify error shown",
    "Guest on Checkout page",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter ABCDEF in Pin Code field\n"
    f"3. Tab out or click PAY NOW\n"
    f"4. Observe validation",
    "Error near Pin Code: Please enter a valid 6-digit PIN code; "
    "alphabetic input is rejected; same validation as logged-in checkout",
    "", "Not Tested", "Negative", "High",
    "Mapped from checklist Postal Code section. Guest PIN validation must match logged-in user validation."
),
(
    "Guest015",
    "Checkout - Guest Checkout",
    "[Negative] Guest enters invalid OTP in OTP modal — verify error and order NOT placed",
    "Guest on OTP verification modal after clicking PLACE ORDER (COD)",
    f"1. As guest, complete all checkout fields; select COD; click PLACE ORDER\n"
    f"2. Wait for OTP modal\n"
    f"3. Enter incorrect 4-digit OTP (e.g. 0000)\n"
    f"4. Click CONFIRM AND PLACE ORDER\n"
    f"5. Observe error handling",
    "Error message in OTP modal: Invalid OTP. Please try again; "
    "order is NOT placed; OTP fields reset or highlighted red; modal stays open",
    "", "Not Tested", "Negative", "Critical",
    "Guest OTP security. Invalid OTP must block order. Same validation as logged-in user OTP flow."
),
(
    "Guest016",
    "Checkout - Guest Checkout",
    "[Negative] Guest leaves OTP fields empty and clicks CONFIRM AND PLACE ORDER — verify validation error",
    "Guest on OTP verification modal",
    f"1. As guest, reach OTP modal\n"
    f"2. Leave all 4 OTP input boxes empty\n"
    f"3. Click CONFIRM AND PLACE ORDER\n"
    f"4. Observe validation",
    "Validation error: Please enter OTP or OTP is required; "
    "OTP fields highlighted; order is NOT placed",
    "", "Not Tested", "Negative", "High",
    "Mandatory OTP validation. Empty OTP must not allow order placement for guest users."
),
(
    "Guest017",
    "Checkout - Guest Checkout",
    "[Negative] Guest cart total > Rs.49,000 — verify Cash on Delivery option is NOT available",
    "Guest user; cart total > Rs.49,000",
    f"1. As guest, add product(s) with total > Rs.49,000 to cart\n"
    f"2. Click CHECKOUT SECURELY then CONTINUE AS GUEST\n"
    f"3. Scroll to Payment Method section\n"
    f"4. Observe available payment options",
    "Only Pay Online is shown; Cash on Delivery is absent from Payment Method section; "
    "button shows PAY NOW; Rs.49,000 COD restriction applies equally to guest checkout",
    "", "Not Tested", "Negative", "Critical",
    "Rs.49,000 COD threshold applies to guest. Mapped from SD_CHECKOUT_004. No bypass for guest users."
),
(
    "Guest018",
    "Checkout - Guest Checkout",
    "[Negative] Guest enters special characters in First Name (e.g. John@#) — verify rejection",
    "Guest on Checkout page",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter John@# in First Name field\n"
    f"3. Tab out\n"
    f"4. Observe validation",
    "First Name rejects special characters; error: First Name should contain alphabets only; "
    "same alpha-only rule applies to guest checkout form",
    "", "Not Tested", "Negative", "High",
    "Mapped from checklist Contact form — alpha-only name validation applies equally to guest checkout."
),

# ── EDGE CASES ────────────────────────────────────────────────────────────────
(
    "Guest019",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest enters email of an existing registered account — verify no silent account merge",
    "Guest on Checkout page; account exists for sreejith.s+4@webandcrafts.com",
    f"1. As guest, navigate to {C}\n"
    f"2. Enter sreejith.s+4@webandcrafts.com (registered account email) in Email field\n"
    f"3. Fill remaining fields and attempt to place order\n"
    f"4. Observe system response",
    "System either: (a) allows guest order with existing email and sends confirmation, "
    "OR (b) prompts user to sign in to existing account; "
    "must NOT silently merge or modify existing account data; no unhandled error",
    "", "Not Tested", "Edge Case", "High",
    "Security edge case: guest using a registered email. No silent account merge without user consent."
),
(
    "Guest020",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest refreshes Checkout page after partially filling address — verify form data handling",
    "Guest on Checkout page with partially filled address form",
    f"1. As guest, navigate to {C}\n"
    f"2. Fill First Name, Last Name, Email, Phone fields\n"
    f"3. Press F5 / browser refresh\n"
    f"4. Observe form state after reload",
    "After refresh: form data preserved via browser autofill or session storage, "
    "OR user notified that data will be lost; cart and Order Summary remain intact; "
    "no unhandled error or blank form crash",
    "", "Not Tested", "Edge Case", "Medium",
    "Guest session state after refresh. No account session — browser/session storage must handle form state."
),
(
    "Guest021",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest clicks browser Back from Checkout to Login page — verify CONTINUE AS GUEST still works",
    "Guest navigated to Checkout via CONTINUE AS GUEST",
    f"1. As guest, proceed to {C} via CONTINUE AS GUEST\n"
    f"2. On Checkout page, click browser Back button\n"
    f"3. Observe page navigated to\n"
    f"4. Click CONTINUE AS GUEST again",
    "Browser back returns to Login page; CONTINUE AS GUEST button still present and clickable; "
    "clicking again returns to Checkout page; cart items are preserved throughout",
    "", "Not Tested", "Edge Case", "Medium",
    "Browser back navigation from guest checkout. Login page must remain with CONTINUE AS GUEST accessible."
),
(
    "Guest022",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest removes last item from Order Summary on Checkout — verify redirect to Cart page",
    "Guest on Checkout page with exactly 1 item in cart",
    f"1. As guest, navigate to {C} via CONTINUE AS GUEST with 1 item\n"
    f"2. In Order Summary, click the X / remove button on the item\n"
    f"3. Observe page behaviour",
    "Item removed; page redirects to Cart page showing empty cart message or EXPLORE PRODUCTS button; "
    "Item Removed from Cart toast displayed; redirect behaviour for guest matches logged-in user (S001)",
    "", "Not Tested", "Edge Case", "High",
    "Mapped from S001 (empty cart redirect). Same behaviour must apply to guest. Prevents empty guest checkout."
),
(
    "Guest023",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest session expires while filling Checkout form — verify graceful handling",
    "Guest on Checkout page; session token expires during form fill",
    f"1. As guest, navigate to {C}\n"
    f"2. Begin filling address form\n"
    f"3. Leave page idle until session expires, or manually clear session cookies in DevTools\n"
    f"4. Attempt to click PAY NOW / PLACE ORDER\n"
    f"5. Observe system response",
    "System handles expired guest session gracefully; either redirects to Login page with message "
    "or re-establishes guest session automatically; no unhandled error; form data not permanently lost",
    "", "Not Tested", "Edge Case", "High",
    "Guest session timeout. Unlike logged-in users, no re-auth available. System must handle gracefully."
),
(
    "Guest024",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest resends OTP after countdown expires — verify new OTP sent and old OTP invalidated",
    "Guest on OTP modal; countdown timer reached 0",
    f"1. As guest, reach OTP modal (COD flow)\n"
    f"2. Wait for countdown timer to expire (~94 seconds)\n"
    f"3. Click Resend a new OTP link\n"
    f"4. Enter the old (expired) OTP and observe\n"
    f"5. Enter the new OTP and observe",
    "After timer expires, Resend OTP becomes clickable; new OTP sent to guest phone; "
    "timer resets; old OTP is invalidated — entering old OTP shows Invalid OTP error; "
    "new OTP allows order placement",
    "", "Not Tested", "Edge Case", "High",
    "Resend OTP for guest. Previous OTP must be invalidated. Same security behaviour as logged-in user."
),
(
    "Guest025",
    "Checkout - Guest Checkout",
    "[Edge Case] Guest opens Checkout in two browser tabs — verify no duplicate order on simultaneous submit",
    "Guest on Checkout page with same cart in two tabs",
    f"1. As guest, navigate to {C} via CONTINUE AS GUEST\n"
    f"2. Open a second tab and navigate to same {C}\n"
    f"3. Fill address and submit order from Tab 1\n"
    f"4. Without refreshing, also click PLACE ORDER from Tab 2\n"
    f"5. Observe whether duplicate orders are created",
    "Only one order is created; second submission either fails gracefully with error "
    "or is blocked by cart/session lock; no duplicate charges or orders",
    "", "Not Tested", "Edge Case", "Medium",
    "Duplicate guest order prevention. No account session makes this higher risk. Cart must be locked after first submit."
),
]

# ── Load workbook ──────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded 'Checkout Page' — current rows: {ws.max_row}")

# ── Find Valid Login TC (TC_CHECKOUT_077) — keep it LAST ──────────────────────
vl_row  = None
vl_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value and str(row[0].value).strip() == "TC_CHECKOUT_077":
        vl_row  = row[0].row
        vl_data = [c.value for c in row]
        break

if vl_row:
    print(f"TC_CHECKOUT_077 (Valid Login) at row {vl_row} — will re-append as last")
    start = vl_row
else:
    start = ws.max_row + 1
    print(f"TC_CHECKOUT_077 not found — appending from row {start}")

# ── Write Guest TCs ────────────────────────────────────────────────────────────
for i, tc in enumerate(TCS):
    r = start + i
    for col, val in enumerate(tc, 1):
        wc(ws, r, col, val, bg(tc[8]))
    print(f"  Written {tc[0]} ({tc[8]}) -> row {r}")

# ── Re-append Valid Login as absolute last ────────────────────────────────────
final = start + len(TCS)
if vl_data:
    for col, val in enumerate(vl_data, 1):
        wc(ws, final, col, val, bg("Positive"))
    print(f"TC_CHECKOUT_077 re-appended at row {final} (LAST)")

# ── Update metadata total TC count ────────────────────────────────────────────
total = final - 5
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            cell.value = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total}', cell.value)
            print(f"Metadata updated — Total TCs: {total}")

# ── Save & replace ─────────────────────────────────────────────────────────────
wb.save(TMP)
print(f"Saved to {TMP}")

for attempt in range(20):
    try:
        os.replace(TMP, FILE)
        print(f"\nSUCCESS: {FILE} updated!")
        print(f"Checkout Page — Total TCs: {total} | Last TC: TC_CHECKOUT_077 (row {final})")
        break
    except PermissionError:
        print(f"Locked ({attempt+1}/20) — please close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"\nFile still locked. Saved as: {TMP}")
    print("Close SunnyDiamonds_v2.xlsx in Excel — file will be replaced on next run.")
