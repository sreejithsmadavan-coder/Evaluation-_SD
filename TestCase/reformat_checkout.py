"""
Reformat SunnyDiamonds_v2.xlsx  —  Checkout Page sheet
Apply EXACT style from SunnyDiamonds.xlsx (the uploaded reference file)
"""
import shutil, os, sys, io, re, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_fmt.xlsx"

# ── Exact colours from reference ──────────────────────────────────────────────
DARK_BLUE   = "FF1F3864"   # banner bg, label text
MID_BLUE    = "FF2E75B6"   # header row bg
WHITE       = "FFFFFFFF"
BLACK       = "FF000000"
LABEL_BG    = "FFD9D9D9"   # metadata label cell bg
VALUE_BG    = "FFF2F2F2"   # metadata value cell bg
POS_BG      = "FFE2EFDA"   # Positive row bg
NEG_BG      = "FFFCE4D6"   # Negative row bg
EDGE_BG     = "FFFFF2CC"   # Edge Case row bg
POS_TEXT    = "FF1F6B2E"   # Positive Test Type text (green)
NEG_TEXT    = "FF7B1E0C"   # Negative Test Type text (dark red)
EDGE_TEXT   = "FF7F6000"   # Edge Case Test Type text (dark yellow)
PRI_HIGH_BG = "FFFFD7D7"   # Priority High bg
PRI_MED_BG  = "FFFFF3CD"   # Priority Medium bg
PRI_CRIT_BG = "FFFFC0C0"   # Priority Critical bg (close to High but deeper)
PRI_LOW_BG  = "FFE2EFDA"   # Priority Low bg

# ── Border helpers ─────────────────────────────────────────────────────────────
def thin():  return Side(style='thin',   color='FF000000')
def med():   return Side(style='medium', color='FF000000')

BORDER_THIN = Border(left=thin(), right=thin(), top=thin(), bottom=thin())
BORDER_HDR  = Border(left=thin(), right=thin(), top=thin(), bottom=med())

# ── Column spec (exact from reference) ────────────────────────────────────────
# Col: A=ModuleName  B=TestCaseID  C=Description  D=Preconditions  E=TestSteps
#      F=TestData    G=ExpectedResult  H=ActualResult  I=Status
#      J=TestType    K=Priority    L=Remarks
COL_WIDTHS = {
    'A': 18, 'B': 14, 'C': 38, 'D': 32, 'E': 56,
    'F': 32, 'G': 42, 'H': 18, 'I': 12, 'J': 14,
    'K': 10, 'L': 30
}
HDR_LABELS = [
    "Module Name", "Test Case ID", "Test Case Description",
    "Preconditions", "Test Steps", "Test Data",
    "Expected Result", "Actual Result", "Status",
    "Test Type", "Priority", "Remarks"
]

# ── Colour maps ────────────────────────────────────────────────────────────────
def row_bg(tc_type):
    t = str(tc_type).strip().lower()
    if "positive" in t: return POS_BG
    if "negative" in t: return NEG_BG
    return EDGE_BG

def type_color(tc_type):
    t = str(tc_type).strip().lower()
    if "positive" in t: return POS_TEXT
    if "negative" in t: return NEG_TEXT
    return EDGE_TEXT

def priority_bg(priority):
    p = str(priority).strip().lower()
    if "critical" in p: return PRI_CRIT_BG
    if "high"     in p: return PRI_HIGH_BG
    if "medium"   in p: return PRI_MED_BG
    return PRI_LOW_BG

# ── Cell writer ────────────────────────────────────────────────────────────────
def set_cell(ws, row, col, value,
             bg=None, font_color=BLACK, bold=False, size=9,
             h_align='left', v_align='top', wrap=True,
             border=BORDER_THIN):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Arial", size=size, bold=bold, color=font_color)
    c.alignment = Alignment(wrap_text=wrap, vertical=v_align, horizontal=h_align)
    c.border    = border
    return c

# ── Load workbook and read existing checkout data ──────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws_old = wb["Checkout Page"]
print(f"Reading existing Checkout Page — rows: {ws_old.max_row}, cols: {ws_old.max_column}")

# Current v2 column mapping (1-indexed):
# 1=TC_ID  2=Module  3=Description  4=Preconditions  5=TestSteps
# 6=Expected  7=Actual  8=Status  9=TestType  10=Priority  11=Remarks
# New structure adds TestData as col 6 (shift Expected→G, Actual→H …)

# Extract all TC rows (start from row 6 in old sheet)
tc_rows = []
for row in ws_old.iter_rows(min_row=6, values_only=True):
    if row[0]:   # has TC_ID
        tc_rows.append({
            'tc_id'   : row[0]  or '',
            'module'  : row[1]  or '',
            'desc'    : row[2]  or '',
            'precond' : row[3]  or '',
            'steps'   : row[4]  or '',
            'expected': row[5]  or '',
            'actual'  : row[6]  or '',
            'status'  : row[7]  or 'Not Tested',
            'type'    : row[8]  or 'Positive',
            'priority': row[9]  or 'Medium',
            'remarks' : row[10] or '',
        })

print(f"Extracted {len(tc_rows)} TC rows")

# ── Delete old sheet and create fresh one ─────────────────────────────────────
del wb["Checkout Page"]
ws = wb.create_sheet("Checkout Page", 0)   # insert at position 0 (first)

# Move to correct position (after Cart Page)
sheet_order = [
    'Registration Test Cases', 'Login Page', 'Home Page',
    'PLP Page', 'PDP Page', 'Cart Page', 'Legend & Field Info', 'Checkout Page'
]
for idx, name in enumerate(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

ws = wb["Checkout Page"]

# ── Column widths ──────────────────────────────────────────────────────────────
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# ── ROW 1 — Banner ────────────────────────────────────────────────────────────
ws.merge_cells('A1:L1')
ws.row_dimensions[1].height = 36
set_cell(ws, 1, 1,
    "SUNNY DIAMONDS \u2014 CHECKOUT PAGE QA TEST CASES",
    bg=DARK_BLUE, font_color=WHITE, bold=True, size=16,
    h_align='center', v_align='center', wrap=False, border=BORDER_THIN)

# ── ROW 2 — Metadata ──────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 22
ws.merge_cells('B2:C2')
ws.merge_cells('E2:F2')
ws.merge_cells('H2:I2')
ws.merge_cells('K2:L2')

meta = [
    (1, "Project URL:",           None),
    (2, "https://qa-sunnydiamonds.webc.in/checkout", None),
    (4, "Module:",                None),
    (5, "Checkout Page",          None),
    (7, "Created By:",            None),
    (8, "Sreejith S Madavan",     None),
    (10,"Created Date:",          None),
    (11,"26-Mar-2026",            None),
]
label_cols = {1, 4, 7, 10}
for col, val, _ in meta:
    is_label = col in label_cols
    set_cell(ws, 2, col, val,
             bg=LABEL_BG if is_label else VALUE_BG,
             font_color=DARK_BLUE,
             bold=is_label, size=10,
             h_align='left', v_align='center', wrap=True,
             border=BORDER_THIN)

# ── ROW 3 — Spacer ────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 6

# ── ROW 4 — Headers ───────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 30
for col_idx, label in enumerate(HDR_LABELS, 1):
    set_cell(ws, 4, col_idx, label,
             bg=MID_BLUE, font_color=WHITE, bold=True, size=11,
             h_align='center', v_align='center', wrap=True,
             border=BORDER_HDR)

# ── ROWS 5+ — Test Cases ──────────────────────────────────────────────────────
for i, tc in enumerate(tc_rows):
    r   = 5 + i
    ws.row_dimensions[r].height = 130
    bg  = row_bg(tc['type'])

    # A: Module Name
    set_cell(ws, r, 1, tc['module'],   bg=bg, size=9)
    # B: TC ID  (bold)
    set_cell(ws, r, 2, tc['tc_id'],    bg=bg, size=9, bold=True)
    # C: Description
    set_cell(ws, r, 3, tc['desc'],     bg=bg, size=9)
    # D: Preconditions
    set_cell(ws, r, 4, tc['precond'],  bg=bg, size=9)
    # E: Test Steps
    set_cell(ws, r, 5, tc['steps'],    bg=bg, size=9)
    # F: Test Data  (new column — blank for now, testers fill during execution)
    set_cell(ws, r, 6, "N/A",          bg=bg, size=9)
    # G: Expected Result
    set_cell(ws, r, 7, tc['expected'], bg=bg, size=9)
    # H: Actual Result
    set_cell(ws, r, 8, tc['actual'],   bg=bg, size=9)
    # I: Status
    set_cell(ws, r, 9, tc['status'],   bg=bg, size=9)
    # J: Test Type  (bold, coloured text)
    set_cell(ws, r, 10, tc['type'],
             bg=bg, font_color=type_color(tc['type']), bold=True, size=9)
    # K: Priority   (bold, coloured bg)
    set_cell(ws, r, 11, tc['priority'],
             bg=priority_bg(tc['priority']), font_color=BLACK, bold=True, size=9)
    # L: Remarks
    set_cell(ws, r, 12, tc['remarks'], bg=bg, size=9)

total_tcs = len(tc_rows)
print(f"Written {total_tcs} TC rows (rows 5 to {4 + total_tcs})")

# ── Update metadata Project URL cell to show total count ─────────────────────
# Overwrite B2 merged cell with URL + TC count info
ws.cell(row=2, column=2).value = \
    f"https://qa-sunnydiamonds.webc.in/checkout    |    Total TCs: {total_tcs}"

# ── Save and replace ──────────────────────────────────────────────────────────
wb.save(TMP)
print(f"Saved formatted file to: {TMP}")

for attempt in range(20):
    try:
        os.replace(TMP, FILE)
        print(f"\nSUCCESS: {FILE} updated with exact reference style!")
        print(f"Checkout Page: {total_tcs} TCs | 12 columns | Style matched to SunnyDiamonds.xlsx")
        break
    except PermissionError:
        print(f"File locked ({attempt+1}/20) — close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"\nFile locked. Saved as: {TMP}")
    print("Close Excel and rename manually.")
