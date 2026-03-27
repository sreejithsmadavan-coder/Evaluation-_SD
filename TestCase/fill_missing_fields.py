"""
Fill ALL missing fields in Checkout Page sheet of SunnyDiamonds_v2.xlsx
Groups:
  A — TC_003,006,008,012,013,014,032,033 : Missing Remarks
  B — TC_053 to TC_091, TC_138          : Missing Preconditions
  C — TC_113 to TC_137                  : Missing Desc/Pre/Steps/Exp/Type/Priority/Remarks
"""
import shutil, os, sys, io, time, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_fill.xlsx"

CHECKOUT = "https://qa-sunnydiamonds.webc.in/checkout"
LOGIN    = "https://qa-sunnydiamonds.webc.in/login"
PLP      = "https://qa-sunnydiamonds.webc.in/jewellery"
NAV      = f"Navigate back to {CHECKOUT}"
COMMON_PRE = (
    "User is logged in (sreejith.s+4@webandcrafts.com / Password); "
    "at least 1 product in cart; "
    f"Checkout page open at {CHECKOUT}"
)
GUEST_PRE = (
    "User is NOT logged in (guest/incognito mode); "
    "at least 1 product added to cart; "
    f"Guest navigated to {CHECKOUT} via CONTINUE AS GUEST on Login page"
)

thin   = Side(style='thin', color='FF000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def wc(ws, row, col, value, keep_bg=True):
    """Write value to cell, preserving existing fill/font style."""
    c = ws.cell(row=row, column=col)
    c.value = value
    if not c.fill or c.fill.fill_type == 'none':
        c.fill = PatternFill("solid", fgColor="FFE2EFDA")
    c.font = Font(name="Arial", size=9,
                  bold=(col == 2),           # TC_ID column bold
                  color="FF000000")
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border = BORDER

# ─────────────────────────────────────────────────────────────────────────────
# A — Remarks for TC_003, 006, 008, 012, 013, 014, 032, 033
# ─────────────────────────────────────────────────────────────────────────────
REMARKS_A = {
    "TC_CHECKOUT_003": (
        "Valid Last Name acceptance. Mapped from checklist Contact form §2: "
        "Last Name must accept alphabetic input without errors."
    ),
    "TC_CHECKOUT_006": (
        "Valid street address acceptance. Mapped from checklist §12: "
        "Address fields must accept valid alphanumeric street addresses."
    ),
    "TC_CHECKOUT_008": (
        "Valid city name acceptance. Mapped from checklist §12: "
        "City field must accept properly formed city names."
    ),
    "TC_CHECKOUT_012": (
        "Billing address checkbox toggle. Mapped from checklist §9: "
        "Unchecking 'same billing address' must reveal a separate billing form."
    ),
    "TC_CHECKOUT_013": (
        "Valid coupon code apply. Mapped from checklist §16: "
        "Discount must be applied and reflected in Order Total immediately."
    ),
    "TC_CHECKOUT_014": (
        "Valid gift card redemption. Redeemed Amount must appear in Order Summary "
        "and reduce the total payable amount correctly."
    ),
    "TC_CHECKOUT_032": (
        "Invalid gift card error handling. Mapped from checklist §16: "
        "Error message must appear near the Gift Card field; no deduction applied."
    ),
    "TC_CHECKOUT_033": (
        "Empty coupon apply validation. Mapped from checklist §1 (mandatory input): "
        "Apply button must not proceed with empty coupon field."
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# B — Preconditions for TC_053–091 and TC_138
# ─────────────────────────────────────────────────────────────────────────────
# These are the additional checklist-mapped TCs; all share the same common precondition.
B_IDS = [f"TC_CHECKOUT_0{n}" for n in range(53, 92)] + ["TC_CHECKOUT_138"]

# ─────────────────────────────────────────────────────────────────────────────
# C — Full data for TC_113–137 (Guest checkout TCs with Playwright results only)
# ─────────────────────────────────────────────────────────────────────────────
# (TC_ID, Desc, Pre, Steps, Expected, Type, Priority, Remarks)
FULL_C = {
    "TC_CHECKOUT_113": (
        "[Positive] Guest user clicks CONTINUE AS GUEST on Login page — verify navigation to Checkout page",
        "User is NOT logged in; at least 1 product in cart; user is on Login page (redirected from cart)",
        f"1. Ensure user is logged out (incognito or clear session)\n"
        f"2. Add a product to cart and click CHECKOUT SECURELY\n"
        f"3. Land on Login page\n"
        f"4. Click CONTINUE AS GUEST button\n"
        f"5. Observe navigation and URL",
        f"User navigates to {CHECKOUT} without credentials; "
        f"full Checkout page loads with Shipping Address form and Order Summary; "
        f"URL is /checkout; no login-required errors",
        "Positive", "Critical",
        "Observed in video: CONTINUE AS GUEST click lands on checkout page. Core guest entry point — no credentials required."
    ),
    "TC_CHECKOUT_114": (
        "[Positive] Guest Checkout page — verify all sections visible (Shipping Address, Order Summary, Payment)",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"2. Verify sections: Shipping Address form (9 fields), "
        f"Coupon Code, Gift Card, Payment Method (COD + Pay Online), Order Summary, PAY NOW/PLACE ORDER button",
        "All sections render correctly for guest; 9 mandatory fields present with asterisks (*); "
        "Order Summary shows cart items; Payment Method shows options based on cart total; no errors",
        "Positive", "High",
        "Observed in video: Full checkout page visible to guest. All sections must render without login session."
    ),
    "TC_CHECKOUT_115": (
        "[Positive] Guest fills all 9 mandatory Shipping Address fields with valid data — verify no errors",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter: First Name=John, Last Name=Doe, Email=johndoe@gmail.com\n"
        f"3. Enter: Phone=9876543210, Address=123 Main Street\n"
        f"4. Enter: Pin Code=682001, City=Kochi\n"
        f"5. Select: State=Kerala (Country defaults to India)\n"
        f"6. Observe all fields accept input without errors",
        "All 9 mandatory fields accept valid input; no validation errors shown; "
        "form is ready for order submission; email used for order confirmation",
        "Positive", "Critical",
        "Guest must fill shipping form. Email is used for order confirmation — not tied to any account."
    ),
    "TC_CHECKOUT_116": (
        "[Positive] Guest cart total < Rs.49,000 — verify Cash on Delivery option is available",
        f"User is NOT logged in; cart total < Rs.49,000 (e.g. Aminah Diamond Ring Rs.30,094); "
        f"guest is on {CHECKOUT} via CONTINUE AS GUEST",
        f"1. As guest, add product with price < Rs.49,000 to cart\n"
        f"2. Proceed to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"3. Scroll to Payment Method section\n"
        f"4. Observe available payment options",
        "Both Cash on Delivery and Pay Online options are visible; "
        "COD is selectable for guest when cart < Rs.49,000; "
        "selecting COD changes button label from PAY NOW to PLACE ORDER",
        "Positive", "Critical",
        "Rs.49,000 COD threshold rule applies to guest checkout. Cart < Rs.49,000 must show COD option."
    ),
    "TC_CHECKOUT_117": (
        "[Positive] Guest clicks PLACE ORDER (COD) — verify Processing order modal displayed",
        f"{GUEST_PRE}; all mandatory fields filled; Cash on Delivery selected",
        f"1. As guest, fill all mandatory Shipping Address fields with valid data\n"
        f"2. Select Cash on Delivery\n"
        f"3. Click PLACE ORDER button\n"
        f"4. Observe immediate system response",
        "Processing modal appears: 'Please wait as we are processing your order, this will only take a while'; "
        "loading spinner displayed; PLACE ORDER button is disabled during processing",
        "Positive", "Critical",
        "Observed in video: Processing modal shown for guest COD order. Same flow as logged-in users."
    ),
    "TC_CHECKOUT_118": (
        "[Positive] Guest COD order — verify Verify Your Mobile Number OTP modal appears",
        f"{GUEST_PRE}; PLACE ORDER clicked; processing modal completed",
        f"1. As guest, fill all fields; select COD; click PLACE ORDER\n"
        f"2. Wait for processing modal to complete\n"
        f"3. Observe OTP modal",
        "Verify Your Mobile Number OTP modal appears; "
        "modal shows masked phone (e.g. +91XXXXXX10); "
        "4-digit OTP input boxes; Resend a new OTP in X seconds timer; "
        "CONFIRM AND PLACE ORDER button",
        "Positive", "Critical",
        "Observed in video: OTP modal triggered for guest COD. OTP sent to phone entered in shipping form."
    ),
    "TC_CHECKOUT_119": (
        "[Positive] Guest enters valid OTP and clicks CONFIRM AND PLACE ORDER — verify Thank You page",
        f"{GUEST_PRE}; OTP verification modal is open; valid OTP received on phone",
        f"1. As guest, complete checkout up to OTP modal\n"
        f"2. Enter valid 4-digit OTP received on phone\n"
        f"3. Click CONFIRM AND PLACE ORDER\n"
        f"4. Observe result",
        "Order confirmed; Thank You For The Purchase! success page displayed; "
        "Confirmation Mail Has Been Sent To Your Mail Id shown; "
        "CONTINUE SHOPPING button visible; confirmation email sent to guest-provided email",
        "Positive", "Critical",
        "End-to-end guest COD checkout happy path. Confirmation email goes to guest email — not an account email."
    ),
    "TC_CHECKOUT_120": (
        "[Positive] Guest selects Pay Online — verify PAY NOW button appears and Razorpay gateway loads",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Fill all mandatory Shipping Address fields\n"
        f"3. Scroll to Payment Method\n"
        f"4. Select Pay Online radio button\n"
        f"5. Click PAY NOW\n"
        f"6. Observe Razorpay gateway behaviour",
        "Pay Online selectable for guest; button shows PAY NOW; "
        "clicking PAY NOW opens Razorpay payment gateway popup; "
        "guest can complete online payment without an account",
        "Positive", "High",
        "Guest Pay Online flow. Razorpay must function for unauthenticated/guest users."
    ),
    "TC_CHECKOUT_121": (
        "[Positive] Login page shows Sign Up link during guest checkout — verify navigation to Registration page",
        "User is NOT logged in; user is on Login page (redirected from cart/checkout)",
        f"1. As guest, add product to cart and click CHECKOUT SECURELY\n"
        f"2. Land on Login page\n"
        f"3. Click Don't have an account? Sign up link\n"
        f"4. Observe navigation",
        "User navigated to Registration/Create Account page; "
        "guest can choose to register before completing checkout; "
        "cart items preserved after registration",
        "Positive", "Medium",
        "Observed in video: Login page shows Sign up link. Alternative flow — guest may register instead."
    ),
    "TC_CHECKOUT_122": (
        "[Negative] Guest accesses /checkout URL directly with empty cart — verify redirect to Login page",
        "User is NOT logged in; cart is empty",
        f"1. Ensure user is logged out and cart is empty\n"
        f"2. Directly type {CHECKOUT} in browser address bar and press Enter\n"
        f"3. Observe page response",
        f"User is redirected to Login page; checkout does not load for guest with empty cart; "
        f"CONTINUE AS GUEST option shown; no unhandled error or blank page",
        "Negative", "High",
        "Direct URL access to /checkout without cart items must redirect to login. Prevents empty checkout sessions."
    ),
    "TC_CHECKOUT_123": (
        "[Negative] Guest submits Checkout form with all mandatory fields empty — verify 9 field-level errors",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"2. Leave all 9 Shipping Address fields empty\n"
        f"3. Select any payment method\n"
        f"4. Click PAY NOW / PLACE ORDER\n"
        f"5. Observe validation errors",
        "Validation errors shown near all 9 mandatory fields; "
        "errors include: First Name is required, Last Name is required, Email is required, "
        "Phone is required, Address is required, Pin Code is required, "
        "City is required, State is required; form does not submit",
        "Negative", "Critical",
        "Mandatory field validation must apply equally to guest checkout. "
        "All 9 errors must display simultaneously on empty submit."
    ),
    "TC_CHECKOUT_124": (
        "[Negative] Guest enters invalid email format (missing @) — verify error message shown",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter 'invalidemail' (no @ symbol) in Email field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error near Email field: 'Invalid email address' or 'Please enter a valid email'; "
        "form does not submit; guest email is critical — used for order confirmation",
        "Negative", "High",
        "Mapped from checklist Email section. Guest email must be valid — used for order confirmation. Must be validated."
    ),
    "TC_CHECKOUT_125": (
        "[Negative] Guest enters phone number with fewer than 10 digits — verify error shown",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter '12345' (5 digits) in Phone Number field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error near Phone field: 'Please enter a valid 10-digit phone number'; "
        "form does not submit; OTP cannot be sent to invalid phone number",
        "Negative", "High",
        "OTP is sent to this phone — invalid phone blocks COD order. Mapped from checklist Phone section."
    ),
    "TC_CHECKOUT_126": (
        "[Negative] Guest enters alphabetic characters in Pin Code — verify error shown",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter 'ABCDEF' in Pin Code field\n"
        f"3. Tab out or click PAY NOW\n"
        f"4. Observe validation",
        "Error near Pin Code: 'Please enter a valid 6-digit PIN code'; "
        "alphabetic input is rejected; same validation as logged-in checkout",
        "Negative", "High",
        "Mapped from checklist Postal Code section. Guest PIN validation must match logged-in user validation."
    ),
    "TC_CHECKOUT_127": (
        "[Negative] Guest enters invalid OTP in OTP verification modal — verify error and order NOT placed",
        f"{GUEST_PRE}; OTP verification modal is open after clicking PLACE ORDER (COD)",
        f"1. As guest, fill all fields; select COD; click PLACE ORDER\n"
        f"2. Wait for OTP modal\n"
        f"3. Enter incorrect 4-digit OTP (e.g. '0000')\n"
        f"4. Click CONFIRM AND PLACE ORDER\n"
        f"5. Observe error handling",
        "Error message in OTP modal: 'Invalid OTP. Please try again'; "
        "order is NOT placed; OTP fields reset or highlighted; modal stays open for retry",
        "Negative", "Critical",
        "Guest OTP security — invalid OTP must block order. Same validation as logged-in user OTP flow."
    ),
    "TC_CHECKOUT_128": (
        "[Negative] Guest leaves OTP fields empty and clicks CONFIRM AND PLACE ORDER — verify validation",
        f"{GUEST_PRE}; OTP verification modal is open",
        f"1. As guest, reach OTP modal\n"
        f"2. Leave all 4 OTP input boxes empty\n"
        f"3. Click CONFIRM AND PLACE ORDER\n"
        f"4. Observe validation",
        "Validation error: 'Please enter OTP' or 'OTP is required'; "
        "OTP fields highlighted; order is NOT placed",
        "Negative", "High",
        "Mandatory OTP field. Empty OTP must not allow order placement for guest users."
    ),
    "TC_CHECKOUT_129": (
        "[Negative] Guest cart total > Rs.49,000 — verify Cash on Delivery option is NOT available",
        "User is NOT logged in; cart total > Rs.49,000 (e.g. diamond jewellery item); "
        f"guest is on {CHECKOUT} via CONTINUE AS GUEST",
        f"1. As guest, add product(s) with total > Rs.49,000 to cart\n"
        f"2. Proceed to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"3. Scroll to Payment Method section\n"
        f"4. Observe available payment options",
        "Only Pay Online is shown in Payment Method; Cash on Delivery is absent; "
        "button shows PAY NOW; Rs.49,000 COD restriction applies equally to guest checkout",
        "Negative", "Critical",
        "Rs.49,000 COD threshold applies to guest. Same business rule as logged-in users. No bypass for guests."
    ),
    "TC_CHECKOUT_130": (
        "[Negative] Guest enters special characters in First Name (e.g. John@#) — verify rejection",
        GUEST_PRE,
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter 'John@#' in First Name field\n"
        f"3. Tab out\n"
        f"4. Observe validation",
        "First Name rejects special characters; error: 'First Name should contain alphabets only'; "
        "same alpha-only rule applies to guest checkout form",
        "Negative", "High",
        "Mapped from checklist Contact form — alpha-only name validation applies equally to guest checkout."
    ),
    "TC_CHECKOUT_131": (
        "[Edge Case] Guest enters email of existing registered account — verify no silent account merge",
        f"Guest is on {CHECKOUT}; account exists for sreejith.s+4@webandcrafts.com",
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Enter 'sreejith.s+4@webandcrafts.com' (existing account email) in Email field\n"
        f"3. Fill remaining fields and attempt to place order\n"
        f"4. Observe system response",
        "System either: (a) allows guest order with existing email and sends confirmation, "
        "OR (b) prompts user to sign in; "
        "must NOT silently merge or modify existing account data; no unhandled error",
        "Edge Case", "High",
        "Security edge case: guest using a registered email. No silent account merge without user consent."
    ),
    "TC_CHECKOUT_132": (
        "[Edge Case] Guest refreshes Checkout page after partially filling address — verify form data handling",
        f"{GUEST_PRE}; shipping address form is partially filled",
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Fill First Name, Last Name, Email, Phone fields\n"
        f"3. Press F5 / browser refresh\n"
        f"4. Observe form state after reload",
        "After refresh: form data preserved via browser autofill or session storage, "
        "OR user notified that data will be lost; cart and Order Summary remain intact; "
        "no unhandled error or blank page",
        "Edge Case", "Medium",
        "Guest session state after refresh. No account session — browser storage must handle form state."
    ),
    "TC_CHECKOUT_133": (
        "[Edge Case] Guest clicks browser Back from Checkout — verify return to Cart page with CHECKOUT SECURELY",
        f"Guest navigated to {CHECKOUT} via CONTINUE AS GUEST",
        f"1. As guest, proceed to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"2. Click browser Back button\n"
        f"3. Observe page navigated to\n"
        f"4. Check CHECKOUT SECURELY button availability",
        "Browser back navigates to Cart page; cart items still displayed; "
        "CHECKOUT SECURELY button still present and functional; "
        "clicking CHECKOUT SECURELY re-shows Login page with CONTINUE AS GUEST",
        "Edge Case", "Medium",
        "Browser back navigation from guest checkout. Cart must remain intact with checkout option available."
    ),
    "TC_CHECKOUT_134": (
        "[Edge Case] Guest removes last item from Order Summary — verify redirect to empty Cart page",
        f"{GUEST_PRE} with exactly 1 item in cart",
        f"1. As guest, navigate to {CHECKOUT} with 1 item\n"
        f"2. In Order Summary, click the X / remove button on the only item\n"
        f"3. Observe page behaviour",
        "Item removed; page redirects to Cart page showing empty cart or EXPLORE PRODUCTS button; "
        "Item Removed from Cart toast displayed; "
        "redirect behaviour for guest matches logged-in user (S001)",
        "Edge Case", "High",
        "Mapped from S001 — empty cart redirect. Same redirect behaviour must apply to guest users."
    ),
    "TC_CHECKOUT_135": (
        "[Edge Case] Guest session expires while filling Checkout form — verify graceful handling",
        f"{GUEST_PRE}; session token expires during form fill",
        f"1. As guest, navigate to {CHECKOUT}\n"
        f"2. Begin filling address form\n"
        f"3. Clear session cookies via DevTools (F12 > Application > Cookies > Clear All)\n"
        f"4. Attempt to click PAY NOW / PLACE ORDER\n"
        f"5. Observe system response",
        "System handles expired guest session gracefully; "
        "redirects to Login page with message OR re-establishes guest session; "
        "no HTTP 500 error or blank screen; form data not permanently lost",
        "Edge Case", "High",
        "Guest session timeout. Unlike logged-in users, no re-auth available. Must handle gracefully."
    ),
    "TC_CHECKOUT_136": (
        "[Edge Case] Guest resends OTP after countdown expires — verify new OTP sent and old OTP invalidated",
        f"{GUEST_PRE}; OTP modal is open; countdown timer has reached 0",
        f"1. As guest, reach OTP modal (COD flow)\n"
        f"2. Wait for countdown timer to expire (~94 seconds)\n"
        f"3. Click Resend a new OTP link\n"
        f"4. Enter the old (expired) OTP → observe error\n"
        f"5. Enter the new OTP → observe success",
        "After timer expires, Resend OTP becomes clickable; new OTP sent to guest phone; "
        "timer resets; old OTP returns 'Invalid OTP' error; "
        "new OTP allows order placement",
        "Edge Case", "High",
        "Resend OTP for guest. Previous OTP must be invalidated when new one is requested."
    ),
    "TC_CHECKOUT_137": (
        "[Edge Case] Guest opens Checkout in two browser tabs simultaneously — verify no duplicate order",
        f"Guest has completed shipping form and selected COD in Tab 1; same cart open in Tab 2",
        f"1. As guest, navigate to {CHECKOUT} via CONTINUE AS GUEST\n"
        f"2. Open a second tab and navigate to same {CHECKOUT}\n"
        f"3. Fill address and submit order from Tab 1\n"
        f"4. Without refreshing, also click PLACE ORDER from Tab 2\n"
        f"5. Observe whether duplicate orders are created",
        "Only one order is created; second submission fails gracefully or is blocked by session lock; "
        "no duplicate charges or orders placed",
        "Edge Case", "Medium",
        "Duplicate guest order prevention. No account session makes this higher risk. Cart must lock after submit."
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# MAIN — Load, find rows, fill data, save
# ─────────────────────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded Checkout Page — rows: {ws.max_row} cols: {ws.max_column}")

# Find header row (look for 'Test Case ID' in col 1)
hdr_row = 5
for r in range(1, 10):
    if ws.cell(r, 1).value and 'Test Case' in str(ws.cell(r, 1).value):
        hdr_row = r; break
print(f"Header row: {hdr_row} | Data starts row: {hdr_row+1}")

# ── Column index map (1-based) ────────────────────────────────────────────────
# From header row:  TC_ID=1, Module=2, Desc=3, Pre=4, Steps=5, Exp=6,
#                   Actual=7, Status=8, Type=9, Priority=10, Remarks=11
COL = dict(tc_id=1, module=2, desc=3, pre=4, steps=5, exp=6,
           actual=7, status=8, tc_type=9, priority=10, remarks=11)

filled_count = 0

for r in range(hdr_row + 1, ws.max_row + 1):
    tc_id = ws.cell(r, COL['tc_id']).value
    if not tc_id or not str(tc_id).strip():
        continue
    tc_id = str(tc_id).strip()

    # ── Group A: Missing Remarks ──────────────────────────────────────────────
    if tc_id in REMARKS_A:
        cell = ws.cell(r, COL['remarks'])
        if not cell.value or str(cell.value).strip() in ('', 'None'):
            wc(ws, r, COL['remarks'], REMARKS_A[tc_id])
            print(f"  [A] {tc_id} → Remarks filled")
            filled_count += 1

    # ── Group B: Missing Preconditions ────────────────────────────────────────
    if tc_id in B_IDS:
        cell = ws.cell(r, COL['pre'])
        if not cell.value or str(cell.value).strip() in ('', 'None'):
            wc(ws, r, COL['pre'], COMMON_PRE)
            print(f"  [B] {tc_id} → Preconditions filled")
            filled_count += 1

    # ── Group C: Full data for TC_113–137 ─────────────────────────────────────
    if tc_id in FULL_C:
        desc, pre, steps, exp, tc_type, priority, remarks = FULL_C[tc_id]

        def fill_if_empty(col_key, value):
            global filled_count
            c = ws.cell(r, COL[col_key])
            if not c.value or str(c.value).strip() in ('', 'None'):
                wc(ws, r, COL[col_key], value)
                filled_count += 1

        fill_if_empty('desc',     desc)
        fill_if_empty('pre',      pre)
        fill_if_empty('steps',    steps)
        fill_if_empty('exp',      exp)
        fill_if_empty('tc_type',  tc_type)
        fill_if_empty('priority', priority)
        fill_if_empty('remarks',  remarks)

        # Apply correct row bg colour based on type
        from openpyxl.styles import PatternFill as PF
        bg_map = {"Positive": "FFE2EFDA", "Negative": "FFFCE4D6", "Edge Case": "FFFFF2CC"}
        bg_col = bg_map.get(tc_type, "FFE2EFDA")
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(r, col_idx)
            if not c.fill or c.fill.fill_type in ('none', None) or \
               c.fill.fgColor.rgb in ('00000000', 'FFFFFFFF', 'FFD9D9D9'):
                c.fill = PF("solid", fgColor=bg_col)

        print(f"  [C] {tc_id} → Full data filled (Type={tc_type})")

print(f"\nTotal cells filled: {filled_count}")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(TMP)
print(f"Saved to {TMP}")

for attempt in range(20):
    try:
        os.replace(TMP, FILE)
        print(f"SUCCESS: {FILE} updated — all missing fields filled!")
        break
    except PermissionError:
        print(f"File locked ({attempt+1}/20) — close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"File still locked. Close Excel and rename {TMP} manually.")
