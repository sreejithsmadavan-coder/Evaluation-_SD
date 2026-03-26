import shutil, os, sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2.xlsx"
TMP  = r"C:\Users\wac.DESKTOP-UU75C42\Documents\Evaluation_SunnyDimonds\TestCase\SunnyDiamonds_v2_sd.xlsx"

GREEN  = "E2EFDA"
RED    = "FCE4D6"
YELLOW = "FFF2CC"

thin   = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    c.border    = border
    return c

def bg_color(tc_type):
    return GREEN if tc_type == "Positive" else (RED if tc_type == "Negative" else YELLOW)

NAV = "Navigate back to https://qa-sunnydiamonds.webc.in/checkout"

# ─────────────────────────────────────────────────────────────────────────────
# Business Rule (observed in video):
#   Cart total < ₹49,000  → COD + Pay Online BOTH available
#   Cart total >= ₹49,000 → ONLY Pay Online (COD option hidden)
#
# Video evidence:
#   18K Rose Gold Eden Diamond Ring ₹24,430 (x1) → Both COD & Pay Online shown
#   Same ring ₹24,430 (x3 = ₹73,290) → Only Pay Online shown, no COD
#
# Test product to use:
#   Below threshold → 18K Rose Gold Eden Diamond Ring (₹24,430 per unit)
#   Above threshold → Same ring qty 3 (₹73,290) OR any product > ₹49,000
# ─────────────────────────────────────────────────────────────────────────────

SD_TCS = [
    # ── A. Cart < ₹49,000 — COD Available ────────────────────────────────────
    (
        "SD_CHECKOUT_001",
        "Checkout - Payment Method Availability",
        "Cart total below ₹49,000 — verify both 'Cash on Delivery' and 'Pay Online' payment options are displayed",
        "User is logged in; cart contains 18K Rose Gold Eden Diamond Ring (₹24,430) — total below ₹49,000 threshold",
        "1. Login to https://qa-sunnydiamonds.webc.in\n"
        "2. Navigate to /jewellery PLP\n"
        "3. Add '18K Rose Gold Eden Diamond Ring' (₹24,430) to cart (qty=1, total < ₹49,000)\n"
        "4. Go to Cart → click 'CHECKOUT SECURELY'\n"
        "5. Scroll down to Payment Method section\n"
        "6. Observe all payment options displayed",
        "Both payment options are visible:\n"
        "• 'Cash on Delivery' radio button — enabled and selectable\n"
        "• 'Pay Online (Secure payment via Razorpay)' radio button — enabled\n"
        "COD option is NOT hidden or disabled when cart total < ₹49,000",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (12s–18s): Cart total ₹24,430 → checkout Payment Method shows BOTH COD and Pay Online options. Business rule: COD available for orders below ₹49,000."
    ),
    (
        "SD_CHECKOUT_002",
        "Checkout - Payment Method Availability",
        "Cart total below ₹49,000 — select 'Cash on Delivery' — verify button label shows 'PLACE ORDER'",
        "User is logged in; cart total < ₹49,000 (e.g., 18K Rose Gold Eden Diamond Ring ₹24,430)",
        "1. " + NAV + " (cart total < ₹49,000)\n"
        "2. Scroll to Payment Method section\n"
        "3. Click 'Cash on Delivery' radio button\n"
        "4. Observe submit button label",
        "Cash on Delivery radio button is selectable; submit button label changes to 'PLACE ORDER'; Pay Online radio becomes deselected",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (18s): Cart ₹24,430 → COD selected → button shows 'PLACE ORDER'. COD is fully functional when total < ₹49,000."
    ),
    (
        "SD_CHECKOUT_003",
        "Checkout - Payment Method Availability",
        "Cart total below ₹49,000 — select 'Pay Online' — verify button label shows 'PAY NOW'",
        "User is logged in; cart total < ₹49,000",
        "1. " + NAV + " (cart total < ₹49,000)\n"
        "2. Scroll to Payment Method section\n"
        "3. Click 'Pay Online' radio button\n"
        "4. Observe submit button label",
        "Pay Online radio is selectable; submit button label shows 'PAY NOW'; Razorpay payment gateway is triggered on click",
        "", "Not Tested", "Positive", "High",
        "Observed in video (12s): Cart ₹24,430 → Pay Online selected → button shows 'PAY NOW'. Both options functional when below threshold."
    ),

    # ── B. Cart >= ₹49,000 — COD NOT Available ───────────────────────────────
    (
        "SD_CHECKOUT_004",
        "Checkout - Payment Method Availability",
        "Cart total above ₹49,000 — verify 'Cash on Delivery' option is NOT displayed on checkout",
        "User is logged in; cart contains items with total > ₹49,000 (e.g., 18K Rose Gold Eden Diamond Ring qty=3, total ₹73,290)",
        "1. Login to https://qa-sunnydiamonds.webc.in\n"
        "2. Add product(s) to cart with combined total exceeding ₹49,000 (e.g., same ring x3 = ₹73,290)\n"
        "3. Go to Cart → click 'CHECKOUT SECURELY'\n"
        "4. Scroll down to Payment Method section\n"
        "5. Observe available payment options",
        "Only 'Pay Online (Secure payment via Razorpay)' is displayed in Payment Method section; 'Cash on Delivery' radio button is completely absent/hidden; submit button shows 'PAY NOW' only",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (29s–32s): Cart total ₹73,290 → checkout Payment Method shows ONLY 'Pay Online' — NO Cash on Delivery option. Business rule: COD disabled for orders >= ₹49,000."
    ),
    (
        "SD_CHECKOUT_005",
        "Checkout - Payment Method Availability",
        "Cart total above ₹49,000 — verify submit button shows 'PAY NOW' (not 'PLACE ORDER')",
        "User is logged in; cart total > ₹49,000",
        "1. " + NAV + " (cart total > ₹49,000)\n"
        "2. Scroll to Payment Method section\n"
        "3. Observe the submit button label without selecting any payment method",
        "Submit button displays 'PAY NOW'; 'PLACE ORDER' label does not appear at any point; only Pay Online is pre-selected or available",
        "", "Not Tested", "Positive", "High",
        "Observed in video (32s): Cart ₹73,290 → button shows 'PAY NOW' only. When COD is hidden, PLACE ORDER label should never appear."
    ),
    (
        "SD_CHECKOUT_006",
        "Checkout - Payment Method Availability",
        "Cart total above ₹49,000 — verify COD option is hidden (not just disabled) from Payment Method section",
        "User is logged in; cart total > ₹49,000",
        "1. " + NAV + " (cart total > ₹49,000)\n"
        "2. Scroll to Payment Method section\n"
        "3. Inspect DOM / page source for 'Cash on Delivery' element\n"
        "4. Verify COD is not present in UI or is hidden with display:none / visibility:hidden",
        "Cash on Delivery option is not visible in the Payment Method section; if present in DOM it must be hidden (display:none or similar); user cannot interact with COD option",
        "", "Not Tested", "Negative", "High",
        "UI verification — COD must not be accessible via DOM manipulation when cart > ₹49,000. Prevents bypassing payment restriction."
    ),

    # ── C. BVA — ₹49,000 Threshold Boundary Tests ───────────────────────────
    (
        "SD_CHECKOUT_007",
        "Checkout - Payment Method BVA",
        "BVA: Cart total = ₹48,999 (just below ₹49,000 threshold) — verify COD IS available",
        "User is logged in; cart items totalling exactly ₹48,999",
        "1. Login and add products to cart such that total is ₹48,999\n"
        "2. Go to Cart → click 'CHECKOUT SECURELY'\n"
        "3. Scroll to Payment Method section\n"
        "4. Observe payment options",
        "Both 'Cash on Delivery' and 'Pay Online' options are displayed; COD is selectable; cart total ₹48,999 is treated as below threshold",
        "", "Not Tested", "Edge Case", "Critical",
        "BVA: threshold-1. ₹48,999 must still show COD. Confirms the boundary is exclusive at ₹49,000 (< not <=)."
    ),
    (
        "SD_CHECKOUT_008",
        "Checkout - Payment Method BVA",
        "BVA: Cart total = ₹49,000 (exactly at threshold) — verify COD is NOT available",
        "User is logged in; cart items totalling exactly ₹49,000",
        "1. Login and add products to cart such that total is exactly ₹49,000\n"
        "2. Go to Cart → click 'CHECKOUT SECURELY'\n"
        "3. Scroll to Payment Method section\n"
        "4. Observe payment options",
        "Only 'Pay Online' is displayed at exactly ₹49,000; Cash on Delivery is hidden; confirms threshold is >= ₹49,000 (inclusive boundary)",
        "", "Not Tested", "Edge Case", "Critical",
        "BVA: exact threshold. ₹49,000 should disable COD. Critical boundary test — confirms whether rule is > or >= 49,000."
    ),
    (
        "SD_CHECKOUT_009",
        "Checkout - Payment Method BVA",
        "BVA: Cart total = ₹49,001 (just above ₹49,000 threshold) — verify COD is NOT available",
        "User is logged in; cart items totalling ₹49,001",
        "1. Login and add products to cart such that total is ₹49,001\n"
        "2. Go to Cart → click 'CHECKOUT SECURELY'\n"
        "3. Scroll to Payment Method section\n"
        "4. Observe payment options",
        "Only 'Pay Online' is displayed; Cash on Delivery is hidden; confirms any amount above threshold restricts COD",
        "", "Not Tested", "Edge Case", "Critical",
        "BVA: threshold+1. ₹49,001 must hide COD. Completes the boundary validation triplet (₹48,999 / ₹49,000 / ₹49,001)."
    ),

    # ── D. Dynamic Behaviour — Cart Modification Crosses Threshold ────────────
    (
        "SD_CHECKOUT_010",
        "Checkout - Payment Method Dynamic",
        "Increase cart quantity from below to above ₹49,000 threshold — verify COD disappears on checkout",
        "User is logged in; cart starts with 1 item (₹24,430) — below threshold",
        "1. Login; add 18K Rose Gold Eden Diamond Ring (₹24,430 x1) to cart\n"
        "2. Navigate to Checkout → verify COD is visible (total < ₹49,000)\n"
        "3. Click 'Edit Cart' → increase quantity to 3 (total = ₹73,290 > ₹49,000)\n"
        "4. Observe 'Item Added to Cart' toast notification\n"
        "5. Click 'CHECKOUT SECURELY' → navigate back to Checkout\n"
        "6. Scroll to Payment Method section",
        "After increasing qty to push total above ₹49,000: COD option is no longer displayed in Payment Method; only 'Pay Online' visible; button shows 'PAY NOW'; Order Summary reflects new total (₹73,290)",
        "", "Not Tested", "Positive", "Critical",
        "Observed in video (3s–32s): User increases ring qty to 3 (total ₹73,290), returns to checkout — COD gone, only Pay Online shown. Core business rule dynamic test."
    ),
    (
        "SD_CHECKOUT_011",
        "Checkout - Payment Method Dynamic",
        "Decrease cart quantity from above to below ₹49,000 threshold — verify COD reappears on checkout",
        "User is logged in; cart starts with 3 items (₹73,290) — above threshold",
        "1. Login; add 18K Rose Gold Eden Diamond Ring (₹24,430 x3 = ₹73,290) to cart\n"
        "2. Navigate to Checkout → verify only Pay Online is shown (total > ₹49,000)\n"
        "3. Click 'Edit Cart' → decrease quantity to 1 (total = ₹24,430 < ₹49,000)\n"
        "4. Click 'CHECKOUT SECURELY' → navigate back to Checkout\n"
        "5. Scroll to Payment Method section",
        "After reducing total below ₹49,000: Both 'Cash on Delivery' and 'Pay Online' options reappear; COD is selectable; Order Summary reflects reduced total (₹24,430)",
        "", "Not Tested", "Positive", "High",
        "Reverse of SD_CHECKOUT_010 — verify COD reappears when cart drops below threshold. Payment options must be recalculated on each checkout page load."
    ),
    (
        "SD_CHECKOUT_012",
        "Checkout - Payment Method Dynamic",
        "Add a high-value product (single item > ₹49,000) to cart — verify COD not available on checkout",
        "User is logged in; product with price above ₹49,000 (e.g., 18K Rose Gold Mix Diamond product ₹49,206)",
        "1. Login to https://qa-sunnydiamonds.webc.in\n"
        "2. Navigate to PLP → add a single product priced above ₹49,000 (e.g., ₹49,206) to cart\n"
        "3. Go to Cart → total is above ₹49,000 with just 1 item\n"
        "4. Click 'CHECKOUT SECURELY'\n"
        "5. Scroll to Payment Method section",
        "Even with a single item priced > ₹49,000: only 'Pay Online' is shown; COD option is hidden; confirms the rule applies to cart total regardless of item count",
        "", "Not Tested", "Positive", "High",
        "Observed on PLP (frame_000): 18K Rose Gold Mix Diamond ₹49,206 (above threshold). Validates rule applies per total — not per item count."
    ),
]

# ── Load workbook ──────────────────────────────────────────────────────────────
shutil.copy2(FILE, TMP)
wb = openpyxl.load_workbook(TMP)
ws = wb["Checkout Page"]
print(f"Loaded 'Checkout Page' — current rows: {ws.max_row}")

# ── Find TC_CHECKOUT_077 (Valid Login — must stay LAST) ───────────────────────
valid_login_row  = None
valid_login_data = None
for row in ws.iter_rows(min_row=6):
    if row[0].value and str(row[0].value).strip() == "TC_CHECKOUT_077":
        valid_login_row  = row[0].row
        valid_login_data = [c.value for c in row]
        break

if valid_login_row:
    print(f"Valid Login TC (TC_CHECKOUT_077) at row {valid_login_row} — will move to last")
    start_row = valid_login_row
else:
    start_row = ws.max_row + 1
    print(f"TC_CHECKOUT_077 not found — appending from row {start_row}")

# ── Write SD TCs ───────────────────────────────────────────────────────────────
for i, tc in enumerate(SD_TCS):
    r = start_row + i
    tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks = tc
    for col, val in enumerate([tc_id, mod, desc, pre, steps, exp, act, status, tc_type, priority, remarks], 1):
        write_cell(ws, r, col, val, bg_color(tc_type))
    print(f"  Written: {tc_id} at row {r}")

# ── Re-append Valid Login as absolute last ────────────────────────────────────
final_row = start_row + len(SD_TCS)
if valid_login_data:
    for col, val in enumerate(valid_login_data, 1):
        write_cell(ws, final_row, col, val, bg_color("Positive"))
    print(f"\nValid Login TC re-appended at row {final_row} as TC_CHECKOUT_077 (LAST)")
else:
    vl = [
        "TC_CHECKOUT_077",
        "Checkout - Authentication",
        "Valid Login — Verify authenticated user (sreejith.s+4@webandcrafts.com) can access Checkout",
        "",
        "1. Navigate to https://qa-sunnydiamonds.webc.in/login\n2. Login with sreejith.s+4@webandcrafts.com / Password\n3. Add product to cart → navigate to Checkout\n4. Verify checkout page loads fully",
        "User authenticated; Checkout page loads with all sections correctly",
        "", "Not Tested", "Positive", "Critical",
        "Valid login TC — ALWAYS LAST. Credentials: sreejith.s+4@webandcrafts.com / Password"
    ]
    for col, val in enumerate(vl, 1):
        write_cell(ws, final_row, col, val, bg_color("Positive"))
    print(f"\nValid Login TC written at row {final_row} (LAST)")

# ── Update metadata ────────────────────────────────────────────────────────────
total_tcs = final_row - 5
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        if cell.value and "Total TCs:" in str(cell.value):
            cell.value = re.sub(r'Total TCs:\s*\d+', f'Total TCs: {total_tcs}', cell.value)
            print(f"Metadata: Total TCs = {total_tcs}")

# ── Save & Replace ─────────────────────────────────────────────────────────────
wb.save(TMP)
print(f"\nSaved to {TMP}")

import time
for attempt in range(15):
    try:
        os.replace(TMP, FILE)
        print(f"SUCCESS: SunnyDiamonds_v2.xlsx updated — {total_tcs} total Checkout TCs")
        print(f"SD TCs: SD_CHECKOUT_001 to SD_CHECKOUT_012 at rows {start_row}–{start_row+len(SD_TCS)-1}")
        print(f"Last TC: TC_CHECKOUT_077 (Valid Login) at row {final_row}")
        break
    except PermissionError:
        print(f"File locked ({attempt+1}/15) — please close SunnyDiamonds_v2.xlsx in Excel...")
        time.sleep(4)
else:
    print(f"Could not replace. Saved as: {TMP}")
    print("Rename SunnyDiamonds_v2_sd.xlsx → SunnyDiamonds_v2.xlsx manually.")
